import discord
import openpyxl
import random
import asyncio
import datetime
import pytz
import requests
from bs4 import BeautifulSoup
from pyfiglet import Figlet
import json
from time import sleep
import timew
import timen
import platform
import psutil
import cpuinfo
import laftel
from animation import *
from urllib.request import urlopen
#from discord_buttons_plugin import *
#from discord.ext import commands
#from discord_slash import SlashCommand

client = discord.Client()




@client.event
async def on_ready():
    now = datetime.datetime.now()
    print("디스코드 봇 로그인이 완료되었습니다.")
    print("디스코드봇 이름:" + client.user.name)
    print("디스코드봇 ID:" + str(client.user.id))
    print("로그인 시간:"+str(datetime.datetime.today()))
    print("디스코드봇 버전:" + str(discord.__version__))
    s = len(client.guilds)
    print(len(client.guilds))
    print(client.guilds)
    print(client.intents)
    print('------')
    await bt(["!도움말 | !help", "!ヘルプ | !帮助", "With {} server | {}서버와 함께".format(s,s)])
    #await bt(["업데이트"])

async def bt(games):
    await client.wait_until_ready()
    while not client.is_closed():
        for g in games:
            await client.change_presence(status=discord.Status.online, activity=discord.Game(g)) # 온라인
            #await client.change_presence(status=discord.Status.dnd, activity=discord.Game(g)) # 방해금지
            #await client.change_presence(status=discord.Status.idle, activity=discord.Game(g)) # 자리비움
            await asyncio.sleep(3)


@client.event
async def on_message(message):



# 개발자용
    if message.author.bot:
        return None
    msg = message.content #msg == "" or message.content == ""

    
    if msg == "?bot on" or msg == "?bot on " or msg == "?boton" or msg == "?boton ":
        mi = 789670002163974145
        if message.author.id == mi:
            ch = client.get_channel(858237994481483777)
            await ch.send("<:MBOT:858655878311313459> M BOT이 온라인이 되었어요!")

    if msg == "?server" or msg == "?server " or msg == "?서버수" or msg == "?서버" or msg == "?서버수 ":
        mi = 789670002163974145
        if message.author.id == mi:
            ch = client.get_channel(859034894251327508)
            s = len(client.guilds)
            await ch.send("{}서버".format(s))

    if message.content.startswith("?패치"):
        mi = 789670002163974145
        if message.author.id == mi:
            ch = client.get_channel(858238395037646848)
            a = message.content[4:]
            embed = discord.Embed(title="[업데이트&패치]", description="패치내용 : {}".format(a),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
            embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
            embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
            embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
            await ch.send(embed=embed)

        
# 핑

    if msg == '엠봇아 핑' or msg == '엠봇아 ping' or msg == "!핑" or msg == "!ping" or msg == "m bot ping" or msg == "mbot ping" or msg == "M bot ping" or msg == "Mbot ping":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try: 
            p = round(client.latency*1000)
            start = timen.getnow()  
            msg = await message.channel.send("Ping test...")
            end = timen.getnow()  
            await msg.delete()
            api_ping = end - start 
            api_ping = round(api_ping * 1000)
            #api_ping = client.ws.ping
            if p < 100:
                embed = discord.Embed(title='''[mbot's ping]''', description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x48ff2a)
                embed.add_field(name="Ping",value="{}ms".format(p),inline=True)
                embed.add_field(name="API",value="{}ms".format(api_ping),inline=True)
                embed.add_field(name="State",value="very good 🟢",inline=True)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)

            elif p < 200:
                embed = discord.Embed(title='''[mbot's ping]''', description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x48ff2a)
                embed.add_field(name="Ping",value="{}ms".format(p),inline=True)
                embed.add_field(name="API",value="{}ms".format(api_ping),inline=True)
                embed.add_field(name="State",value="good 🟢",inline=True)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed) 

            elif p >= 200 and p < 400:
                embed = discord.Embed(title='''[mbot's ping]''', description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xffe82a)
                embed.add_field(name="Ping",value="{}ms".format(p),inline=True)
                embed.add_field(name="API",value="{}ms".format(api_ping),inline=True)
                embed.add_field(name="State",value="usually 🟡",inline=True)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)
            
            else:
                embed = discord.Embed(title='''[mbot's ping]''', description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff2a2a)
                embed.add_field(name="핑",value="{}ms".format(p),inline=True)
                embed.add_field(name="API",value="{}ms".format(api_ping),inline=True)
                embed.add_field(name="State",value="bad 🔴",inline=True)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return





# 기본-------------------------------------------------


    if msg == "엠봇아 안녕" or msg == "엠봇아 ㅎㅇ":
        await message.channel.send("안녕하세요!:wave:")
        m = message.author.dm
        if m == True:
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "엠봇아 잘가" or msg == "엠봇아 ㅂㅂ" or msg == "엠봇아 ㅂㅇ":
        await message.channel.send("잘가세요!")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "엠봇아" or msg == "엠봇" or msg == "mbot":
        a = random.randrange(1,6) 
        if a == 1:
            await message.channel.send("네!")
        elif a == 2:
            await message.channel.send("넹!")
        elif a == 3:
            await message.channel.send("!")
        elif a == 4:
            await message.channel.send("네?")
        elif a == 5:
            await message.channel.send("네")

        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "엠봇아 제작자" or msg == "엠봇 제작자" or msg == "M BOT 제작자":
        await message.channel.send("`마인잡지#0001`님이에요!")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "엠봇아 놀자":
        await message.channel.send("`!도움말`로 확인해 보세요!")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "엠봇아 마인잡지" or msg == "엠봇아 마잡" or msg == "엠봇아 마인잡체" or msg == "엠봇아 잡체" or msg == "엠봇아 잡지":
        await message.channel.send("엠봇 제작자!")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "알고리즘" or msg == "유튜브 알고리즘" or msg == "엠봇아 유튜브 알고리즘":
        await message.channel.send("🤖 삐빅 나는야 엠파고 ~~번역기~~너튜브 알고리즘이지 삐빅 🤖")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "엠봇아 시간" or msg == "엠봇 시간" or msg == "mbot time" or msg == "!time" or msg == "!시간":
        now = datetime.datetime.now()
        embedtime = discord.Embed(title=':alarm_clock: 현재 시간 :alarm_clock:', description=str(now.year)+"년 "+str(now.month)+"월 "+str(now.day)+"일 "+str(now.hour)+"시 "+str(now.minute)+"분 "+str(now.second)+"초", color=0x83fff8)
        await message.channel.send(embed=embedtime)
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "!hellothisisverification":
        await message.channel.send("마인잡지#0001 / ID : 789670002163974145")
        print("개발자 {} {}".format(message.author.id,message.author.name))
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if message.content == "!프사":
        embed = discord.Embed(title="{}님의 프사입니다!".format(message.author), description="", color=0xbf6b4e)
        embed.set_image(url=message.author.avatar_url)
        await message.channel.send(embed=embed)
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if message.content.startswith("!시간"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

        now = datetime.datetime.now()

        nh = now.hour
        nm = now.minute
        ns = now.second
        embed = discord.Embed(title = "대한민국 서울특별시 기준 시각", description="{}시{}분{}초 입니다".format(nh,nm,ns), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x579aff)
        await message.channel.send(embed = embed)

    if message.content.startswith("!주사위"):
        r = random.randrange(1,7)
        embed = discord.Embed(title = "주사위", description="{}".format(r), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x579aff)
        await message.channel.send(embed = embed)

    if message.content.startswith("안녕") or message.content.startswith("안뇽") or message.content.startswith("안농"):
        await message.channel.send(message.author.mention+"님 "+message.guild.name+"서버에 오신것을 환영합니다!")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id,message.content,message.channel,message.channel.id,message.guild.name,message.guild.id,datetime.datetime.today()))

    if msg == "?" or msg == "??" or msg == "???" or msg == "????" or msg == "?????" or msg == "??????" or msg == "????????" or msg == "????????" or msg == "?????????" or msg == "??????????" or msg == "???????????" or msg == "????????????" or msg == "?????????????":
        await message.channel.send("?")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
            message.author.name, message.author.id, message.content, message.channel, message.channel.id,
            message.guild.name, message.guild.id, datetime.datetime.today()))

    if message.content.startswith("ㅋ"):
        r = random.randrange(2,6)
        if r == 2:
            await message.channel.send("ㅋㅋ")
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
                message.author.name, message.author.id, message.content, message.channel, message.channel.id,
                message.guild.name, message.guild.id, datetime.datetime.today()))

        elif r == 3:
            await message.channel.send("ㅋㅋㅋ")
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
                message.author.name, message.author.id, message.content, message.channel, message.channel.id,
                message.guild.name, message.guild.id, datetime.datetime.today()))

        elif r == 4:
            await message.channel.send("ㅋㅋㅋㅋ")
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
                message.author.name, message.author.id, message.content, message.channel, message.channel.id,
                message.guild.name, message.guild.id, datetime.datetime.today()))

        else:
            await message.channel.send("ㅋㅋㅋㅋㅋ")
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
                message.author.name, message.author.id, message.content, message.channel, message.channel.id,
                message.guild.name, message.guild.id, datetime.datetime.today()))

    if message.content.startswith("ㄷ"):
        await message.channel.send("ㄷㄷ")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
            message.author.name, message.author.id, message.content, message.channel, message.channel.id,
            message.guild.name, message.guild.id, datetime.datetime.today()))

    if message.content.startswith("ㅠ"):
        await message.channel.send("ㅠㅠ")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(
            message.author.name, message.author.id, message.content, message.channel, message.channel.id,
            message.guild.name, message.guild.id, datetime.datetime.today()))

# 도움말 ----------------------------------------------
    if msg == "!도움말" or msg == "!도움말 " or msg == "!도움" or msg == "!도움 " or msg == '엠봇아 도움말' or msg == "엠봇아 도움말 " or msg == "엠봇아 도움" or msg == "엠봇아 도움 " or msg == "!help" or msg == "!ヘルプ" or msg == "!帮助":
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):

                if sheet["Y" + str(i)].value == 0:
                    t = sheet["B" + str(i)].value
                    embed = discord.Embed(title="도움말", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️기본**", value="`!ping`, `!언어`, `!언어 <언어이름>`, `!초대`, `!서포터`, `!홈페이지`, `!출석`, `엠봇아 <할말>`, `!기억 <이름> <기억할것>`,\n`!기억삭제 <이름>`, `엠봇아 <이름>`", inline=False)
                    embed.add_field(name="**🙍‍♀️프로필**", value="`!가입`, `!프로필`, `!상메 <내용>`, `!업적`", inline=False)
                    embed.add_field(name="**💰돈**", value="`!돈`, `!돈받기`, `!은행`, `!계좌개설`, `!입금 <돈>`,`!출금 <돈>`,`!계좌`,`!구걸`, `!채광`", inline=False)
                    embed.add_field(name="**🎮게임**", value="`!가위바위보`, `!뽑기`, `!도박 <돈>`, `!올인`, `!주사위`", inline=False)
                    embed.add_field(name="**🎮전적**", value="`!롤 <이름>`", inline=False)
                    embed.add_field(name="**🏪상점**", value="`!상점`, `!전자상점`, `!음식상점`", inline=False)
                    embed.add_field(name="**💻컴퓨터**", value="`!컴퓨터`, `!부팅`, `!화면`, `!os설치`", inline=False)
                    embed.add_field(name="**🔍검색**", value="`!코로나`, `!뉴스`, `!웹툰`", inline=False)
                    embed.add_field(name="**🛠관리자**", value="`!공지 <내용>`, `!투표 <내용1>/<내용2>`, `!삭제 <삭제할 메세지 수>`, `!슬로우 <초>`", inline=False)
                    embed.add_field(name="**🎈기타**", value="`!봇 정보`, `!내 정보`, `!한강`, `!타이머 <초>`, `!출력 <메세지(영어로)>`,\n`!계산 <수> <연산자> <수>`, `!원주율`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠주의사항⚠**", value="1. M BOT의 모든 기능을 사용하시려면`!가입`을 해야됩니다.\n2. 관리자 명령어는 서버에서 관리자 역할이 있어야만 실행됩니다.", inline=False)
                    embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 도움말", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                elif sheet["Y" + str(i)].value == 1:
                    t = sheet["B" + str(i)].value
                    embed = discord.Embed(title="Help", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️Basic**", value="`!ping`, `!language`, `!language <language name>`, `!invite`, `!supporter`, `!website`, `!attendance`, `mbot <Key>`, `!memory <Key> <Value>`,\n`!erasememory <이름>`, `mbot <say>`", inline=False)
                    embed.add_field(name="**🙍‍♀️Profile**", value="`!join`, `!profile`, `!statusmessage <Contents>`, `!Achievements`", inline=False)
                    embed.add_field(name="**💰Money**", value="`!money`, `!get money`, `!bank`, `!accountopening`, `!deposit <money>`,`!withdraw <money>`,`!account`,`!beg`, `!mining`", inline=False)
                    embed.add_field(name="**🎮Game**", value="`!Rock Paper Scissors`, `!draw`, `!gambling <money>`, `!all`", inline=False)
                    embed.add_field(name="**🎮 Retrieve player information**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪Shop**", value="`!shop`, `!e-shop`, `!foodshop`", inline=False)
                    embed.add_field(name="**💻Computer**", value="`!computer`, `!boot`, `!screen`, `!os install`", inline=False)
                    embed.add_field(name="**🔍Search**", value="`!covid-19`, `!news`, `!webtoon`", inline=False)
                    embed.add_field(name="**🛠Administrator**", value="`!notice <Content>`, `!vote <Content 1>/<Content 2>`, `!delete <number>`, `!slow <second>`", inline=False)
                    embed.add_field(name="**🎈Guitar**", value="`!botinfo`, `!userinfo`, `!hanriver`, `!timer <second>`, `!print <message>`,\n`!calculate <number> <+ - * / ^> <number>`, `!pi`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠Caution⚠**", value="1. To use all the features of M BOT, you must sign up for '!join`\n2. Administrator commands are executed only if you have an administrator role on the server.", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT Help", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                elif sheet["Y" + str(i)].value == 2:
                    t = sheet["B" + str(i)].value
                    embed = discord.Embed(title="ヘルプ", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️基本**", value="`!ping`, `!言語`, `!言語 <言語名>`, `!招待`, `!サポーター`, `!ウェブサイト`, `!出席`, `エムボトア <Key>`, `!記憶 <Key> <Value>`,\n`!記憶の削除 <Key>`, `エムボトア <発言>`", inline=False)
                    embed.add_field(name="**🙍‍♀️プロフィール**", value="`!登録`, `!プロフィール`, `!状態メッセージ <内容>`, `!業績`", inline=False)
                    embed.add_field(name="**💰お金**", value="`!お金`, `!お金を受け取る`, `!銀行`, `!口座開設`, `!入金 <お金>`,`!出金 <お金>`,`!口座`,`!頼む`, `!採鉱`", inline=False)
                    embed.add_field(name="**🎮ゲーム**", value="`!じゃんけん`, `!抜く`, `!ギャンブル <お金>`, `!オールイン`", inline=False)
                    embed.add_field(name="**🎮プレーヤー情報検索**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪店**", value="`!店`, `!電子商店`, `!食料品店`", inline=False)
                    embed.add_field(name="**💻コンピュータ**", value="`!コンピュータ`, `!起動`, `!画面`, `!osのインストール`", inline=False)
                    embed.add_field(name="**🔍検索**", value="`!コロナ19`, `!ニュース`, `!ウェプトゥン`", inline=False)
                    embed.add_field(name="**🛠管理者**", value="`!お知らせ <内容>`, `!投票 <内容1>/<内容2>`, `!削除 <数字>`, `!スロー <秒>`", inline=False)
                    embed.add_field(name="**🎈その他**", value="`!ボットについて`, `!私の情報`, `!漢江`, `!タイマー <秒>`, `!出力 <message>`,\n`!計算 <数> <+ - * / ^> <数>`, `!円周率`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠注意事項⚠**", value="1. M BOTのすべての機能を使用するには、`！登録`をしなければならなります。\n2.管理者のコマンドは、サーバーの管理者の役割があってこそ実行されます。", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOTヘルプ", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                elif sheet["Y" + str(i)].value == 3:
                    t = sheet["B" + str(i)].value
                    embed = discord.Embed(title="帮助", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️基础**", value="`!ping`, `!语`, `!语 <语言名称>`, `!邀请`, `!支持者`, `!网站`, `!出勤率`, `机器人 <Key>`, `!记忆 <Key> <Value>`,\n`!擦除记忆 <Key>`, `机器人 <说>`", inline=False)
                    embed.add_field(name="**🙍‍♀️简介**", value="`!加入`, `!轮廓`, `!状态消息 <内容>`, `!成就`", inline=False)
                    embed.add_field(name="**💰钱**", value="`!钱`, `!拿钱`, `!银行`, `!开户`, `!订金 <钱>`,`!提取 <钱>`,`!帐户`,`!乞讨`, `!矿业`", inline=False)
                    embed.add_field(name="**🎮游戏**", value="`!剪刀石头布`, `!画`, `!赌博 <钱>`, `!全在`", inline=False)
                    embed.add_field(name="**🎮检索玩家信息**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪店铺**", value="`!店铺`, `!网上商店`, `!食品店`", inline=False)
                    embed.add_field(name="**💻电脑**", value="`!计算机`, `!开机`, `!屏幕`, `!操作系统安装`", inline=False)
                    embed.add_field(name="**🔍搜索**", value="`!新冠肺炎`, `!消息`, `!网络卡通`", inline=False)
                    embed.add_field(name="**🛠管理员**", value="`!注意 <内容>`, `!投票 <内容1>/<内容2>`, `!删除 <数字>`, `!慢 <秒>`", inline=False)
                    embed.add_field(name="**🎈吉他**", value="`!机器人信息`, `!我的信息`, `!汉江`, `!计时器 <秒>`, `!打印 <message>`,\n`!计算 <数字> <+ - * / ^> <数字>`, `!周长`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠警告**", value="1.要使用M BOT的所有功能，您需要注册'!加入'。\n2. 只有当您在服务器上具有管理员角色时，才会执行管理员命令。", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 帮助", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
        
        try:
            msg = await message.channel.send(embed=embed)
            await msg.add_reaction("🇰🇷")
            await msg.add_reaction("🇺🇸")
            await msg.add_reaction("🇯🇵")
            await msg.add_reaction("🇨🇳")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '🇰🇷' or user == message.author and str(reaction.emoji) == '🇺🇸' or user == message.author and str(reaction.emoji) == '🇯🇵' or user == message.author and str(reaction.emoji) == '🇨🇳' or user == message.author and str(reaction.emoji) == '❌'


            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                await msg.delete()
                await message.channel.send(embed=embed)

            else:
                if str(reaction.emoji) == "❌":
                    await msg.delete()
                    await message.channel.send(embed=embed)


                elif str(reaction.emoji) == "🇰🇷":
                    embed = discord.Embed(title="도움말", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️기본**", value="`!ping`, `!언어`, `!언어 <언어이름>`, `!초대`, `!서포터`, `!홈페이지`, `!출석`, `엠봇아 <할말>`, `!기억 <이름> <기억할것>`,\n`!기억삭제 <이름>`, `엠봇아 <이름>`", inline=False)
                    embed.add_field(name="**🙍‍♀️프로필**", value="`!가입`, `!프로필`, `!상메 <내용>`, `!업적`", inline=False)
                    embed.add_field(name="**💰돈**", value="`!돈`, `!돈받기`, `!은행`, `!계좌개설`, `!입금 <돈>`,`!출금 <돈>`,`!계좌`,`!구걸`, `!채광`", inline=False)
                    embed.add_field(name="**🎮게임**", value="`!가위바위보`, `!뽑기`, `!도박 <돈>`, `!올인`, `!주사위`", inline=False)
                    embed.add_field(name="**🎮전적**", value="`!롤 <이름>`", inline=False)
                    embed.add_field(name="**🏪상점**", value="`!상점`, `!전자상점`, `!음식상점`", inline=False)
                    embed.add_field(name="**💻컴퓨터**", value="`!컴퓨터`, `!부팅`, `!화면`, `!os설치`", inline=False)
                    embed.add_field(name="**🔍검색**", value="`!코로나`, `!뉴스`, `!웹툰`", inline=False)
                    embed.add_field(name="**🛠관리자**", value="`!공지 <내용>`, `!투표 <내용1>/<내용2>`, `!삭제 <삭제할 메세지 수>`, `!슬로우 <초>`", inline=False)
                    embed.add_field(name="**🎈기타**", value="`!봇 정보`, `!내 정보`, `!한강`, `!타이머 <초>`, `!출력 <메세지(영어로)>`,\n`!계산 <수> <연산자> <수>`, `!원주율`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠주의사항⚠**", value="1. M BOT의 모든 기능을 사용하시려면`!가입`을 해야됩니다.\n2. 관리자 명령어는 서버에서 관리자 역할이 있어야만 실행됩니다.", inline=False)
                    embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 도움말", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)


                elif str(reaction.emoji) == "🇺🇸":
                    embed = discord.Embed(title="Help", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️Basic**", value="`!ping`, `!language`, `!language <language name>`, `!invite`, `!supporter`, `!website`, `!attendance`, `mbot <Key>`, `!memory <Key> <Value>`,\n`!erasememory <이름>`, `mbot <say>`", inline=False)
                    embed.add_field(name="**🙍‍♀️Profile**", value="`!join`, `!profile`, `!statusmessage <Contents>`, `!Achievements`", inline=False)
                    embed.add_field(name="**💰Money**", value="`!money`, `!get money`, `!bank`, `!accountopening`, `!deposit <money>`,`!withdraw <money>`,`!account`,`!beg`, `!mining`", inline=False)
                    embed.add_field(name="**🎮Game**", value="`!Rock Paper Scissors`, `!draw`, `!gambling <money>`, `!all`", inline=False)
                    embed.add_field(name="**🎮 Retrieve player information**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪Shop**", value="`!shop`, `!e-shop`, `!foodshop`", inline=False)
                    embed.add_field(name="**💻Computer**", value="`!computer`, `!boot`, `!screen`, `!os install`", inline=False)
                    embed.add_field(name="**🔍Search**", value="`!covid-19`, `!news`, `!webtoon`", inline=False)
                    embed.add_field(name="**🛠Administrator**", value="`!notice <Content>`, `!vote <Content 1>/<Content 2>`, `!delete <number>`, `!slow <second>`", inline=False)
                    embed.add_field(name="**🎈Guitar**", value="`!botinfo`, `!userinfo`, `!hanriver`, `!timer <second>`, `!print <message>`,\n`!calculate <number> <+ - * / ^> <number>`, `!pi`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠Caution⚠**", value="1. To use all the features of M BOT, you must sign up for '!join`\n2. Administrator commands are executed only if you have an administrator role on the server.", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT Help", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)

                elif str(reaction.emoji) == "🇯🇵":
                    embed = discord.Embed(title="ヘルプ", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️基本**", value="`!ping`, `!言語`, `!言語 <言語名>`, `!招待`, `!サポーター`, `!ウェブサイト`, `!出席`, `エムボトア <Key>`, `!記憶 <Key> <Value>`,\n`!記憶の削除 <Key>`, `エムボトア <発言>`", inline=False)
                    embed.add_field(name="**🙍‍♀️プロフィール**", value="`!登録`, `!プロフィール`, `!状態メッセージ <内容>`, `!業績`", inline=False)
                    embed.add_field(name="**💰お金**", value="`!お金`, `!お金を受け取る`, `!銀行`, `!口座開設`, `!入金 <お金>`,`!出金 <お金>`,`!口座`,`!頼む`, `!採鉱`", inline=False)
                    embed.add_field(name="**🎮ゲーム**", value="`!じゃんけん`, `!抜く`, `!ギャンブル <お金>`, `!オールイン`", inline=False)
                    embed.add_field(name="**🎮プレーヤー情報検索**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪店**", value="`!店`, `!電子商店`, `!食料品店`", inline=False)
                    embed.add_field(name="**💻コンピュータ**", value="`!コンピュータ`, `!起動`, `!画面`, `!osのインストール`", inline=False)
                    embed.add_field(name="**🔍検索**", value="`!コロナ19`, `!ニュース`, `!ウェプトゥン`", inline=False)
                    embed.add_field(name="**🛠管理者**", value="`!お知らせ <内容>`, `!投票 <内容1>/<内容2>`, `!削除 <数字>`, `!スロー <秒>`", inline=False)
                    embed.add_field(name="**🎈その他**", value="`!ボットについて`, `!私の情報`, `!漢江`, `!タイマー <秒>`, `!出力 <message>`,\n`!計算 <数> <+ - * / ^> <数>`, `!円周率`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠注意事項⚠**", value="1. M BOTのすべての機能を使用するには、`！登録`をしなければならなります。\n2.管理者のコマンドは、サーバーの管理者の役割があってこそ実行されます。", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOTヘルプ", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)

                elif str(reaction.emoji) == "🇨🇳":
                    await msg.delete()
                    embed = discord.Embed(title="帮助", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️基础**", value="`!ping`, `!语`, `!语 <语言名称>`, `!邀请`, `!支持者`, `!网站`, `!出勤率`, `机器人 <Key>`, `!记忆 <Key> <Value>`,\n`!擦除记忆 <Key>`, `机器人 <说>`", inline=False)
                    embed.add_field(name="**🙍‍♀️简介**", value="`!加入`, `!轮廓`, `!状态消息 <内容>`, `!成就`", inline=False)
                    embed.add_field(name="**💰钱**", value="`!钱`, `!拿钱`, `!银行`, `!开户`, `!订金 <钱>`,`!提取 <钱>`,`!帐户`,`!乞讨`, `!矿业`", inline=False)
                    embed.add_field(name="**🎮游戏**", value="`!剪刀石头布`, `!画`, `!赌博 <钱>`, `!全在`", inline=False)
                    embed.add_field(name="**🎮检索玩家信息**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪店铺**", value="`!店铺`, `!网上商店`, `!食品店`", inline=False)
                    embed.add_field(name="**💻电脑**", value="`!计算机`, `!开机`, `!屏幕`, `!操作系统安装`", inline=False)
                    embed.add_field(name="**🔍搜索**", value="`!新冠肺炎`, `!消息`, `!网络卡通`", inline=False)
                    embed.add_field(name="**🛠管理员**", value="`!注意 <内容>`, `!投票 <内容1>/<内容2>`, `!删除 <数字>`, `!慢 <秒>`", inline=False)
                    embed.add_field(name="**🎈吉他**", value="`!机器人信息`, `!我的信息`, `!汉江`, `!计时器 <秒>`, `!打印 <message>`,\n`!计算 <数字> <+ - * / ^> <数字>`, `!周长`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠警告**", value="1.要使用M BOT的所有功能，您需要注册'!加入'。\n2. 只有当您在服务器上具有管理员角色时，才会执行管理员命令。", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 帮助", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)

            print(t) # 가입 확인용 엑셀 조회 

        except:
            embed = discord.Embed(title="도움말", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
            embed.add_field(name="**⚙️기본**", value="`!ping`, `!언어`, `!언어 <언어이름>`, `!초대`, `!서포터`, `!홈페이지`, `!출석`, `엠봇아 <할말>`, `!기억 <이름> <기억할것>`,\n`!기억삭제 <이름>`, `엠봇아 <이름>`", inline=False)
            embed.add_field(name="**🙍‍♀️프로필**", value="`!가입`, `!프로필`, `!상메 <내용>`, `!업적`", inline=False)
            embed.add_field(name="**💰돈**", value="`!돈`, `!돈받기`, `!은행`, `!계좌개설`, `!입금 <돈>`,`!출금 <돈>`,`!계좌`,`!구걸`, `!채광`", inline=False)
            embed.add_field(name="**🎮게임**", value="`!가위바위보`, `!뽑기`, `!도박 <돈>`, `!올인`, `!주사위`", inline=False)
            embed.add_field(name="**🎮전적**", value="`!롤 <이름>`", inline=False)
            embed.add_field(name="**🏪상점**", value="`!상점`, `!전자상점`, `!음식상점`", inline=False)
            embed.add_field(name="**💻컴퓨터**", value="`!컴퓨터`, `!부팅`, `!화면`, `!os설치`", inline=False)
            embed.add_field(name="**🔍검색**", value="`!코로나`, `!뉴스`, `!웹툰`", inline=False)
            embed.add_field(name="**🛠관리자**", value="`!공지 <내용>`, `!투표 <내용1>/<내용2>`, `!삭제 <삭제할 메세지 수>`, `!슬로우 <초>`", inline=False)
            embed.add_field(name="**🎈기타**", value="`!봇 정보`, `!내 정보`, `!한강`, `!타이머 <초>`, `!출력 <메세지(영어로)>`,\n`!계산 <수> <연산자> <수>`, `!원주율`", inline=False)
            embed.add_field(name="** **", value="** **", inline=False)
            embed.add_field(name="**⚠주의사항⚠**", value="1. M BOT의 모든 기능을 사용하시려면`!가입`을 해야됩니다.\n2. 관리자 명령어는 서버에서 관리자 역할이 있어야만 실행됩니다.", inline=False)
            embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
            embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
            embed.set_author(name="M BOT 도움말", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
            msg = await message.channel.send(embed=embed)
            await msg.add_reaction("🇰🇷")
            await msg.add_reaction("🇺🇸")
            await msg.add_reaction("🇯🇵")
            await msg.add_reaction("🇨🇳")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '🇰🇷' or user == message.author and str(reaction.emoji) == '🇺🇸' or user == message.author and str(reaction.emoji) == '🇯🇵' or user == message.author and str(reaction.emoji) == '🇨🇳' or user == message.author and str(reaction.emoji) == '❌'


            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                await msg.delete()
                await message.channel.send(embed=embed)

            else:
                if str(reaction.emoji) == "❌":
                    await msg.delete()
                    await message.channel.send(embed=embed)


                elif str(reaction.emoji) == "🇰🇷":
                    embed = discord.Embed(title="도움말", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️기본**", value="`!ping`, `!언어`, `!언어 <언어이름>`, `!초대`, `!서포터`, `!홈페이지`, `!출석`, `엠봇아 <할말>`, `!기억 <이름> <기억할것>`,\n`!기억삭제 <이름>`, `엠봇아 <이름>`", inline=False)
                    embed.add_field(name="**🙍‍♀️프로필**", value="`!가입`, `!프로필`, `!상메 <내용>`, `!업적`", inline=False)
                    embed.add_field(name="**💰돈**", value="`!돈`, `!돈받기`, `!은행`, `!계좌개설`, `!입금 <돈>`,`!출금 <돈>`,`!계좌`,`!구걸`, `!채광`", inline=False)
                    embed.add_field(name="**🎮게임**", value="`!가위바위보`, `!뽑기`, `!도박 <돈>`, `!올인`, `!주사위`", inline=False)
                    embed.add_field(name="**🎮전적**", value="`!롤 <이름>`", inline=False)
                    embed.add_field(name="**🏪상점**", value="`!상점`, `!전자상점`, `!음식상점`", inline=False)
                    embed.add_field(name="**💻컴퓨터**", value="`!컴퓨터`, `!부팅`, `!화면`, `!os설치`", inline=False)
                    embed.add_field(name="**🔍검색**", value="`!코로나`, `!뉴스`, `!웹툰`", inline=False)
                    embed.add_field(name="**🛠관리자**", value="`!공지 <내용>`, `!투표 <내용1>/<내용2>`, `!삭제 <삭제할 메세지 수>`, `!슬로우 <초>`", inline=False)
                    embed.add_field(name="**🎈기타**", value="`!봇 정보`, `!내 정보`, `!한강`, `!타이머 <초>`, `!출력 <메세지(영어로)>`,\n`!계산 <수> <연산자> <수>`, `!원주율`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠주의사항⚠**", value="1. M BOT의 모든 기능을 사용하시려면`!가입`을 해야됩니다.\n2. 관리자 명령어는 서버에서 관리자 역할이 있어야만 실행됩니다.", inline=False)
                    embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 도움말", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)

                elif str(reaction.emoji) == "🇺🇸":
                    embed = discord.Embed(title="Help", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️Basic**", value="`!ping`, `!language`, `!language <language name>`, `!invite`, `!supporter`, `!website`, `!attendance`, `mbot <Key>`, `!memory <Key> <Value>`,\n`!erasememory <이름>`, `mbot <say>`", inline=False)
                    embed.add_field(name="**🙍‍♀️Profile**", value="`!join`, `!profile`, `!statusmessage <Contents>`, `!Achievements`", inline=False)
                    embed.add_field(name="**💰Money**", value="`!money`, `!get money`, `!bank`, `!accountopening`, `!deposit <money>`,`!withdraw <money>`,`!account`,`!beg`, `!mining`", inline=False)
                    embed.add_field(name="**🎮Game**", value="`!Rock Paper Scissors`, `!draw`, `!gambling <money>`, `!all`", inline=False)
                    embed.add_field(name="**🎮 Retrieve player information**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪Shop**", value="`!shop`, `!e-shop`, `!foodshop`", inline=False)
                    embed.add_field(name="**💻Computer**", value="`!computer`, `!boot`, `!screen`, `!os install`", inline=False)
                    embed.add_field(name="**🔍Search**", value="`!covid-19`, `!news`, `!webtoon`", inline=False)
                    embed.add_field(name="**🛠Administrator**", value="`!notice <Content>`, `!vote <Content 1>/<Content 2>`, `!delete <number>`, `!slow <second>`", inline=False)
                    embed.add_field(name="**🎈Guitar**", value="`!botinfo`, `!userinfo`, `!hanriver`, `!timer <second>`, `!print <message>`,\n`!calculate <number> <+ - * / ^> <number>`, `!pi`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠Caution⚠**", value="1. To use all the features of M BOT, you must sign up for '!join`\n2. Administrator commands are executed only if you have an administrator role on the server.", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT Help", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)

                elif str(reaction.emoji) == "🇯🇵":
                    embed = discord.Embed(title="ヘルプ", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️基本**", value="`!ping`, `!言語`, `!言語 <言語名>`, `!招待`, `!サポーター`, `!ウェブサイト`, `!出席`, `エムボトア <Key>`, `!記憶 <Key> <Value>`,\n`!記憶の削除 <Key>`, `エムボトア <発言>`", inline=False)
                    embed.add_field(name="**🙍‍♀️プロフィール**", value="`!登録`, `!プロフィール`, `!状態メッセージ <内容>`, `!業績`", inline=False)
                    embed.add_field(name="**💰お金**", value="`!お金`, `!お金を受け取る`, `!銀行`, `!口座開設`, `!入金 <お金>`,`!出金 <お金>`,`!口座`,`!頼む`, `!採鉱`", inline=False)
                    embed.add_field(name="**🎮ゲーム**", value="`!じゃんけん`, `!抜く`, `!ギャンブル <お金>`, `!オールイン`", inline=False)
                    embed.add_field(name="**🎮プレーヤー情報検索**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪店**", value="`!店`, `!電子商店`, `!食料品店`", inline=False)
                    embed.add_field(name="**💻コンピュータ**", value="`!コンピュータ`, `!起動`, `!画面`, `!osのインストール`", inline=False)
                    embed.add_field(name="**🔍検索**", value="`!コロナ19`, `!ニュース`, `!ウェプトゥン`", inline=False)
                    embed.add_field(name="**🛠管理者**", value="`!お知らせ <内容>`, `!投票 <内容1>/<内容2>`, `!削除 <数字>`, `!スロー <秒>`", inline=False)
                    embed.add_field(name="**🎈その他**", value="`!ボットについて`, `!私の情報`, `!漢江`, `!タイマー <秒>`, `!出力 <message>`,\n`!計算 <数> <+ - * / ^> <数>`, `!円周率`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠注意事項⚠**", value="1. M BOTのすべての機能を使用するには、`！登録`をしなければならなります。\n2.管理者のコマンドは、サーバーの管理者の役割があってこそ実行されます。", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOTヘルプ", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)

                elif str(reaction.emoji) == "🇨🇳":
                    await msg.delete()
                    embed = discord.Embed(title="帮助", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                    embed.add_field(name="**⚙️基础**", value="`!ping`, `!语`, `!语 <语言名称>`, `!邀请`, `!支持者`, `!网站`, `!出勤率`, `机器人 <Key>`, `!记忆 <Key> <Value>`,\n`!擦除记忆 <Key>`, `机器人 <说>`", inline=False)
                    embed.add_field(name="**🙍‍♀️简介**", value="`!加入`, `!轮廓`, `!状态消息 <内容>`, `!成就`", inline=False)
                    embed.add_field(name="**💰钱**", value="`!钱`, `!拿钱`, `!银行`, `!开户`, `!订金 <钱>`,`!提取 <钱>`,`!帐户`,`!乞讨`, `!矿业`", inline=False)
                    embed.add_field(name="**🎮游戏**", value="`!剪刀石头布`, `!画`, `!赌博 <钱>`, `!全在`", inline=False)
                    embed.add_field(name="**🎮检索玩家信息**", value="`!lol <Nickname>`", inline=False)
                    embed.add_field(name="**🏪店铺**", value="`!店铺`, `!网上商店`, `!食品店`", inline=False)
                    embed.add_field(name="**💻电脑**", value="`!计算机`, `!开机`, `!屏幕`, `!操作系统安装`", inline=False)
                    embed.add_field(name="**🔍搜索**", value="`!新冠肺炎`, `!消息`, `!网络卡通`", inline=False)
                    embed.add_field(name="**🛠管理员**", value="`!注意 <内容>`, `!投票 <内容1>/<内容2>`, `!删除 <数字>`, `!慢 <秒>`", inline=False)
                    embed.add_field(name="**🎈吉他**", value="`!机器人信息`, `!我的信息`, `!汉江`, `!计时器 <秒>`, `!打印 <message>`,\n`!计算 <数字> <+ - * / ^> <数字>`, `!周长`", inline=False)
                    embed.add_field(name="** **", value="** **", inline=False)
                    embed.add_field(name="**⚠警告**", value="1.要使用M BOT的所有功能，您需要注册'!加入'。\n2. 只有当您在服务器上具有管理员角色时，才会执行管理员命令。", inline=False)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 帮助", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await msg.delete()
                    await message.channel.send(embed=embed)




                

    #if message.content.startswith("!도움")

# 출력
    if message.content.startswith("!출력"):
        try:
            g = message.content[4:]
            f = Figlet(font='slant')
            await message.channel.send("```" + f.renderText(g) + "```") 
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return

    if message.content.startswith("!print"):
        try:
            g = message.content[7:]
            f = Figlet(font='slant')
            await message.channel.send("```" + f.renderText(g) + "```") 
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return

    if message.content.startswith("!出力"):
        try:
            g = message.content[4:]
            f = Figlet(font='slant')
            await message.channel.send("```" + f.renderText(g) + "```") 
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return
    
    if message.content.startswith("!打印"):
        try:
            g = message.content[4:]
            f = Figlet(font='slant')
            await message.channel.send("```" + f.renderText(g) + "```") 
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


# 코로나-------------------------------------------------------------
    if msg == "엠봇아 코로나" or msg == "엠봇아 코로나19" or msg == "엠봇아 covid" or msg == "엠봇아 covid19":

        req = requests.get("http://ncov.mohw.go.kr/")

        soup = BeautifulSoup(req.text, "html.parser")

        k, w = soup.find("div", class_="liveNum_today_new").find_all("span", class_="data")
        t = soup.find("div", class_="liveNumOuter").find_all("span",class_="livedate")
        all = soup.find("div", class_="liveNum").find_all("span",class_="num")
        wan = soup.find("div", class_="liveNum").find_all("span",class_="before") #격리해제(완치)
        wan = str(wan)
        wany = wan[27:34] #전일대비 증가
        wann = wan[64:71] #완치
        wanc = wan[101:108] #치료
        wank = wan[138:143] #전일대비 사망
        t = str(t)
        t = t[25:37] #날짜
        all = str(all)
        alln = all[49:56] #누적
        allw = all[83:90] #완치
        allc = all[117:122] #치료중
        alld = all[149:154] #사망


        k = k.text #국내
        w = w.text #해외
        k = k.replace(",","")
        k = int(k)
        w = int(w)
        a = k+w
        us = message.author.avatar_url
        wann = wann[3:6]
        embed = discord.Embed(title="대한민국 코로나19 확진 현황\n{}".format(t), description="🇰🇷 **국내발생** : {}\n✈ **해외유입** : {}\n🚑 **합계** : {}\n🎈 **완치** : {}\n\n🚑 **누적** : {}{}\n🎆 **누적완치** : {}{}\n🏥 **치료중** : {}{}\n■ **사망** : {}{}".format(k,w,a,wann,alln,wany,allw,wann,allc,wanc,alld,wank), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.add_field(name="ㅤ", value="[오류제보](<https://discord.gg/upt73pxKHX>)", inline=True)
        embed.add_field(name="ㅤ", value="[출처](<http://ncov.mohw.go.kr/>)", inline=True)
        embed.set_image(url="https://media.discordapp.net/attachments/832849028873191516/837500708915904562/covid19.jpg?width=902&height=638")
        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/837504706896330782/covid.png?width=638&height=638")
        embed.set_footer(text="{}".format(message.author),icon_url=us)

        await message.channel.send(embed=embed)
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

    if msg == "!코로나 출처" or msg == "!covid source" or msg == "!コロナ源" or msg == "!新冠肺炎源":
        await message.channel.send("http://ncov.mohw.go.kr/")
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

    if msg == "!코로나" or msg == "!코로나19" or msg == "!covid" or msg == "!covid19" or msg == "!covid-19" or msg == "!コロナ19" or msg == "!コロナ" or msg == "!新冠肺炎":

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id

        req = requests.get("http://ncov.mohw.go.kr/")

        soup = BeautifulSoup(req.text, "html.parser")

        k, w = soup.find("div", class_="liveNum_today_new").find_all("span", class_="data")
        t = soup.find("div", class_="liveNumOuter").find_all("span",class_="livedate")
        all = soup.find("div", class_="liveNum").find_all("span",class_="num")
        wan = soup.find("div", class_="liveNum").find_all("span",class_="before") #격리해제(완치)
        
        wany = wan[0]
        wany = str(wany)
        wany = wany[26:-7]

        wann = wan[1]
        wann = str(wann)
        wann = wann[21:-7]

        wanc = wan[2]
        wanc = str(wanc)
        wanc = wanc[21:-7]

        wank = wan[3]
        wank = str(wank)
        wank = wank[21:-7]

        t = str(t)
        t = t[25:37] #날짜

        alln = all[0] 
        alln = str(alln)
        alln = alln[48:-7]

        allw = all[1]
        allw = str(allw)
        allw = allw[18:-7]


        allc = all[2]
        allc = str(allc)
        allc = allc[18:-7]


        alld = all[3]
        alld = str(alld)
        alld = alld[18:-7]


        k = k.text #국내
        w = w.text #해외
        k = k.replace(",","")
        k = int(k)
        w = int(w)
        a = k+w #합계
        us = message.author.avatar_url
        wanna = wann[3:6]


        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    embed = discord.Embed(title="대한민국 코로나19 확진 현황\n{}".format(t), description="🇰🇷 **국내발생** : {}\n✈ **해외유입** : {}\n🚑 **합계** : {}\n🎈 **완치** : {}\n\n🚑 **누적** : {}{}\n🎆 **누적완치** : {}{}\n🏥 **치료중** : {}{}\n■ **사망** : {}{}".format(k,w,a,wanna,alln,wany,allw,wann,allc,wanc,alld,wank), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                    embed.add_field(name="ㅤ", value="[오류제보](<https://discord.gg/fEEHHM4eHq>)", inline=True)
                    embed.add_field(name="ㅤ", value="[출처](<http://ncov.mohw.go.kr/>)", inline=True)
                    embed.set_image(url="https://media.discordapp.net/attachments/832849028873191516/837500708915904562/covid19.jpg?width=902&height=638")
                    embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/837504706896330782/covid.png?width=638&height=638")
                    embed.set_footer(text="{}".format(message.author),icon_url=us)

                elif sheet["Y" + str(i)].value == 1:
                    embed = discord.Embed(title="Korea's Covid-19 Confirmation Status\n{}".format(t), description="🇰🇷 **Domestic Occurrence** : {}\n✈ **Foreign import** : {}\n🚑 **Total** : {}\n🎈 **Cure** : {}\n\n🚑 **Accumulate** : {}{}\n🎆 **Cumulative cure** : {}{}\n🏥 **Under treatment** : {}{}\n■ **Dead** : {}{}".format(k,w,a,wanna,alln,wany,allw,wann,allc,wanc,alld,wank), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                    embed.add_field(name="ㅤ", value="[Error Report](<https://discord.gg/fEEHHM4eHq>)", inline=True)
                    embed.add_field(name="ㅤ", value="[Source](<http://ncov.mohw.go.kr/>)", inline=True)
                    embed.set_image(url="https://media.discordapp.net/attachments/832849028873191516/837500708915904562/covid19.jpg?width=902&height=638")
                    embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/837504706896330782/covid.png?width=638&height=638")
                    embed.set_footer(text="{}".format(message.author),icon_url=us)

                elif sheet["Y" + str(i)].value == 2:
                    embed = discord.Embed(title="大韓民国コロナ19確定現況\n{}".format(t), description="🇰🇷 **国内発生** : {}\n✈ **海外流入** : {}\n🚑 **合計** : {}\n🎈 **完治** : {}\n\n🚑 **累積** : {}{}\n🎆 **累積完治** : {}{}\n🏥 **治療中** : {}{}\n■ **死亡** : {}{}".format(k,w,a,wanna,alln,wany,allw,wann,allc,wanc,alld,wank), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                    embed.add_field(name="ㅤ", value="[エラー情報提供](<https://discord.gg/fEEHHM4eHq>)", inline=True)
                    embed.add_field(name="ㅤ", value="[ソース](<http://ncov.mohw.go.kr/>)", inline=True)
                    embed.set_image(url="https://media.discordapp.net/attachments/832849028873191516/837500708915904562/covid19.jpg?width=902&height=638")
                    embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/837504706896330782/covid.png?width=638&height=638")
                    embed.set_footer(text="{}".format(message.author),icon_url=us)

                elif sheet["Y" + str(i)].value == 3:
                    embed = discord.Embed(title="韩国的 新冠肺炎 确认状态\n{}".format(t), description="🇰🇷 **国内发生** : {}\n✈ **国外进口** : {}\n🚑 **和** : {}\n🎈 **治愈** : {}\n\n🚑 **积累** : {}{}\n🎆 **累积治愈** : {}{}\n🏥 **治疗中** : {}{}\n■ **死的** : {}{}".format(k,w,a,wanna,alln,wany,allw,wann,allc,wanc,alld,wank), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                    embed.add_field(name="ㅤ", value="[错误报告](<https://discord.gg/fEEHHM4eHq>)", inline=True)
                    embed.add_field(name="ㅤ", value="[来源](<http://ncov.mohw.go.kr/>)", inline=True)
                    embed.set_image(url="https://media.discordapp.net/attachments/832849028873191516/837500708915904562/covid19.jpg?width=902&height=638")
                    embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/837504706896330782/covid.png?width=638&height=638")
                    embed.set_footer(text="{}".format(message.author),icon_url=us)

        try:
            await message.channel.send(embed=embed)
        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)



        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))


# 뉴스
    if message.content.startswith("!뉴스") or msg == "!news" or msg == "!News" or msg == "!NEWS" or msg == "!ニュース" or msg == "!消息":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    wnecom = "어느 언론사로 하시겠습니까?"
                    errep = "오류제보"
                    remos = "자세히 보기"
                    sour = "제공"
                    look = "보러가기"
                    discd = "Discord"
                    
                elif sheet["Y" + str(i)].value == 1:
                    wnecom = "Which press would you like to work for?"
                    errep = "Error Report"
                    remos = "Read more"
                    sour = "Offer"
                    look = "Go to see"
                    discd = "Discord"

                elif sheet["Y" + str(i)].value == 2:
                    wnecom = "どの報道機関にしますか？"
                    errep = "エラー情報提供"
                    remos = "続きを読む"
                    sour = "提供"
                    look = "見に行く"
                    discd = "Discord"
                    
                elif sheet["Y" + str(i)].value == 3:
                    wnecom = "你想为哪家出版社工作？"
                    errep = "错误报告"
                    remos = "阅读更多"
                    sour = "提供"
                    look = "去看"
                    discd = "Discord"
                    
        try:
            embed = discord.Embed(title="NEWS", description="{}".format(wnecom), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)
            embed.add_field(name="YTN", value="<:YTN_News_logo:840799636569980949>", inline=True)
            #embed.add_field(name="MBC", value="<:MBC_News_logo:840800374083944449>", inline=True)
            embed.add_field(name="KBS", value="<:KBS_News_Logo:840800750292434944>", inline=True)

            embed.add_field(name="ㅤ", value="[{}](<https://discord.gg/upt73pxKHX>)".format(errep), inline=True)
            embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
            msg = await message.channel.send(embed=embed)
            await msg.add_reaction("<:YTN_News_logo:840799636569980949>")
            #await msg.add_reaction("<:MBC_News_logo:840800374083944449>")
            await msg.add_reaction("<:KBS_News_Logo:840800750292434944>")
            await msg.add_reaction("❌")
        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)




        def check(reaction, user):
            if user.bot == 1:
                return None
            return user == message.author and str(
                reaction.emoji) == '<:YTN_News_logo:840799636569980949>' or user == message.author and str(reaction.emoji) == '<:MBC_News_logo:840800374083944449>' or user == message.author and str(reaction.emoji) == '<:KBS_News_Logo:840800750292434944>' or user == message.author and str(reaction.emoji) == '❌'
        try:
            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
        except asyncio.TimeoutError:
            await msg.delete()
        else:
            if str(reaction.emoji) == "<:YTN_News_logo:840799636569980949>":
                await msg.delete()
                embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                msg = await message.channel.send(embed=embed1)
                html = requests.get("https://www.ytn.co.kr/")
                soup = BeautifulSoup(html.text, "html.parser")
                html.close()
                datal_list = soup.findAll('div', {'class': 'top_li2'})
                #datal_list = soup.findAll('li', {'class': 'YTN_CSA_topnews1'})
                li_list = []
                for datal in datal_list:
                    li_list.extend(datal.findAll('li'))
                li_list = li_list[0:1]
                for li in li_list:
                    img = li.find('img')
                    title = img['alt']
                    img_src = img['src']
                    l1 = li.find('a')
                    l1 = l1['href']
                    embed = discord.Embed(title="**[YTN NEWS]**", description=title,timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x179de9)
                    embed.add_field(name=remos, value=f"[{look}](<{l1}>)", inline=True)
                    embed.add_field(name=errep, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                    embed.add_field(name=sour, value="[YTN](<https://www.ytn.co.kr/>)",inline=True)
                    embed.set_thumbnail(url="https://www.ytn.co.kr/img/info/new_img_company01.gif")
                    embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                    embed.set_image(url=img_src)
                    await msg.edit(embed=embed)





            if str(reaction.emoji) == "<:KBS_News_Logo:840800750292434944>":
                await msg.delete()
                embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                msg = await message.channel.send(embed=embed1)
                html = requests.get("https://news.kbs.co.kr/common/main.html")
                soup = BeautifulSoup(html.text, "html.parser")
                html.close()
                #print(soup)
                datal_list = soup.findAll('div', {'class': "section-box"})
                datal_list = soup.findAll('ul', {'class': "list-type list-thumb"})
                li_list = []
                for datal in datal_list:
                    li_list.extend(datal.findAll('li'))
                    li_list = li_list[1:2]
                for li in li_list:
                    img = li.find('img')
                    title = img['alt']
                    img_src = img['src']
                    l1 = li.find('a')
                    l1 = l1['href']
                    title = title[8:]
                    l1 = "https://news.kbs.co.kr"+l1
                    embed = discord.Embed(title="**[KBS NEWS]**", description=title,timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x179de9)
                    embed.add_field(name=remos, value=f"[{look}](<{l1}>)", inline=True)
                    embed.add_field(name=errep, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                    embed.add_field(name=sour, value="[YTN](<https://www.ytn.co.kr/>)",inline=True)
                    embed.set_thumbnail(url="https://blog.kakaocdn.net/dn/yPTjk/btqwNuhXB9M/diZOWtur8epy0zE0KcGRGK/img.jpg")
                    embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                    embed.set_image(url="https://news.kbs.co.kr"+img_src)
                    await msg.edit(embed=embed)




            if str(reaction.emoji) == "❌":
                await msg.delete()


# 웹툰-------------------------------
    if message.content.startswith("!요일웹툰") or message.content.startswith("!웹툰추천") or message.content.startswith("!네이버웹툰") or message.content.startswith("!네이버 웹툰") or message.content.startswith("!오늘의 웹툰") or message.content.startswith("!요일 웹툰") or message.content.startswith("!웹툰 추천") or message.content.startswith("!웹툰") or msg == "!网络卡通" or msg == "!ウェプトゥン" or msg == "!Web toon" or msg == "!Webtoon" or msg == "!webtoon":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        week = timew.getDay()  # 오늘 (timew 에 getday 함수 불러오기)

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    monweb = "월요일의 웹툰" # 월요일의 웹툰
                    tueweb = "화요일의 웹툰" # 화요일의 웹툰
                    wedweb = "수요일의 웹툰" # 수요일의 웹툰
                    thuweb = "목요일의 웹툰" # 목요일의 웹툰
                    friweb = "금요일의 웹툰" # 금요일의 웹툰
                    satweb = "토요일의 웹툰" # 토요일의 웹툰
                    sunweb = "일요일의 웹툰" # 일요일의 웹툰

                    st = "1위" # 1위
                    nd = "2위" # 2위
                    rd = "3위" # 3위
                    postv = "자세이 보기" # 자세이 보기

                    errorpo = "오류제보" # 오류제보
                    discd = "디스코드" # 디스코드

                    offer = "제공" # 제공
                    naver = "네이버웹툰" # 네이버웹툰

                    webpostv = "웹툰 자세히 보기" # 웹툰 자세히 보기
                    webpostv_1 = "자세히 보고싶은 웹툰을 이모지로 반응해 주세요!" # 자세히 보고싶은 웹툰을 이모지로 반응해 주세요!
                    webpostv_2 = "컷툰의 경우 작가명이\n제대로 나오지 않을수 있습니다." # 컷툰의 경우 작가명이\n제대로 나오지 않을수 있습니다.

                    # _ 자제히

                    wtitle = "제목" # 제목
                    wauthor = "작가" # 작가
                    wgenre = "장르" # 장르
                    wage = "연령" # 연령
                    wlook = "보러가기" #보러가기

                elif sheet["Y" + str(i)].value == 1:
                    monweb = "Monday's Webtoon" # 월요일의 웹툰
                    tueweb = "Tuesday's Webtoon" # 화요일의 웹툰
                    wedweb = "Wednesday's Webtoon" # 수요일의 웹툰
                    thuweb = "Thursday's Webtoon" # 목요일의 웹툰
                    friweb = "Friday's Webtoon" # 금요일의 웹툰
                    satweb = "Saturday's Webtoon" # 토요일의 웹툰
                    sunweb = "Sunday's Webtoon" # 일요일의 웹툰

                    st = "1st" # 1위
                    nd = "2nd place" # 2위
                    rd = "3rd place" # 3위
                    postv = "Posture view" # 자세이 보기

                    errorpo = "Error report" # 오류제보
                    discd = "Discord" # 디스코드

                    offer = "Offer" # 제공
                    naver = "Naver Webtoon" # 네이버웹툰

                    webpostv = "Learn more about webtoons" # 웹툰 자세히 보기
                    webpostv_1 = "Please respond to the webtoon you want to see in detail with an emoji!" # 자세히 보고싶은 웹툰을 이모지로 반응해 주세요!
                    webpostv_2 = "In the case of cuttoons, the name of the artist\nmay not appear properly." # 컷툰의 경우 작가명이\n제대로 나오지 않을수 있습니다.

                    # _ 자제히

                    wtitle = "Title" # 제목
                    wauthor = "Writer" # 작가
                    wgenre = "Genre" # 장르
                    wage = "Age" # 연령
                    wlook = "Go to see" # 보러가기

                elif sheet["Y" + str(i)].value == 2:
                    monweb = "月曜日のウェプトゥン" # 월요일의 웹툰
                    tueweb = "火曜日のウェプトゥン" # 화요일의 웹툰
                    wedweb = "水曜日のウェプトゥン" # 수요일의 웹툰
                    thuweb = "木曜日のウェプトゥン" # 목요일의 웹툰
                    friweb = "金曜日のウェプトゥン" # 금요일의 웹툰
                    satweb = "土曜日のウェプトゥン" # 토요일의 웹툰
                    sunweb = "日曜日のウェプトゥン" # 일요일의 웹툰

                    st = "1位" # 1위
                    nd = "2位" # 2위
                    rd = "3位" # 3위
                    postv = "ジャセイ表示" # 자세이 보기

                    errorpo = "エラー情報提供" # 오류제보
                    discd = "ディスコード" # 디스코드

                    offer = "提供" # 제공
                    naver = "ネイバーウェプトゥン" # 네이버웹툰

                    webpostv = "ウェプトゥン続きを読む" # 웹툰 자세히 보기
                    webpostv_1 = "詳細見たいウェプトゥンを叔母地反応してください！" # 자세히 보고싶은 웹툰을 이모지로 반응해 주세요!
                    webpostv_2 = "コトツンの場合作家名\n正しく出ざるを得ています。" # 컷툰의 경우 작가명이\n제대로 나오지 않을수 있습니다.

                    # _ 자제히

                    wtitle = "タイトル" # 제목
                    wauthor = "作家" # 작가
                    wgenre = "ジャンル" # 장르
                    wage = "年齢" # 연령
                    wlook = "見に行く" # 보러가기
                    
                elif sheet["Y" + str(i)].value == 3:
                    monweb = "星期一的网络漫画" # 월요일의 웹툰
                    tueweb = "周二的网络漫画" # 화요일의 웹툰
                    wedweb = "周三的网络漫画" # 수요일의 웹툰
                    thuweb = "星期四的网络漫画" # 목요일의 웹툰
                    friweb = "周五的网络漫画" # 금요일의 웹툰
                    satweb = "星期六的网络漫画" # 토요일의 웹툰
                    sunweb = "周日的网络漫画" # 일요일의 웹툰

                    st = "第一" # 1위
                    nd = "第二名" # 2위
                    rd = "第三名" # 3위
                    postv = "姿势观" # 자세이 보기

                    errorpo = "错误报告" # 오류제보
                    discd = "不和谐" # 디스코드

                    offer = "提供" # 제공
                    naver = "Naver 网络漫画" # 네이버웹툰

                    webpostv = "了解有关网络漫画的更多信息" # 웹툰 자세히 보기
                    webpostv_1 = "请用表情符号回复您想详细查看的网络漫画！" # 자세히 보고싶은 웹툰을 이모지로 반응해 주세요!
                    webpostv_2 = "对于卡通，艺术家姓名\n可能无法正确显示。" # 컷툰의 경우 작가명이\n제대로 나오지 않을수 있습니다.

                    # _ 자제히

                    wtitle = "标题" # 제목
                    wauthor = "作者" # 작가
                    wgenre = "类型" # 장르
                    wage = "年龄" # 연령
                    wlook = "去看" # 보러가기


        if week == "월":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=mon")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list = soup.findAll('div', {'class': 'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))

            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']

                        embed = discord.Embed(title="**{}**".format(monweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")

                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'

                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(monweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xaaaaaa)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")

                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(reaction.emoji) == '🥇' or user == message.author and str(reaction.emoji) == '🥈' or user == message.author and str(reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com" + ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com" + ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src2)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com" + ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:  # 내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:  # 장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:  # 작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title3,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title3, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src3)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll3), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()

                            if str(reaction.emoji) == "❌":
                                await msg.delete()
                                
        if week == "화":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=tue")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list = soup.findAll('div', {'class': 'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))

            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']

                        embed = discord.Embed(title="**{}**".format(tueweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")

                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(
                                reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'

                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(tueweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x30a0ff)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")

                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(reaction.emoji) == '🥇' or user == message.author and str(reaction.emoji) == '🥈' or user == message.author and str(reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com" + ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xe33434)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com" + ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xe33434)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com" + ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:  # 내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:  # 장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:  # 작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xe33434)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()

                            if str(reaction.emoji) == "❌":
                                await msg.delete()
                                
        if week == "수":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=wed")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list = soup.findAll('div', {'class': 'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))
            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']

                        embed = discord.Embed(title="**{}**".format(wedweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")

                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'

                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(wedweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x30a0ff)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")

                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(
                                        reaction.emoji) == '🥇' or user == message.author and str(
                                        reaction.emoji) == '🥈' or user == message.author and str(
                                        reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com" + ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x346de3)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com" + ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src2)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x346de3)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com" + ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:  # 내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:  # 장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:  # 작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title3,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title3, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src3)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll3), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x346de3)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()

                            if str(reaction.emoji) == "❌":
                                await msg.delete()
                                
        if week == "목":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=thu")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list = soup.findAll('div', {'class': 'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))

            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']

                        embed = discord.Embed(title="**{}**".format(thuweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")

                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(
                                reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'

                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(thuweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x30a0ff)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")

                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(reaction.emoji) == '🥇' or user == message.author and str(reaction.emoji) == '🥈' or user == message.author and str(reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com" + ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x30ce5c)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com" + ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src2)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x30ce5c)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com" + ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:  # 내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:  # 장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:  # 작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title3,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title3, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src3)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll3), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x30ce5c)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()

                            if str(reaction.emoji) == "❌":
                                await msg.delete()
        
        if week == "금":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=fri")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list = soup.findAll('div', {'class': 'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))

            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']

                        embed = discord.Embed(title="**{}**".format(friweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")

                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'

                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(friweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x30a0ff)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")

                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(reaction.emoji) == '🥇' or user == message.author and str(reaction.emoji) == '🥈' or user == message.author and str(reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com" + ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",
                                                                           timestamp=datetime.datetime.now(
                                                                               pytz.timezone('UTC')), color=0xf7be64)
                                                    embed1.set_image(
                                                        url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com" + ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src2)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xf7be64)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com" + ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:  # 내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:  # 장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:  # 작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()
                                                    embed = discord.Embed(title=title3,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title3, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src3)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll3), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",
                                                                           timestamp=datetime.datetime.now(
                                                                               pytz.timezone('UTC')), color=0xf7be64)
                                                    embed1.set_image(
                                                        url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()

                            if str(reaction.emoji) == "❌":
                                await msg.delete()

        if week == "토":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=sat")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list=soup.findAll('div',{'class':'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))

            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']


                        embed = discord.Embed(title="**{}**".format(satweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")
                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(
                                reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'
                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(satweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x30a0ff)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")



                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(reaction.emoji) == '🥇' or user == message.author and str(reaction.emoji) == '🥈' or user == message.author and str(reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com"+ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()


                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x996e24)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    embed1.set_footer(text="{}".format(message.author),
                                                                     icon_url=message.author.avatar_url)
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com"+ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()




                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src2)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x996e24)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    embed1.set_footer(text="{}".format(message.author),
                                                                     icon_url=message.author.avatar_url)
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com"+ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list: #내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list: #장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]


                                                uwj = uw_list1.find("span", class_="genre") #장르
                                                uwo = uw_list1.find("span", class_="age") #나이
                                                uwj = uwj.text
                                                uwo = uwo.text


                                                for u in u_list: #작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()


                                                    embed = discord.Embed(title=title3,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title3, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src3)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll3), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                                                    embed.set_footer(text="{}".format(message.author),icon_url=message.author.avatar_url)


                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x996e24)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    embed1.set_footer(text="{}".format(message.author),
                                                                     icon_url=message.author.avatar_url)

                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()
                                        
                            if str(reaction.emoji) == "❌":
                                await msg.delete()
                                
        if week == "일":
            html = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week=sun")
            soup = BeautifulSoup(html.text, "html.parser")
            html.close()
            datal_list = soup.findAll('div', {'class': 'list_area daily_img'})
            li_list = []
            for datal in datal_list:
                li_list.extend(datal.findAll('li'))

            li_list = li_list[0:3]
            li_list1 = li_list[0:1]
            li_list2 = li_list[1:2]
            li_list3 = li_list[2:3]
            for li1 in li_list1:
                img1 = li1.find('img')
                title1 = img1['title']
                img_src1 = img1['src']
                l1 = li1.find('a')
                ll1 = l1['href']
                for li2 in li_list2:
                    img2 = li2.find('img')
                    title2 = img2['title']
                    img_src2 = img2['src']
                    l2 = li2.find('a')
                    ll2 = l2['href']
                    for li3 in li_list3:
                        img3 = li3.find('img')
                        title3 = img3['title']
                        img_src3 = img3['src']
                        l3 = li3.find('a')
                        ll3 = l3['href']

                        embed = discord.Embed(title="**{}**".format(sunweb),description="🥇{} : {}\n🥈{} : {}\n🥉 {} : {}\n\n{} 🔍".format(st, title1, nd, title2, rd, title3, postv),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                        embed.set_thumbnail(url="https://media.discordapp.net/attachments/832849028873191516/840419651182919680/Naver_Line_Webtoon_logo.png")
                        embed.add_field(name=errorpo, value="[{}](<https://discord.gg/upt73pxKHX>)".format(discd), inline=True)
                        embed.add_field(name=offer, value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                        embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("🔍")
                        await msg.add_reaction("❌")

                        def check(reaction, user):
                            if user.bot == 1:
                                return None
                            return user == message.author and str(
                                reaction.emoji) == '🔍' or user == message.author and str(reaction.emoji) == '❌'

                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                        except asyncio.TimeoutError:
                            await msg.delete()
                        else:
                            if str(reaction.emoji) == "🔍":
                                await msg.delete()
                                embed = discord.Embed(title="**{}**".format(sunweb), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x30a0ff)
                                embed.add_field(name=webpostv,value="{}\n🥇 | {}\n🥈 | {}\n🥉 | {}\n\n**{}**".format(webpostv_1, title1,title2,title3, webpostv_2),inline=False)
                                embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                                msg = await message.channel.send(embed=embed)
                                await msg.add_reaction("🥇")
                                await msg.add_reaction("🥈")
                                await msg.add_reaction("🥉")
                                await msg.add_reaction("❌")

                                def check(reaction, user):
                                    if user.bot == 1:
                                        return None
                                    return user == message.author and str(reaction.emoji) == '🥇' or user == message.author and str(reaction.emoji) == '🥈' or user == message.author and str(reaction.emoji) == '🥉' or user == message.author and str(reaction.emoji) == '❌'

                                try:
                                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                                except asyncio.TimeoutError:
                                    await msg.delete()
                                else:
                                    if str(reaction.emoji) == "🥇":
                                        await msg.delete()
                                        lll1 = "https://comic.naver.com" + ll1

                                        html = requests.get(lll1)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                uw_list1 = uw_list1.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title1,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title1, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src1)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll1), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                                                    embed.set_footer(text="{}".format(message.author),icon_url=message.author.avatar_url)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x3b7dc1)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    embed1.set_footer(text="{}".format(message.author),icon_url=message.author.avatar_url)
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥈":
                                        await msg.delete()
                                        lll2 = "https://comic.naver.com" + ll2

                                        html = requests.get(lll2)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()

                                                    embed = discord.Embed(title=title2,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title2, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src2)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll2), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                                                    embed.set_footer(text="{}".format(message.author),icon_url=message.author.avatar_url)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x3b7dc1)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    embed1.set_footer(text="{}".format(message.author),
                                                                     icon_url=message.author.avatar_url)
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "🥉":
                                        await msg.delete()
                                        lll3 = "https://comic.naver.com" + ll3

                                        html = requests.get(lll3)
                                        soup = BeautifulSoup(html.text, "html.parser")
                                        html.close()
                                        datal_list = soup.findAll('div', {'class': 'detail'})
                                        u_list = []
                                        uw_list = []
                                        uw_list1 = []

                                        for datal in datal_list:
                                            u_list.extend(datal.findAll('h2'))

                                        for datal in datal_list:  # 내용
                                            uw_list.extend(datal.findAll('p'))
                                            uw_list = uw_list[0]
                                            uw_list = uw_list.text

                                            for datal in datal_list:  # 장르,연령
                                                uw_list1.extend(datal.findAll('p'))
                                                uw_list1 = uw_list1[1]

                                                uwj = uw_list1.find("span", class_="genre")  # 장르
                                                uwo = uw_list1.find("span", class_="age")  # 나이
                                                uwj = uwj.text
                                                uwo = uwo.text

                                                for u in u_list:  # 작가명
                                                    uu = u.find("span", class_="wrt_nm")
                                                    uu = uu.text
                                                    uu = uu.lstrip()
                                                    embed = discord.Embed(title=title3,description="{} : **{}**\n{} : **{}**\n\n**{}**\n\n{} : **{}**\n{} : **{}**".format(wtitle, title3, wauthor, uu, uw_list, wgenre, uwj, wage, uwo),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xaaaaaa)
                                                    embed.set_thumbnail(url=img_src3)
                                                    embed.add_field(name=wlook, value="[{}](<{}>)".format(wlook, lll3), inline=True)
                                                    embed.add_field(name=errorpo,value="[{}](<https://discord.gg/fEEHHM4eHq>)".format(discd),inline=True)
                                                    embed.add_field(name=offer,value="[{}](<https://comic.naver.com/webtoon/weekday.nhn>)".format(naver),inline=True)
                                                    embed.set_footer(text="{}".format(message.author),icon_url=message.author.avatar_url)

                                                    embed1 = discord.Embed(title="Loading..", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x3b7dc1)
                                                    embed1.set_image(url="https://media.discordapp.net/attachments/832849028873191516/840427068411215922/loading_img.gif")
                                                    embed1.set_footer(text="{}".format(message.author),icon_url=message.author.avatar_url)
                                                    msg = await message.channel.send(embed=embed1)
                                                    await msg.edit(embed=embed)

                                    elif str(reaction.emoji) == "❌":
                                        await msg.delete()

                            if str(reaction.emoji) == "❌":
                                await msg.delete()


# 프로필 -------------------------------------------------------
    if message.content.startswith("!상메"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                te = message.content[3:]
                sheet["D" + str(i)].value = str(te)
                if sheet["Y" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다\n내용 :" + te)
                    break

                elif sheet["Y" + str(i)].value == 1:
                    await message.channel.send(message.author.mention + "Registration is complete.\nContents:" + te)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    await message.channel.send(message.author.mention + "登録が完了しました.\n内容：" + te)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    await message.channel.send(message.author.mention + "注册完成.\n内容 :" + te)
                    break

                else:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다.\n 내용 :" + te)
                    break

        file.save("user.xlsx")

    if message.content.startswith("!상태메세지"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                te = message.content[3:]
                sheet["D" + str(i)].value = str(te)
                if sheet["Y" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다\n내용 :" + te)
                    break

                elif sheet["Y" + str(i)].value == 1:
                    await message.channel.send(message.author.mention + "Registration is complete.\nContents:" + te)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    await message.channel.send(message.author.mention + "登録が完了しました.\n内容：" + te)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    await message.channel.send(message.author.mention + "注册完成.\n内容 :" + te)
                    break

                else:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다.\n 내용 :" + te)
                    break

                
        file.save("user.xlsx")

    if message.content.startswith("!statusmessage"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                te = message.content[15:]
                sheet["D" + str(i)].value = str(te)
                if sheet["Y" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다\n내용 :" + te)
                    break

                elif sheet["Y" + str(i)].value == 1:
                    await message.channel.send(message.author.mention + "Registration is complete.\nContents:" + te)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    await message.channel.send(message.author.mention + "登録が完了しました.\n内容：" + te)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    await message.channel.send(message.author.mention + "注册完成.\n内容 :" + te)
                    break

                else:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다.\n 내용 :" + te)
                    break
        file.save("user.xlsx")

    if message.content.startswith("!状態メッセージ"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                te = message.content[9:]
                sheet["D" + str(i)].value = str(te)
                if sheet["Y" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다\n내용 :" + te)
                    break

                elif sheet["Y" + str(i)].value == 1:
                    await message.channel.send(message.author.mention + "Registration is complete.\nContents:" + te)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    await message.channel.send(message.author.mention + "登録が完了しました.\n内容：" + te)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    await message.channel.send(message.author.mention + "注册完成.\n内容 :" + te)
                    break

                else:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다.\n 내용 :" + te)
                    break
        file.save("user.xlsx")

    if message.content.startswith("!状态消息"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                te = message.content[6:]
                sheet["D" + str(i)].value = str(te)
                if sheet["Y" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다\n내용 :" + te)
                    break

                elif sheet["Y" + str(i)].value == 1:
                    await message.channel.send(message.author.mention + "Registration is complete.\nContents:" + te)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    await message.channel.send(message.author.mention + "登録が完了しました.\n内容：" + te)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    await message.channel.send(message.author.mention + "注册完成.\n内容 :" + te)
                    break

                else:
                    await message.channel.send(message.author.mention + "등록이 완료되었습니다.\n 내용 :" + te)
                    break
        file.save("user.xlsx")

    #_______________________________

    if message.content.startswith("!프로필 등록") or message.content.startswith("!프로필 가입") or message.content.startswith("!가입") or message.content.startswith("!join") or message.content.startswith("!登録") or message.content.startswith("!加入"):
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        us = message.author.avatar_url
        now = datetime.datetime.now()
        ny = now.year
        nm = now.month
        nd = now.day
        nh = now.hour
        nmm = now.minute
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        #엑셀시트 F = 년 / G = 월 / H = 일

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                await message.channel.send(message.author.mention + "님 이미 가입되었습니다")
                break
            else:
                if sheet["A" + str(i)].value == "-":
                    if sheet["B" + str(i)].value == "-":
                        sheet["A" + str(i)].value = str(un)
                        sheet["B" + str(i)].value = str(ui)
                        sheet["C" + str(i)].value = 5000
                        sheet["D" + str(i)].value = "안녕하세요"
                        sheet["E" + str(i)].value = str(us)
                        sheet["F" + str(i)].value = ny
                        sheet["G" + str(i)].value = nm
                        sheet["H" + str(i)].value = nd
                        sheet["I" + str(i)].value = ny
                        sheet["J" + str(i)].value = nm
                        sheet["K" + str(i)].value = nd
                        sheet["L" + str(i)].value = nh
                        sheet["M" + str(i)].value = nmm
                        sheet["N" + str(i)].value = 0
                        sheet["O" + str(i)].value = 0
                        sheet["AA" + str(i)].value = 0
                        sheet["AB" + str(i)].value = 0
                        sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                        embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。\n[이용약관](https://discord-mbot.netlify.app/use.html)", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        await message.channel.send(message.author.mention, embed=embed)
                        embed = discord.Embed(title="《업적달성》",description="『유저』 - 가입하기! !업적", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        await message.channel.send(message.author.mention, embed=embed)
                        sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                        file.save("user.xlsx")
                        break
                break










    if msg == "!프로필" or msg == "!profile" or msg == "!プロフィール" or msg == "!轮廓" and not message.content.startswith("!프로필 등록") and not message.content.startswith("!프로필 가입") and not message.content.startswith("!프로필 <@"):
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=un, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="사용중인 언어", value="한국어 🇰🇷", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embeduser.set_thumbnail(url=message.author.avatar_url)

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 1:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"""'s profile""",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="Nickname", value=un, inline=True)
                    embeduser.add_field(name="Discord ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="Status message", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="Achievements", value=u, inline=False)
                    embeduser.add_field(name="Member since", value="{}/{}/{}, {}:{} KST (UTC+09:00) Asia/Seoul".format(um, ud ,uy ,uh, umm), inline=True)
                    embeduser.add_field(name="language in use", value="English 🇺🇸", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embeduser.set_thumbnail(url=message.author.avatar_url)

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"さんのプロフィール",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="ニックネーム", value=un, inline=True)
                    embeduser.add_field(name="ユーザ名", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状態メッセージ", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="業績", value=u, inline=False)
                    embeduser.add_field(name="登録日", value="{}年{}月{}日、{}時{}分 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用している言語", value="日本 🇯🇵", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embeduser.set_thumbnail(url=message.author.avatar_url)

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"的个人资料",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="名称", value=un, inline=True)
                    embeduser.add_field(name="ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状态消息", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="成就", value=u, inline=False)
                    embeduser.add_field(name="会员自", value=str(uy)+"{}年 {}月 {}天, {}小时 {}分钟 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用的语言", value="中国人 🇨🇳", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embeduser.set_thumbnail(url=message.author.avatar_url)

                    await message.channel.send(embed=embeduser)
                    break

                else:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=un, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="사용중인 언어", value="한국어 🇰🇷", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embeduser.set_thumbnail(url=message.author.avatar_url)

                    await message.channel.send(embed=embeduser)
                    break




        #엑셀시트 F = 년 / G = 월 / H = 일

    if message.content.startswith("!프로필"):
        id = message.content[8:-1]
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        un = message.author.name
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                if sheet["Y" + str(i)].value == 0:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=un, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="사용중인 언어", value="한국어 🇰🇷", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break


                elif sheet["Y" + str(i)].value == 1:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"""'s profile""",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="Nickname", value=un, inline=True)
                    embeduser.add_field(name="Discord ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="Status message", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="Achievements", value=u, inline=False)
                    embeduser.add_field(name="Member since", value="{}/{}/{}, {}:{} KST (UTC+09:00) Asia/Seoul".format(um, ud ,uy ,uh, umm), inline=True)
                    embeduser.add_field(name="language in use", value="English 🇺🇸", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"さんのプロフィール",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="ニックネーム", value=un, inline=True)
                    embeduser.add_field(name="ユーザ名", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状態メッセージ", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="業績", value=u, inline=False)
                    embeduser.add_field(name="登録日", value="{}年{}月{}日、{}時{}分 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用している言語", value="日本 🇯🇵", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"的个人资料",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="名称", value=un, inline=True)
                    embeduser.add_field(name="ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状态消息", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="成就", value=u, inline=False)
                    embeduser.add_field(name="会员自", value=str(uy)+"{}年 {}月 {}天, {}小时 {}分钟 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用的语言", value="中国人 🇨🇳", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                else:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=sun, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break
    
    if message.content.startswith("!profile"):
        id = message.content[12:-1]
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        un = message.author.name

        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                if sheet["Y" + str(i)].value == 0:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=un, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="사용중인 언어", value="한국어 🇰🇷", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break


                elif sheet["Y" + str(i)].value == 1:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"""'s profile""",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="Nickname", value=un, inline=True)
                    embeduser.add_field(name="Discord ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="Status message", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="Achievements", value=u, inline=False)
                    embeduser.add_field(name="Member since", value="{}/{}/{}, {}:{} KST (UTC+09:00) Asia/Seoul".format(um, ud ,uy ,uh, umm), inline=True)
                    embeduser.add_field(name="language in use", value="English 🇺🇸", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"さんのプロフィール",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="ニックネーム", value=un, inline=True)
                    embeduser.add_field(name="ユーザ名", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状態メッセージ", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="業績", value=u, inline=False)
                    embeduser.add_field(name="登録日", value="{}年{}月{}日、{}時{}分 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用している言語", value="日本 🇯🇵", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"的个人资料",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="名称", value=un, inline=True)
                    embeduser.add_field(name="ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状态消息", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="成就", value=u, inline=False)
                    embeduser.add_field(name="会员自", value=str(uy)+"{}年 {}月 {}天, {}小时 {}分钟 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用的语言", value="中国人 🇨🇳", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                else:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=sun, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

    if message.content.startswith("!プロフィール"):
        id = message.content[12:-1]
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        un = message.author.name

        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                if sheet["Y" + str(i)].value == 0:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=un, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value=str(uy)+"년 "+str(um)+"월 "+str(ud)+"일 "+str(uh)+"시 "+str(umm)+"분", inline=True)
                    embeduser.add_field(name="사용중인 언어", value="한국어 🇰🇷", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break


                elif sheet["Y" + str(i)].value == 1:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"""'s profile""",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="Nickname", value=un, inline=True)
                    embeduser.add_field(name="Discord ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="Status message", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="Achievements", value=u, inline=False)
                    embeduser.add_field(name="Member since", value="{}/{}/{}, {}:{} KST (UTC+09:00) Asia/Seoul".format(um, ud ,uy ,uh, umm), inline=True)
                    embeduser.add_field(name="language in use", value="English 🇺🇸", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"さんのプロフィール",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="ニックネーム", value=un, inline=True)
                    embeduser.add_field(name="ユーザ名", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状態メッセージ", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="業績", value=u, inline=False)
                    embeduser.add_field(name="登録日", value="{}年{}月{}日、{}時{}分 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用している言語", value="日本 🇯🇵", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"的个人资料",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="名称", value=un, inline=True)
                    embeduser.add_field(name="ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状态消息", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="成就", value=u, inline=False)
                    embeduser.add_field(name="会员自", value=str(uy)+"{}年 {}月 {}天, {}小时 {}分钟 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用的语言", value="中国人 🇨🇳", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                else:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=sun, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

    if message.content.startswith("!profile"):
        id = message.content[7:-1]
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        un = message.author.name

        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                if sheet["Y" + str(i)].value == 0:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=un, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="사용중인 언어", value="한국어 🇰🇷", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break


                elif sheet["Y" + str(i)].value == 1:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"""'s profile""",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="Nickname", value=un, inline=True)
                    embeduser.add_field(name="Discord ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="Status message", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="Achievements", value=u, inline=False)
                    embeduser.add_field(name="Member since", value="{}/{}/{}, {}:{} KST (UTC+09:00) Asia/Seoul".format(um, ud ,uy ,uh, umm), inline=True)
                    embeduser.add_field(name="language in use", value="English 🇺🇸", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 2:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"さんのプロフィール",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="ニックネーム", value=un, inline=True)
                    embeduser.add_field(name="ユーザ名", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状態メッセージ", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="業績", value=u, inline=False)
                    embeduser.add_field(name="登録日", value="{}年{}月{}日、{}時{}分 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用している言語", value="日本 🇯🇵", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                elif sheet["Y" + str(i)].value == 3:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=un+"的个人资料",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="名称", value=un, inline=True)
                    embeduser.add_field(name="ID", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="状态消息", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="成就", value=u, inline=False)
                    embeduser.add_field(name="会员自", value=str(uy)+"{}年 {}月 {}天, {}小时 {}分钟 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.add_field(name="使用的语言", value="中国人 🇨🇳", inline=False)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

                else:
                    sun = sheet["A" + str(i)].value
                    sui = sheet["B" + str(i)].value
                    sum = sheet["C" + str(i)].value

                    uy = sheet["I" + str(i)].value
                    um = sheet["J" + str(i)].value
                    ud = sheet["K" + str(i)].value
                    uh = sheet["L" + str(i)].value
                    umm = sheet["M" + str(i)].value
                    u = sheet["W" + str(i)].value #업적
                    sheet["E" + str(i)].value = str(message.author.avatar_url)
                    c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                    rc = str(random.sample(c, 1))
                    rct = rc[1:8]
                    rct = int(rct)

                    embeduser = discord.Embed(title=sun+"님의 프로필",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                    embeduser.add_field(name="닉네임", value=sun, inline=True)
                    embeduser.add_field(name="아이디", value=sui, inline=False)
                    embeduser.add_field(name="M COIN", value=sum, inline=False)
                    embeduser.add_field(name="상태메세지", value=sheet["D" + str(i)].value, inline=False)
                    embeduser.add_field(name="업적", value=u, inline=False)
                    embeduser.add_field(name="가입일", value="{}년 {}월 {}일, {}시{}분 KST (UTC+09:00) Asia/Seoul".format(uy, um, ud, uh, umm), inline=True)
                    embeduser.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래

                    await message.channel.send(embed=embeduser)
                    break

    
    #_______________________________


# 업적
    if message.content.startswith("!업적") or msg == "!Achievements" or msg == "!業績" or msg == "!成就":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                u = sheet["W" + str(i)].value
                c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff,0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
                rc = str(random.sample(c, 1))
                rct = rc[1:8]
                rct = int(rct)
                if sheet["Y" + str(i)].value == 0:
                    la = "님의 업적"
                    
                elif sheet["Y" + str(i)].value == 1:
                    la = """'s achievements"""

                elif sheet["Y" + str(i)].value == 2:
                    la = "さんの業績"

                elif sheet["Y" + str(i)].value == 3:
                    la = "的成就"

                else:
                    la = "님의 업적"

                embed = discord.Embed(title=un+la,description=u,timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                embed.set_thumbnail(url=message.author.avatar_url)

                await message.channel.send(embed=embed)


# 돈받기----------------------------------------------------
    if message.content.startswith("!돈받기") or message.content.startswith("!get money") or message.content.startswith("!お金を受け取る") or message.content.startswith("!拿钱"):
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        now = datetime.datetime.now()
        ny = now.year
        nm = now.month
        nd = now.day
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["F" + str(i)].value <= ny:
                    if sheet["G" + str(i)].value <= nm:
                        if sheet["H" + str(i)].value < nd:
                            sc = sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1000
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["E" + str(i)].value = str(message.author.avatar_url)
                            bc = sheet["P" + str(i)].value

                            if sheet["Y" + str(i)].value == 0:
                                la = "님 1000M COIN이 지급되었습니다"
                                m = "돈"
                                b = "은행"
                                d = "최근 받은 날짜"
                            elif sheet["Y" + str(i)].value == 1:
                                la = """'s 1000M COIN has been paid"""
                                m = "Money"
                                b = "Bank"
                                d = "Last received date"
                            elif sheet["Y" + str(i)].value == 2:
                                la = "様 1000M COINが支給されました"
                                m = "お金"
                                b = "銀行"
                                d = "最近受信した日付"
                            elif sheet["Y" + str(i)].value == 3:
                                la = "已支付 1000M COIN"
                                m = "钱"
                                b = "银行"
                                d = "最后收到日期"
                            else:
                                la = "님 1000M COIN이 지급되었습니다"
                                m = "돈"
                                b = "은행"
                                d = "최근 받은 날짜"

                            embeduserm = discord.Embed(title=un +la,description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            embeduserm.add_field(name="💵 | {}".format(m), value=str(sc) + "M COIN", inline=False)
                            embeduserm.add_field(name="💳 | {}".format(b), value=str(bc)+"M COIN", inline=False)
                            embeduserm.add_field(name="💰 | {}".format(d),value="{}/{}/{} KST (UTC+09:00) Asia/Seoul".format(ny, nm, nd),inline=False)
                            
                            embeduserm.set_footer(text="{}".format(message.author.name), icon_url=message.author.avatar_url)
                            await message.channel.send(embed=embeduserm)
                        else:
                            if sheet["Y" + str(i)].value == 0:
                                d = "최근 받은 날짜"
                                t = "오늘 이미 받았습니다"

                            elif sheet["Y" + str(i)].value == 1:
                                d = "Last received date"
                                t = "I already got it today"

                            elif sheet["Y" + str(i)].value == 2:
                                d = "最近受信した日付"
                                t = "今日すでに受けました"

                            elif sheet["Y" + str(i)].value == 3:
                                d = "最后收到日期"
                                t = "我今天已经收到了"

                            else:
                                d = "최근 받은 날짜"
                                t = "오늘 이미 받았습니다"
                            await message.channel.send(message.author.mention + "{}.\n{} : {}/{}/{} KST (UTC+09:00) Asia/Seoul".format(t,d,ny, nm, nd))
                        break
        file.save("user.xlsx")

    if message.content.startswith("!구걸") or message.content.startswith("!beg") or message.content.startswith("!頼む") or message.content.startswith("!乞讨") or msg == "!Beg":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                r = random.randrange(0,76)
                sheet["C" + str(i)].value = sheet["C" + str(i)].value + r
                sheet["V" + str(i)].value = sheet["V" + str(i)].value + 1

                if sheet["Y" + str(i)].value == 0:
                    n = "님"
                    nn = " M COIN이 지급되었습니다"
                    ad = "업적달성"
                    val = "『거지』 - 구걸 100번 하기! `!업적`"

                elif sheet["Y" + str(i)].value == 1:
                    n = "'s"
                    nn = " M COIN has been paid"
                    ad = "achievement"
                    val = "『거지』 - Beg 100 times! `!Achievements`"

                elif sheet["Y" + str(i)].value == 2:
                    n = "様"
                    nn = " M COINが支給されました"
                    ad = "業績達成"
                    val = "『거지』 - 頼む 100回する！ `!業績`"

                elif sheet["Y" + str(i)].value == 3:
                    n = "先生"
                    nn = " M COIN已支付"
                    ad = "成就"
                    val = "『거지』 - 乞讨 100次！ `!成就`"

                else:
                    n = "님"
                    nn = " M COIN이 지급되었습니다"
                    ad = "업적달성"
                    val = "『거지』 - 구걸 100번 하기! `!업적`"

                await message.channel.send(message.author.mention+("{} {}{}".format(n, r, nn)))
                if sheet["V" + str(i)].value == 100:
                        embed = discord.Embed(title="**《{}》**".format(ad),description="{}".format(val), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        await message.channel.send(message.author.mention,embed=embed)
                        sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『거지』 "
                break
        file.save("user.xlsx")


# 돈설정 관리자용
    if message.content.startswith("!돈설정"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        mi = 789670002163974145
        c = message.content
        cc = int(c[5:])
        if message.author.id == mi:
            for i in range(1, 1001):
                sheet["C" + str(i)].value = cc
                await message.channel.send(message.author.mention + "님 M COIN 이`"+str(cc)+"M COIN`으로 설정되었습니다")
                break
            file.save("user.xlsx")
        else:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : 권한부족", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.channel.send(embed=embed)


# 돈확인

    if msg == ("!돈") or msg == "!ehs" or msg == "!M COIN" or msg == "!엠코인" or msg == "!M 코인" or msg == "!엠 COIN" or msg == "!MCOIN" or msg == "!m coin" or msg == "!m 코인" or msg == "!엠 coin" or msg == "!money" or msg == "!Money" or msg == "!钱" or msg == "!お金" and not msg == ("!돈받기") and not message.content.startswith("!돈 <@"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        now = datetime.datetime.now()
        ny = now.year
        nm = now.month
        nd = now.day
        c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
        rc = str(random.sample(c, 1))
        rct = rc[1:8]
        rct = int(rct)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"
                elif sheet["Y" + str(i)].value == 1:
                    la = """'s M COIN"""
                    m = "Money"
                    b = "Bank"
                    d = "Last received date"
                elif sheet["Y" + str(i)].value == 2:
                    la = "さんのM COIN"
                    m = "お金"
                    b = "銀行"
                    d = "最近受信した日付"
                elif sheet["Y" + str(i)].value == 3:
                    la = " M COIN"
                    m = "钱"
                    b = "银行"
                    d = "最后收到日期"
                else:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"

                embed = discord.Embed(title=un+"{}".format(la),description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.add_field(name="💵 | {}".format(m), value=str(sc)+"M COIN", inline=False)
                embed.add_field(name="💳 | {}".format(b), value=str(bc)+"M COIN", inline=False)
                embed.add_field(name="💰 | {}".format(d), value="{}/{}/{} KST (UTC+09:00) Asia/Seoul".format(ny, nm, nd),inline=False)
                embed.set_thumbnail(url=message.author.avatar_url) # 프사
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                await message.channel.send(embed=embed)
                break
        file.save("user.xlsx")

    #___

    if message.content.startswith("!돈") and not msg == "!돈" and not msg == "!돈 ":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        id = message.content[6:-1]
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        now = datetime.datetime.now()
        ny = now.year
        nm = now.month
        nd = now.day
        un = message.author.name

        c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
        rc = str(random.sample(c, 1))
        rct = rc[1:8]
        rct = int(rct)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                name = sheet["A" + str(i)].value
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value
                if sheet["Y" + str(i)].value == 0:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"
                elif sheet["Y" + str(i)].value == 1:
                    la = """'s M COIN"""
                    m = "Money"
                    b = "Bank"
                    d = "Last received date"
                elif sheet["Y" + str(i)].value == 2:
                    la = "さんのM COIN"
                    m = "お金"
                    b = "銀行"
                    d = "最近受信した日付"
                elif sheet["Y" + str(i)].value == 3:
                    la = " M COIN"
                    m = "钱"
                    b = "银行"
                    d = "最后收到日期"
                else:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"

                embed = discord.Embed(title=un+"{}".format(la),description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.add_field(name="💵 | {}".format(m), value=str(sc)+"M COIN", inline=False)
                embed.add_field(name="💳 | {}".format(b), value=str(bc)+"M COIN", inline=False)
                embed.add_field(name="💰 | {}".format(d), value="{}/{}/{} KST (UTC+09:00) Asia/Seoul".format(ny, nm, nd),inline=False)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)
                break
        file.save("user.xlsx")

    if message.content.startswith("!Money") or message.content.startswith("!money") and not msg == "!Money" and not msg == "!money":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        id = message.content[10:-1]
        file = openpyxl.load_workbook("user.xlsx")
        un = message.author.name
        sheet = file.active
        now = datetime.datetime.now()
        un = message.author.name

        ny = now.year
        nm = now.month
        nd = now.day
        c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
        rc = str(random.sample(c, 1))
        rct = rc[1:8]
        rct = int(rct)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                name = sheet["A" + str(i)].value
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value
                if sheet["Y" + str(i)].value == 0:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"
                elif sheet["Y" + str(i)].value == 1:
                    la = """'s M COIN"""
                    m = "Money"
                    b = "Bank"
                    d = "Last received date"
                elif sheet["Y" + str(i)].value == 2:
                    la = "さんのM COIN"
                    m = "お金"
                    b = "銀行"
                    d = "最近受信した日付"
                elif sheet["Y" + str(i)].value == 3:
                    la = " M COIN"
                    m = "钱"
                    b = "银行"
                    d = "最后收到日期"
                else:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"

                embed = discord.Embed(title=un+"{}".format(la),description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.add_field(name="💵 | {}".format(m), value=str(sc)+"M COIN", inline=False)
                embed.add_field(name="💳 | {}".format(b), value=str(bc)+"M COIN", inline=False)
                embed.add_field(name="💰 | {}".format(d), value="{}/{}/{} KST (UTC+09:00) Asia/Seoul".format(ny, nm, nd),inline=False)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)
                break
        file.save("user.xlsx")

    if message.content.startswith("!お金") and not msg == "!お金" and not msg == "!お金 ":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        id = message.content[7:-1]
        un = message.author.name
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        now = datetime.datetime.now()
        un = message.author.name

        ny = now.year
        nm = now.month
        nd = now.day
        c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
        rc = str(random.sample(c, 1))
        rct = rc[1:8]
        rct = int(rct)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                name = sheet["A" + str(i)].value
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value
                if sheet["Y" + str(i)].value == 0:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"
                elif sheet["Y" + str(i)].value == 1:
                    la = """'s M COIN"""
                    m = "Money"
                    b = "Bank"
                    d = "Last received date"
                elif sheet["Y" + str(i)].value == 2:
                    la = "さんのM COIN"
                    m = "お金"
                    b = "銀行"
                    d = "最近受信した日付"
                elif sheet["Y" + str(i)].value == 3:
                    la = " M COIN"
                    m = "钱"
                    b = "银行"
                    d = "最后收到日期"
                else:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"

                embed = discord.Embed(title=un+"{}".format(la),description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.add_field(name="💵 | {}".format(m), value=str(sc)+"M COIN", inline=False)
                embed.add_field(name="💳 | {}".format(b), value=str(bc)+"M COIN", inline=False)
                embed.add_field(name="💰 | {}".format(d), value="{}/{}/{} KST (UTC+09:00) Asia/Seoul".format(ny, nm, nd),inline=False)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)
                break
        file.save("user.xlsx")

    if message.content.startswith("!钱") and not msg == "!钱" and not msg == "!钱 ":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        id = message.content[6:-1]
        un = message.author.name
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        now = datetime.datetime.now()
        un = message.author.name

        ny = now.year
        nm = now.month
        nd = now.day
        c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363, 0xff9663)
        rc = str(random.sample(c, 1))
        rct = rc[1:8]
        rct = int(rct)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(id):
                name = sheet["A" + str(i)].value
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value
                if sheet["Y" + str(i)].value == 0:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"
                elif sheet["Y" + str(i)].value == 1:
                    la = """'s M COIN"""
                    m = "Money"
                    b = "Bank"
                    d = "Last received date"
                elif sheet["Y" + str(i)].value == 2:
                    la = "さんのM COIN"
                    m = "お金"
                    b = "銀行"
                    d = "最近受信した日付"
                elif sheet["Y" + str(i)].value == 3:
                    la = " M COIN"
                    m = "钱"
                    b = "银行"
                    d = "最后收到日期"
                else:
                    la = "님의 M COIN"
                    m = "돈"
                    b = "은행"
                    d = "최근 받은 날짜"

                embed = discord.Embed(title=un+"{}".format(la),description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.add_field(name="💵 | {}".format(m), value=str(sc)+"M COIN", inline=False)
                embed.add_field(name="💳 | {}".format(b), value=str(bc)+"M COIN", inline=False)
                embed.add_field(name="💰 | {}".format(d), value="{}/{}/{} KST (UTC+09:00) Asia/Seoul".format(ny, nm, nd),inline=False)
                embed.set_footer(text=message.author, icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)
                break
        file.save("user.xlsx")


# 기억 --------------------------------------------------

    if message.content.startswith("!기억") or message.content.startswith("!배워") or message.content.startswith("!学び") or message.content.startswith("!記憶") or message.content.startswith("!记忆") or message.content.startswith("!学习") and not message.content.startswith("!기억삭제") and not message.content.startswith("!기억 제작자") and not message.content.startswith("!기억 엠봇아 핑") and not message.content.startswith("!기억 마인잡지")  and not message.content.startswith("!기억 마잡") and not message.content.startswith("!기억 마인잡체") and not message.content.startswith("!기억 잡체") and not message.content.startswith("!기억 잡지"):
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:
                    w = "경고!"
                    a = "이라고 등록하겠습니까?"
                    c = "취소!"
                    o = "완료!"
                elif sheet["Y" + str(i)].value == 1:
                    w = "Warning!"
                    a = "Would you like to register?"
                    c = "Cancel!"
                    o = "Complete!"
                elif sheet["Y" + str(i)].value == 2:
                    w = "警告！"
                    a = "と登録しますか？"
                    c = "キャンセル！"
                    o = "完了！"
                elif sheet["Y" + str(i)].value == 3:
                    w = "警告！"
                    a = "你想注册吗？"
                    c = "取消！"
                    o = "完全的！"
                else:
                    w = "경고!"
                    a = "이라고 등록하겠습니까?"
                    c = "취소!"
                    o = "완료!"


        file = openpyxl.load_workbook("mbot.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        ui = message.author.id
        ban = 0
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        try:
            if ui != ban:
                for i in range(1, 1001):
                    if len(learn) == 3:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = ""
                            sheet["D" + str(i)].value = ""
                            sheet["E" + str(i)].value = ""
                            learn_name = sheet["F" + str(i)].value
                            embedlearn1 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+"`"+" **{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn1)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")
                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o)+"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))


                    elif len(learn) == 4:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = learn[3]
                            sheet["D" + str(i)].value = ""
                            sheet["E" + str(i)].value = ""
                            embedlearn2 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+" "+learn[3]+"`"+ "**{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn2)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(
                                    reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o)+"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+" "+learn[3]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))
                            break


                    elif len(learn) == 5:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = learn[3]
                            sheet["D" + str(i)].value = learn[4]
                            sheet["E" + str(i)].value = ""
                            embedlearn3 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+"`"+ "**{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn3)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(
                                    reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o)+"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))
                            break


                    elif len(learn) == 6:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = learn[3]
                            sheet["D" + str(i)].value = learn[4]
                            sheet["E" + str(i)].value = learn[5]
                            embedlearn4 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+" "+learn[5]+"`"+ "**{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn4)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o) +"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+" "+learn[5]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))
                            break

            elif ui == ban:
                await message.channel.send("엠봇 기억서비스를 사용할수없는 사용자 입니다\nYou are a user who cannot use the mbot memory service")
        
        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)


    if message.content.startswith("!memory") and not message.content.startswith("!erase memory") and not message.content.startswith("!기억 제작자") and not message.content.startswith("!erasememory") and not message.content.startswith("!memory 마인잡지")  and not message.content.startswith("!memory 마잡") and not message.content.startswith("!memory 마인잡체") and not message.content.startswith("!memory 잡체") and not message.content.startswith("!memory 잡지"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:
                    w = "경고!"
                    a = "이라고 등록하겠습니까?"
                    c = "취소!"
                    o = "완료!"
                elif sheet["Y" + str(i)].value == 1:
                    w = "Warning!"
                    a = "Would you like to register?"
                    c = "Cancel!"
                    o = "Complete!"
                elif sheet["Y" + str(i)].value == 2:
                    w = "警告！"
                    a = "と登録しますか？"
                    c = "キャンセル！"
                    o = "完了！"
                elif sheet["Y" + str(i)].value == 3:
                    w = "警告！"
                    a = "你想注册吗？"
                    c = "取消！"
                    o = "完全的！"
        file = openpyxl.load_workbook("mbot.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        ui = message.author.id
        ban = 0

        try:
            if ui != ban:
                for i in range(1, 1001):
                    if len(learn) == 3:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = ""
                            sheet["D" + str(i)].value = ""
                            sheet["E" + str(i)].value = ""
                            learn_name = sheet["F" + str(i)].value
                            embedlearn1 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+"`"+" **{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn1)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")
                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o)+"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))


                    elif len(learn) == 4:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = learn[3]
                            sheet["D" + str(i)].value = ""
                            sheet["E" + str(i)].value = ""
                            embedlearn2 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+" "+learn[3]+"`"+ "**{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn2)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(
                                    reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o)+"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+" "+learn[3]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))
                            break


                    elif len(learn) == 5:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = learn[3]
                            sheet["D" + str(i)].value = learn[4]
                            sheet["E" + str(i)].value = ""
                            embedlearn3 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+"`"+ "**{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn3)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(
                                    reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o)+"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))
                            break


                    elif len(learn) == 6:
                        if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                            sheet["A" + str(i)].value = learn[1]
                            sheet["B" + str(i)].value = learn[2]
                            sheet["C" + str(i)].value = learn[3]
                            sheet["D" + str(i)].value = learn[4]
                            sheet["E" + str(i)].value = learn[5]
                            embedlearn4 = discord.Embed(title="**{}**".format(w), description="`"+learn[1]+"`"+" **~>** "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+" "+learn[5]+"`"+ "**{}**".format(a), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                            msg = await message.channel.send(embed=embedlearn4)
                            await msg.add_reaction("⭕")
                            await msg.add_reaction("❌")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(reaction.emoji) == '❌'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                                await message.channel.send(message.author.mention + " {}".format(c))
                                break
                            else:
                                if str(reaction.emoji) == "⭕":
                                    await reaction.message.channel.send(message.author.mention + " {} ".format(o) +"["+"`"+learn[1]+"`"+" ~> "+"`"+learn[2]+" "+learn[3]+" "+learn[4]+" "+learn[5]+"`"+"]")
                                    ui = list(str(user.id))
                                    sheet["F" + str(i)].value = user.name+str("#")+ui[0]+ui[1]+ui[2]+ui[3]+ui[4]+ui[5]
                                    sheet["G" + str(i)].value = user.id
                                    file.save("mbot.xlsx")
                                    break
                                elif str(reaction.emoji) =='❌':
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + " {}".format(c))
                            break

            elif ui == ban:
                await message.channel.send("엠봇 기억서비스를 사용할수없는 사용자 입니다\nYou are a user who cannot use the mbot memory service")

        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)
        



    if message.content.startswith("엠봇아 ") or message.content.startswith("mbot ") or message.content.startswith("エムボトア ") or message.content.startswith("机器人 ") and not message.content.startswith("엠봇아 도움말") and not msg == "엠봇아 help" and not msg == "엠봇아 Help" and not msg == "엠봇아 코로나" and not msg == "엠봇아 코로나19" and not msg == "엠봇아 covid" and not msg == "엠봇아 covid19" and not msg == "코로나" and not msg == "코로나19" and not msg == "covid" and not msg == "covid19" and not msg == "엠봇아 안녕" and not msg == "엠봇아 ㅎㅇ" and not msg == "엠봇아 잘가" and not msg == "엠봇아 ㅂㅂ" and not msg == "엠봇아 ㅂㅇ" and not msg == "엠봇아" and not msg == "엠봇" and not msg == "mbot" and not msg == '핑' and not msg == 'ping' and not msg == "엠봇아 가위바위보" and not message.content.startswith("엠봇아 마인잡지")  and not message.content.startswith("엠봇아 마잡") and not message.content.startswith("엠봇아 마인잡체") and not message.content.startswith("엠봇아 잡체") and not message.content.startswith("엠봇아 잡지") and not msg == "엠봇아 머랭이" and not msg == "엠봇아 랭" and not msg == "엠봇아 멀랭이" and not msg == "엠봇아 " and not  msg == "엠봇아 무말랭이" and not  msg == "엠봇아 머랭" and not msg == "엠봇아 도움"and not msg == "마인매거진" and not msg == "Mine메거진" and not message.content.startswith("마인Magazine") and not msg == "MINEMAGAZINE" and not msg == "엠봇아 제작자":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:
                    w = "등록자"

                elif sheet["Y" + str(i)].value == 1:
                    w = "Registrant"

                elif sheet["Y" + str(i)].value == 2:
                    w = "登録者"

                elif sheet["Y" + str(i)].value == 3:
                    w = "注册人"

                else:
                    w = "등록자"


        file = openpyxl.load_workbook("mbot.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A" + str(i)].value == memory[1]:
                sb = sheet["B" + str(i)].value or " "
                sc = sheet["C" + str(i)].value or " "
                sd = sheet["D" + str(i)].value or " "
                se = sheet["E" + str(i)].value or " "
                learn_name = sheet["F" + str(i)].value
                try:
                    await message.channel.send(sb + " " + sc + " " + sd + " " + se + "\n```{} : ".format(w)+learn_name+""+"```")
                    break
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)


    if message.content.startswith("!기억삭제") or message.content.startswith("!erasememory") or message.content.startswith("!記憶の削除") or message.content.startswith("!擦除记忆") or message.content.startswith("!erase"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:
                    w = "기억이 삭제되었습니다."
                    d = "자신이 등록한것만 삭제할수 있습니다!"

                elif sheet["Y" + str(i)].value == 1:
                    w = "Memory has been erased."
                    d = "You can only delete what you have registered!"

                elif sheet["Y" + str(i)].value == 2:
                    w = "記憶が削除されました。"
                    d = "自分が登録しただけ削除することができます！"

                elif sheet["Y" + str(i)].value == 3:
                    w = "记忆已被抹去。"
                    d = "您只能删除已注册的内容！"

                else:
                    w = "기억이 삭제되었습니다."
                    d = "자신이 등록한것만 삭제할수 있습니다!"


        file = openpyxl.load_workbook("mbot.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        ui = message.author.id
        mi = 789670002163974145

        for i in range(1, 1001):
            if sheet["A" + str(i)].value == str(memory[1]):
                if sheet["G" + str(i)].value == ui or sheet["G" + str(i)].value == mi:
                    sheet["A" + str(i)].value = "-"
                    sheet["B" + str(i)].value = " "
                    sheet["C" + str(i)].value = " "
                    sheet["D" + str(i)].value = " "
                    sheet["E" + str(i)].value = " "
                    sheet["F" + str(i)].value = " "
                    sheet["G" + str(i)].value = " "
                    try:
                        await message.channel.send("{}".format(w))
                        
                    except:
                        await message.channel.send("기억이 삭제되었습니다.")
                    
                    file.save("mbot.xlsx")
                    
                elif message.author.id == mi:
                    sheet["A" + str(i)].value = "-"
                    sheet["B" + str(i)].value = " "
                    sheet["C" + str(i)].value = " "
                    sheet["D" + str(i)].value = " "
                    sheet["E" + str(i)].value = " "
                    sheet["F" + str(i)].value = " "
                    sheet["G" + str(i)].value = " "
                    await message.channel.send("관리자 권한으로 기억이 삭제되었습니다!")
                    file.save("mbot.xlsx")
                else:
                    try:
                        await message.channel.send(message.author.mention + f'{(d)}')
                        break
                    except:
                        await message.channel.send(message.author.mention + "자신이 등록한것만 삭제할수 있습니다!")
                        break


# 가위바위보 -------------------------------------------------------

    if message.content.startswith('!가위바위보') or message.content.startswith('!Rock Paper Scissors') or message.content.startswith('!じゃんけん') or message.content.startswith('!剪刀石头布'):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:                    
                    rsp = ["가위", "바위", "보"]
                    embed = discord.Embed(title="가위바위보", description="가위바위보를 합니다 10초내로 (가위/바위/보)를 써주세요!\n승리 100 M COIN | 무승부 50 M COIN | 패배 10 M COIN", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x9e3fe0)
                    channel = message.channel
                    msg1 = await message.channel.send(embed=embed)
                    def check(m):
                        return m.author == message.author and m.channel == channel
                    try:
                        msg2 = await client.wait_for('message', timeout=10.0, check=check)
                    except asyncio.TimeoutError:
                        await msg1.delete()
                        embed = discord.Embed(title="가위바위보", description="10초가 지났네요...!", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                        await message.channel.send(embed=embed)
                        return
                    else:
                        await msg1.delete()
                        bot_rsp = str(random.choice(rsp))
                        user_rsp = str(msg2.content)
                        answer = ""
                        if bot_rsp == user_rsp:
                            embed = discord.Embed(title=bot_rsp + " VS " + user_rsp, description="아쉽지만 비겼습니다\n+50 M COIN.\n만약 아직 엠봇서비스에 가입하지 않았다면\n`!프로필 가입`을 이용해주세요", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x9c9c9c)
                            await message.channel.send(embed=embed)
                            file = openpyxl.load_workbook("user.xlsx")
                            sheet = file.active
                            ui = message.author.id
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value = sheet["C" + str(i)].value+50
                            file.save("user.xlsx")

                        elif (bot_rsp == "가위" and user_rsp == "바위") or (bot_rsp == "보" and user_rsp == "가위") or (bot_rsp == "바위" and user_rsp == "보"):
                            embed = discord.Embed(title=bot_rsp + " VS " + user_rsp, description=message.author.name+"님  🎉축하드립니다!🎉\n+100 M COIN\n만약 아직 엠봇서비스에 가입하지 않았다면\n`!프로필 가입`을 이용해주세요",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xe4d923)
                            await message.channel.send(message.author.mention,embed=embed)
                            file = openpyxl.load_workbook("user.xlsx")
                            sheet = file.active
                            ui = message.author.id
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value = sheet["C" + str(i)].value+100
                            file.save("user.xlsx")

                        elif (bot_rsp == "바위" and user_rsp == "가위") or (bot_rsp == "가위" and user_rsp == "보") or (bot_rsp == "보" and user_rsp == "바위"):
                            embed = discord.Embed(title=bot_rsp + " VS " + user_rsp, description="제가 이겼습니다!\n+10 M COIN\n만약 아직 엠봇서비스에 가입하지 않았다면\n`!프로필 가입`을 이용해주세요", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x000000)
                            await message.channel.send(message.author.mention,embed=embed)
                            file = openpyxl.load_workbook("user.xlsx")
                            sheet = file.active
                            ui = message.author.id
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value = sheet["C" + str(i)].value+10
                            file.save("user.xlsx")

                        else:
                            embed = discord.Embed(title="가위바위보", description="가위, 바위, 보 중에서만 내셔야죠...", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xff0000)
                            await message.channel.send(message.author.mention,embed=embed)
                            return

                elif sheet["Y" + str(i)].value == 1:
                    embed = discord.Embed(title="**⚠️Warning!⚠️**", description="Rock Paper Scissors is supported only when the default language is Korean!",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff600)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(message.author.mention, embed=embed)

                elif sheet["Y" + str(i)].value == 2:
                    embed = discord.Embed(title="**⚠️警告！⚠️**", description="じゃんけん見るデフォルトの言語が韓国語である状態でのみサポートされます！",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff600)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(message.author.mention, embed=embed)

                elif sheet["Y" + str(i)].value == 3:
                    embed = discord.Embed(title="**⚠️警告！⚠️**", description="剪刀石头布 仅在默认语言为韩语时才支持！",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff600)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(message.author.mention, embed=embed)

                else:
                    rsp = ["가위", "바위", "보"]
                    embed = discord.Embed(title="가위바위보", description="가위바위보를 합니다 10초내로 (가위/바위/보)를 써주세요!\n승리 100 M COIN | 무승부 50 M COIN | 패배 10 M COIN", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x9e3fe0)
                    channel = message.channel
                    msg1 = await message.channel.send(embed=embed)
                    def check(m):
                        return m.author == message.author and m.channel == channel
                    try:
                        msg2 = await client.wait_for('message', timeout=10.0, check=check)
                    except asyncio.TimeoutError:
                        await msg1.delete()
                        embed = discord.Embed(title="가위바위보", description="10초가 지났네요...!", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                        await message.channel.send(embed=embed)
                        return
                    else:
                        await msg1.delete()
                        bot_rsp = str(random.choice(rsp))
                        user_rsp = str(msg2.content)
                        answer = ""
                        if bot_rsp == user_rsp:
                            embed = discord.Embed(title=bot_rsp + " VS " + user_rsp, description="아쉽지만 비겼습니다\n+50 M COIN.\n만약 아직 엠봇서비스에 가입하지 않았다면\n`!프로필 가입`을 이용해주세요", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x9c9c9c)
                            await message.channel.send(embed=embed)
                            file = openpyxl.load_workbook("user.xlsx")
                            sheet = file.active
                            ui = message.author.id
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value = sheet["C" + str(i)].value+50
                            file.save("user.xlsx")

                        elif (bot_rsp == "가위" and user_rsp == "바위") or (bot_rsp == "보" and user_rsp == "가위") or (bot_rsp == "바위" and user_rsp == "보"):
                            embed = discord.Embed(title=bot_rsp + " VS " + user_rsp, description=message.author.name+"님  🎉축하드립니다!🎉\n+100 M COIN\n만약 아직 엠봇서비스에 가입하지 않았다면\n`!프로필 가입`을 이용해주세요",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xe4d923)
                            await message.channel.send(message.author.mention,embed=embed)
                            file = openpyxl.load_workbook("user.xlsx")
                            sheet = file.active
                            ui = message.author.id
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value = sheet["C" + str(i)].value+100
                            file.save("user.xlsx")

                        elif (bot_rsp == "바위" and user_rsp == "가위") or (bot_rsp == "가위" and user_rsp == "보") or (bot_rsp == "보" and user_rsp == "바위"):
                            embed = discord.Embed(title=bot_rsp + " VS " + user_rsp, description="제가 이겼습니다!\n+10 M COIN\n만약 아직 엠봇서비스에 가입하지 않았다면\n`!프로필 가입`을 이용해주세요", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x000000)
                            await message.channel.send(message.author.mention,embed=embed)
                            file = openpyxl.load_workbook("user.xlsx")
                            sheet = file.active
                            ui = message.author.id
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value = sheet["C" + str(i)].value+10
                            file.save("user.xlsx")

                        else:
                            embed = discord.Embed(title="가위바위보", description="가위, 바위, 보 중에서만 내셔야죠...", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xff0000)
                            await message.channel.send(message.author.mention,embed=embed)
                            return


# 상점 -------------------------

    if message.content.startswith("!상점") or message.content.startswith("!shop") or message.content.startswith("!店") or message.content.startswith("!店铺") or message.content.startswith("!mart"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):

                if sheet["Y" + str(i)].value == 0:
                    embed = discord.Embed(title="**[상점]**",description="상점 목록", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="전자제품 상점(전자상점)💻", value="전자제품 상점 입니다 !전자상점 또는\n아래 이모티콘을 클릭해 주세요", inline=False)
                    embed.add_field(name="음식상점 🍱", value="음식상점 입니다 !음식상점 또는\n아래 이모티콘을 클릭해 주세요", inline=False)


                    #__ 전자상점
                    embedc = discord.Embed(title="**[전자제품 상점]**", description="전자제품 목록",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedc.add_field(name="컴퓨터💻", value="컴퓨터 입니다 50,000 M COIN", inline=False)
                    embedc.add_field(name="M DOW <:MDOW:841548472854511616>",value="컴퓨터를 사용하기 위한 운영체체(os) 입니다 1,000 M COIN", inline=False)


                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="라면 🍜", value="맛있다 [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="아이스크림 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="피자 🍕", value="치즈ㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="햄버거 🍔", value="굿!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="핫도그 🌭", value="맛있다! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="우유 🥛", value="!!우유~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="추가중...", value="추가중...", inline=False)


                elif sheet["Y" + str(i)].value == 1:
                    embed = discord.Embed(title="**[shop]**",description="store list", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="Electronics store 💻", value="It's an electronics store. !e-shop or\nClick the emoticon below", inline=False)
                    embed.add_field(name="food shop 🍱", value="It is a food store. !food shop or\nClick the emoji below", inline=False)


                    #__ 전자상점
                    embedc = discord.Embed(title="**[electronics store]**", description="Electronics List",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedc.add_field(name="computer 💻", value="it's a computer 50,000 M COIN", inline=False)
                    embedc.add_field(name="M DOW <:MDOW:841548472854511616>",value="It is the operating system (OS) for using the computer. 1,000 M COIN", inline=False)


                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP List-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="Ramen 🍜", value="Best! [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="Ice cream 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="Pizza 🍕", value="Cheeseㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="Next | Before", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP List-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="Hamburger 🍔", value="GOOD!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="Hotdog 🌭", value="Delicious!! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="Milk 🥛", value="!!milk~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="Next | Before", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP List-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="being added...", value="being added...", inline=False)

                    
                elif sheet["Y" + str(i)].value == 2:
                    embed = discord.Embed(title="**[店]**",description="店リスト", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="電子製品店💻", value="電子製品店です。 ！電子商店 や\n下の絵文字をクリックしてください", inline=False)
                    embed.add_field(name="食料品店 🍱", value="食料品店です。 !食料品店 または\ n下の絵文字をクリックしてください", inline=False)


                    #__ 전자상점
                    embedc = discord.Embed(title="**[電子製品店]**", description="電子製品のリスト",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedc.add_field(name="コンピュータ 💻", value="コンピュータです 50,000 M COIN", inline=False)
                    embedc.add_field(name="M DOW <:MDOW:841548472854511616>",value="コンピュータを使用するための大量のオペレーティングシステム（os）です。 1,000 M COIN", inline=False)
                    
                    
                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食品のリスト-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="ラーメン 🍜", value="おいしい! [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="アイスクリーム 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="ピザ 🍕", value="チーズーー! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="次の | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食品のリスト-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="ハンバーガー 🍔", value="グッド!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="ホットドッグ 🌭", value="美味しい!! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="牛乳 🥛", value="!!牛乳~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="次の | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食品のリスト-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="追加中...", value="追加中...", inline=False)


                elif sheet["Y" + str(i)].value == 3:
                    embed = discord.Embed(title="**[店铺]**",description="商店清单", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="电子商店 💻", value="这是一家电子产品商店。 !电子商店 或\n点击下方表情", inline=False)
                    embed.add_field(name="食品店 🍱", value="这是一家食品店。 !食品店 或\n点击下面的表情符号", inline=False)


                    #__ 전자상점
                    embedc = discord.Embed(title="**[电子商店]**", description="电子产品清单",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedc.add_field(name="计算机 💻", value="这是一台电脑 50,000 M COIN", inline=False)
                    embedc.add_field(name="M DOW <:MDOW:841548472854511616>",value="它是用于使用计算机的操作系统 (OS)。 1,000 M COIN", inline=False)
                                        

                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食物清单-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="拉面 🍜", value="很好！ [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="冰淇淋 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="比萨 🍕", value="起司ㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="下一个 | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食物清单-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="汉堡包 🍔", value="好的!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="热狗 🌭", value="很好! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="牛奶 🥛", value="!!牛奶~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="下一个 | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食物清单-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="添加...", value="添加...", inline=False)


                else:
                    embed = discord.Embed(title="**[상점]**",description="상점 목록", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="전자제품 상점(전자상점)💻", value="전자제품 상점 입니다 !전자상점 또는\n아래 이모티콘을 클릭해 주세요", inline=False)
                    embed.add_field(name="음식상점 🍱", value="음식상점 입니다 !음식상점 또는\n아래 이모티콘을 클릭해 주세요", inline=False)


                    #__ 전자상점
                    embed = discord.Embed(title="**[전자제품 상점]**", description="전자제품 목록",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="컴퓨터 💻", value="컴퓨터 입니다 50,000 M COIN", inline=False)
                    embed.add_field(name="M DOW <:MDOW:841548472854511616>",value="컴퓨터를 사용하기 위한 운영체체(os) 입니다 1,000 M COIN", inline=False)


                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="라면🍜", value="맛있다 [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="아이스크림🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="피자🍕", value="치즈ㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="햄버거🍔", value="굿!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="핫도그🌭", value="맛있다! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="우유🥛", value="!!우유~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="추가중...", value="추가중...", inline=False)



        msg = await message.channel.send(embed=embed)
        await msg.add_reaction("💻")
        await msg.add_reaction("🍱")
        await msg.add_reaction("❌")
        def check(reaction, user):
            if user.bot == 1:  # 봇이면 패스
                return None
            return user == message.author and str(reaction.emoji) == '💻' or user == message.author and str(reaction.emoji) == '🍱' or user == message.author and str(reaction.emoji) == '❌'

        try:
            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
        except asyncio.TimeoutError:
            embed = discord.Embed(title="Cancel", description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
            await msg.delete()
            await message.channel.send(embed=embed)

        else:
            if str(reaction.emoji) == "💻":

                await msg.delete()
                msg = await message.channel.send(embed=embedc)
                await msg.add_reaction("💻")
                await msg.add_reaction("<:MDOW:841548472854511616>")
                await msg.add_reaction("❌")
                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                ui = message.author.id

                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '💻' or user == message.author and str(reaction.emoji) == '<:MDOW:841548472854511616>' or user == message.author and str(reaction.emoji) == '❌'

                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="Cancel", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    await msg.delete()
                    await message.channel.send(embed=embed)

                else:
                    if str(reaction.emoji) == "💻":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                com = sheet["AA" + str(i)].value

                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "컴퓨터를 구매"
                                    a_211 = "남은"

                                    a_3 = "님 이미 컴퓨터가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy a computer"
                                    a_211 = "remaining "

                                    a_3 = "already have a computer.\nPlease use `!computer`"
                                    a_4 = "does not have enough M COIN. Please use `!get money`"
                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "コンピュータを購入する"
                                    a_211 = "残り "

                                    a_3 = "様のコンピュータが既にあります。\n`!コンピュータ` を使用してください"
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"
                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买电脑"
                                    a_211 = "剩余 "

                                    a_3 = "您的计算机已经存在。\n请用 `!计算机`"
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"
                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "컴퓨터를 구매"
                                    a_211 = "남은"

                                    a_3 = "님 이미 컴퓨터가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                if sum >= 50000 and com != 1:  # 돈o 컴x
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 50000
                                    sheet["AA" + str(i)].value = 1
                                    embedsnackshop2 = discord.Embed(title="💻 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} 💻 {} [-50,000 M COIN]\n  {}M COIN : ".format(a_2, a_21, a_211) + str(summ))

                                elif sum >= 50000 and com == 1:  # 돈o 컴o
                                    await message.channel.send(message.author.mention + (a_3))

                                elif sum <= 50000 and com == 1:  # 돈x 컴o
                                    await message.channel.send(message.author.mention + (a_3))

                                else:  # 돈x 컴x
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")


                    if str(reaction.emoji) == "<:MDOW:841548472854511616>":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                os = sheet["AB" + str(i)].value


                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "M DOW os를 구매"
                                    a_211 = "남은"

                                    a_3 = "님 이미 M DOW os가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy M dow os"
                                    a_211 = "remaining "

                                    a_3 = "already have M DOW os.\nPlease use `!os install` or `!computer`"
                                    a_4 = "does not have enough M COIN. Please use `!get money`"
                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "M DOW osを購入"
                                    a_211 = "残り "

                                    a_3 = "さんはM DOW osがあります。\n`!osのインストール` または `!コンピュータ` 使用してください"
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"
                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "购买 M dow os"
                                    a_211 = "剩余 "

                                    a_3 = "您已经拥有 M DOW 操作系统。\n请用 `!操作系统安装` 要么 `!计算机`"
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"
                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "M DOW os를 구매"
                                    a_211 = "남은"

                                    a_3 = "님 이미 M DOW os가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                                if sum >= 1000 and os != 1:  # 돈o 컴x
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 1000
                                    sheet["AB" + str(i)].value = 1
                                    embedsnackshop2 = discord.Embed(title="<:MDOW:841548472854511616> {}".format(a_1),description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} <:MDOW:841548472854511616> {} [-1,000 M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))

                                elif sum >= 1000 and os == 1:  # 돈o MDOW o
                                    await message.channel.send(message.author.mention + (a_3))

                                elif sum <= 1000 and os == 1:  # 돈x MDOW o
                                    await message.channel.send(message.author.mention + (a_3))

                                else:  # 돈x MDOWx
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")


                    elif str(reaction.emoji) == "❌":
                        embed = discord.Embed(title="Cancel", description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                        await msg.delete()
                        await message.channel.send(embed=embed)


            elif str(reaction.emoji) == "🍱":
                

                msg = await message.channel.send(embed=embedsnackshop)
                await msg.add_reaction("🍜")
                await msg.add_reaction("🍦")
                await msg.add_reaction("🍕")
                await msg.add_reaction("▶")

                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                ui = message.author.id

                def check(reaction, user):
                    if user.bot == 1:
                        return None
                    return user == message.author and str(reaction.emoji) == '🍜' or user == message.author and str(
                        reaction.emoji) == '🍦' or user == message.author and str(
                        reaction.emoji) == '🍕' or user == message.author and str(reaction.emoji) == '▶'

                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=30.0, check=check)
                except asyncio.TimeoutError:
                    embedsnackshopn = discord.Embed(title="Cancel", description="",
                                                    timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                                    color=0x83fff8)
                    await msg.delete()
                    await message.channel.send(embed=embedsnackshopn)

                else:
                    if str(reaction.emoji) == "🍜":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                
                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "라면을 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy ramen"
                                    a_211 = "remaining "
                                    a_4 = "does not have enough M COIN. Please use `!get money`"

                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "ラーメンを購入"
                                    a_211 = "残り "
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买拉面"
                                    a_211 = "剩余 "
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "라면을 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                if sum >= 10:
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 10
                                    embedsnackshop2 = discord.Embed(title="🍜 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(
                                        user.name + "{} 🍜 {} [-10M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                                    break
                                else:
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")


                    elif str(reaction.emoji) == "🍦":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "아이스크림을 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy ice cream"
                                    a_211 = "remaining "
                                    a_4 = "does not have enough M COIN. Please use `!get money`"

                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "アイスクリームを購入"
                                    a_211 = "残り "
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买冰淇淋"
                                    a_211 = "剩余 "
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "아이스크림을 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                                if sum >= 5:
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 5
                                    embedsnackshop2 = discord.Embed(title="🍦 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} 🍦 {} [-5M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                                    break
                                else:
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")


                    elif str(reaction.emoji) == "🍕":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "피자를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy a pizza"
                                    a_211 = "remaining "
                                    a_4 = "does not have enough M COIN. Please use `!get money`"

                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "ピザを購入"
                                    a_211 = "残り "
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买一个披萨"
                                    a_211 = "剩余 "
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "피자를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                                if sum >= 70:
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 70
                                    embedsnackshop2 = discord.Embed(title="🍕 {}".format(a_1), description="",
                                                                    timestamp=datetime.datetime.now(
                                                                        pytz.timezone('UTC')), color=0x83fff8)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} 🍕 {} [-70M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                                    break
                                else:
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")

                    if str(reaction.emoji) == "▶":
                        await msg.delete()
                        msg = await message.channel.send(embed=embedsnackshop2)
                        await msg.add_reaction("🍔")
                        await msg.add_reaction("🌭")
                        await msg.add_reaction("🥛")
                        await msg.add_reaction("▶")

                # p2

                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '🍔' or user == message.author and str(reaction.emoji) == '🌭' or user == message.author and str(reaction.emoji) == '🥛' or user == message.author and str(reaction.emoji) == '▶'

                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=20.0, check=check)
                except asyncio.TimeoutError:
                    embedsnackshopn = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83ffb5)
                    await msg.delete()
                    await message.channel.send(embed=embedsnackshopn)

                else:

                    if str(reaction.emoji) == "🍔":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "햄버거를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy a hamburger"
                                    a_211 = "remaining "
                                    a_4 = "does not have enough M COIN. Please use `!get money`"

                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "ハンバーガーを購入"
                                    a_211 = "残り "
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买一个汉堡"
                                    a_211 = "剩余 "
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "햄버거를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                                if sum >= 40:
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 40
                                    embedsnackshop2 = discord.Embed(title="🍔 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} 🍔 {} [-40M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                                    break
                                else:
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")

                    elif str(reaction.emoji) == "🌭":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "핫도그를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy a hot dog"
                                    a_211 = "remaining "
                                    a_4 = "does not have enough M COIN. Please use `!get money`"

                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "ホットドッグを購入"
                                    a_211 = "残り "
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买热狗"
                                    a_211 = "剩余 "
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "핫도그를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"
                                    
                                if sum >= 30:
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 30
                                    embedsnackshop2 = discord.Embed(title="🌭 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} 🌭 {} [-30M COIN]\n{} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                                    break
                                else:
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")

                    elif str(reaction.emoji) == "🥛":
                        for i in range(1, 1001):
                            if sheet["B" + str(i)].value == str(ui):
                                sum = sheet["C" + str(i)].value
                                if sheet["Y" + str(i)].value == 0:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "우유를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                                elif sheet["Y" + str(i)].value == 1:
                                    a_1 = "Purchase complete!"
                                    a_2 = """'s"""
                                    a_21 = "buy milk"
                                    a_211 = "remaining "
                                    a_4 = "does not have enough M COIN. Please use `!get money`"

                                    
                                elif sheet["Y" + str(i)].value == 2:
                                    a_1 = "購入完了！"
                                    a_2 = "様が"
                                    a_21 = "牛乳を購入"
                                    a_211 = "残り "
                                    a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                                    
                                elif sheet["Y" + str(i)].value == 3:
                                    a_1 = "购买完成！"
                                    a_2 = "尼姆"
                                    a_21 = "买牛奶"
                                    a_211 = "剩余 "
                                    a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                                    
                                else:
                                    a_1 = "구매완료!"
                                    a_2 = "님이"
                                    a_21 = "우유를 구매"
                                    a_211 = "남은"
                                    a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"
                                    
                                if sum >= 30:
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 10
                                    embedsnackshop2 = discord.Embed(title="🥛 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                                    await msg.delete()
                                    await message.channel.send(embed=embedsnackshop2)
                                    await reaction.message.channel.send(user.name + "{} 🥛 {} [-10M COIN]\n{} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                                    break
                                else:
                                    await msg.delete()
                                    await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                        file.save("user.xlsx")

                    elif str(reaction.emoji) == "▶":
                        await msg.delete()
                        msg = await message.channel.send(embed=embedsnackshop3)
                        await msg.add_reaction("⛔")

                # p3

                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '⛔'

                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
                except asyncio.TimeoutError:
                    embedsnackshopn = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xcfff83)
                    await msg.delete()
                    await message.channel.send(embed=embedsnackshopn)

                else:
                    if str(reaction.emoji) == "⛔":
                        embedsnackshopn = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xcfff83)
                        await msg.delete()
                        await message.channel.send(embed=embedsnackshopn)


            elif str(reaction.emoji) == "❌":
                embed = discord.Embed(title="Cancel", description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                await msg.delete()
                await message.channel.send(embed=embed)


    if message.content.startswith("!전자상점") or message.content.startswith("!전자제품 상점") or message.content.startswith("!전자 상점") or message.content.startswith("!e-shop") or message.content.startswith("!電子商店") or message.content.startswith("!电子商店") or message.content.startswith("!e shop"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                sc = sheet["C" + str(i)].value
                bc = sheet["P" + str(i)].value

                if sheet["Y" + str(i)].value == 0:
                    embed = discord.Embed(title="**[전자제품 상점]**", description="전자제품 목록",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="컴퓨터💻", value="컴퓨터 입니다 50,000 M COIN", inline=False)
                    embed.add_field(name="M DOW <:MDOW:841548472854511616>",value="컴퓨터를 사용하기 위한 운영체체(os) 입니다 1,000 M COIN", inline=False)

                elif sheet["Y" + str(i)].value == 1:
                    embed = discord.Embed(title="**[electronics store]**", description="Electronics List",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="computer 💻", value="it's a computer 50,000 M COIN", inline=False)
                    embed.add_field(name="M DOW <:MDOW:841548472854511616>",value="It is the operating system (OS) for using the computer. 1,000 M COIN", inline=False)
                    
                elif sheet["Y" + str(i)].value == 2:
                    embed = discord.Embed(title="**[電子製品店]**", description="電子製品のリスト",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="コンピュータ 💻", value="コンピュータです 50,000 M COIN", inline=False)
                    embed.add_field(name="M DOW <:MDOW:841548472854511616>",value="コンピュータを使用するための大量のオペレーティングシステム（os）です。 1,000 M COIN", inline=False)
                    
                elif sheet["Y" + str(i)].value == 3:
                    embed = discord.Embed(title="**[电子商店]**", description="电子产品清单",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="计算机 💻", value="这是一台电脑 50,000 M COIN", inline=False)
                    embed.add_field(name="M DOW <:MDOW:841548472854511616>",value="它是用于使用计算机的操作系统 (OS)。 1,000 M COIN", inline=False)
                    
                else:
                    embed = discord.Embed(title="**[전자제품 상점]**", description="전자제품 목록",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.add_field(name="컴퓨터 💻", value="컴퓨터 입니다 50,000 M COIN", inline=False)
                    embed.add_field(name="M DOW <:MDOW:841548472854511616>",value="컴퓨터를 사용하기 위한 운영체체(os) 입니다 1,000 M COIN", inline=False)

        msg = await message.channel.send(embed=embed)
        await msg.add_reaction("💻")
        await msg.add_reaction("<:MDOW:841548472854511616>")
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        def check(reaction, user):
            if user.bot == 1:  # 봇이면 패스
                return None
            return user == message.author and str(reaction.emoji) == '💻' or user == message.author and str(reaction.emoji) == '<:MDOW:841548472854511616>'
        try:
            reaction, user = await client.wait_for('reaction_add', timeout=10000.0, check=check)
        except asyncio.TimeoutError:
            embed = discord.Embed(title="구매취소!", description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
            await msg.delete()
            await message.channel.send(embed=embed)

        else:
            if str(reaction.emoji) == "💻":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        com = sheet["AA" + str(i)].value

                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "컴퓨터를 구매"
                            a_211 = "남은"

                            a_3 = "님 이미 컴퓨터가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy a computer"
                            a_211 = "remaining "

                            a_3 = "already have a computer.\nPlease use `!computer`"
                            a_4 = "does not have enough M COIN. Please use `!get money`"
                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "コンピュータを購入する"
                            a_211 = "残り "

                            a_3 = "様のコンピュータが既にあります。\n`!コンピュータ` を使用してください"
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"
                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买电脑"
                            a_211 = "剩余 "

                            a_3 = "您的计算机已经存在。\n请用 `!计算机`"
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"
                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "컴퓨터를 구매"
                            a_211 = "남은"

                            a_3 = "님 이미 컴퓨터가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        if sum >= 50000 and com != 1:  # 돈o 컴x
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 50000
                            sheet["AA" + str(i)].value = 1
                            embedsnackshop2 = discord.Embed(title="💻 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} 💻 {} [-50,000 M COIN]\n  {}M COIN : ".format(a_2, a_21, a_211) + str(summ))

                        elif sum >= 50000 and com == 1:  # 돈o 컴o
                            await message.channel.send(message.author.mention + (a_3))

                        elif sum <= 50000 and com == 1:  # 돈x 컴o
                            await message.channel.send(message.author.mention + (a_3))

                        else:  # 돈x 컴x
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")


            if str(reaction.emoji) == "<:MDOW:841548472854511616>":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        os = sheet["AB" + str(i)].value


                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "M DOW os를 구매"
                            a_211 = "남은"

                            a_3 = "님 이미 M DOW os가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy M dow os"
                            a_211 = "remaining "

                            a_3 = "already have M DOW os.\nPlease use `!os install` or `!computer`"
                            a_4 = "does not have enough M COIN. Please use `!get money`"
                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "M DOW osを購入"
                            a_211 = "残り "

                            a_3 = "さんはM DOW osがあります。\n`!osのインストール` または `!コンピュータ` 使用してください"
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"
                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "购买 M dow os"
                            a_211 = "剩余 "

                            a_3 = "您已经拥有 M DOW 操作系统。\n请用 `!操作系统安装` 要么 `!计算机`"
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"
                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "M DOW os를 구매"
                            a_211 = "남은"

                            a_3 = "님 이미 M DOW os가 있습니다.\n`!컴퓨터` 를 사용해 주세요"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                        if sum >= 1000 and os != 1:  # 돈o 컴x
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 1000
                            sheet["AB" + str(i)].value = 1
                            embedsnackshop2 = discord.Embed(title="<:MDOW:841548472854511616> {}".format(a_1),description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fff8)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} <:MDOW:841548472854511616> {} [-1,000 M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))

                        elif sum >= 1000 and os == 1:  # 돈o MDOW o
                            await message.channel.send(message.author.mention + (a_3))

                        elif sum <= 1000 and os == 1:  # 돈x MDOW o
                            await message.channel.send(message.author.mention + (a_3))

                        else:  # 돈x MDOWx
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")


    if message.content.startswith('!Food shop') or msg == "!음식 상점" or msg == '!food shop' or msg == '!음식상점' or msg == "!食品店" or msg == "!食料品店":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):

                if sheet["Y" + str(i)].value == 0:

                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="라면 🍜", value="맛있다 [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="아이스크림 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="피자 🍕", value="치즈ㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="햄버거 🍔", value="굿!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="핫도그 🌭", value="맛있다! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="우유 🥛", value="!!우유~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="추가중...", value="추가중...", inline=False)


                elif sheet["Y" + str(i)].value == 1:



                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP List-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="Ramen 🍜", value="Best! [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="Ice cream 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="Pizza 🍕", value="Cheeseㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="Next | Before", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP List-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="Hamburger 🍔", value="GOOD!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="Hotdog 🌭", value="Delicious!! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="Milk 🥛", value="!!milk~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="Next | Before", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP List-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="being added...", value="being added...", inline=False)

                    
                elif sheet["Y" + str(i)].value == 2:

                    
                    
                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食品のリスト-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="ラーメン 🍜", value="おいしい! [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="アイスクリーム 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="ピザ 🍕", value="チーズーー! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="次の | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食品のリスト-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="ハンバーガー 🍔", value="グッド!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="ホットドッグ 🌭", value="美味しい!! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="牛乳 🥛", value="!!牛乳~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="次の | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食品のリスト-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="追加中...", value="追加中...", inline=False)


                elif sheet["Y" + str(i)].value == 3:

                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食物清单-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="拉面 🍜", value="很好！ [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="冰淇淋 🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="比萨 🍕", value="起司ㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="下一个 | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食物清单-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="汉堡包 🍔", value="好的!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="热狗 🌭", value="很好! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="牛奶 🥛", value="!!牛奶~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="下一个 | 前", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 食物清单-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="添加...", value="添加...", inline=False)


                else:
                    #__ 음식상점
                    embedsnackshop = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-1",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embedsnackshop.add_field(name="라면🍜", value="맛있다 [10M COIN]", inline=False)
                    embedsnackshop.add_field(name="아이스크림🍦", value="!!! [5M COIN]", inline=False)
                    embedsnackshop.add_field(name="피자🍕", value="치즈ㅡㅡ! [70M COIN]", inline=False)
                    embedsnackshop.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop2 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-2",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                    embedsnackshop2.add_field(name="햄버거🍔", value="굿!! [40M COIN]", inline=False)
                    embedsnackshop2.add_field(name="핫도그🌭", value="맛있다! [30M COIN]", inline=False)
                    embedsnackshop2.add_field(name="우유🥛", value="!!우유~! [10M COIN]", inline=False)
                    embedsnackshop2.add_field(name="다음 | 이전", value="▶ㅤ|ㅤ◀", inline=False)

                    embedsnackshop3 = discord.Embed(title="FOOD SHOP BOT", description="FOOD SHOP 음식 목록-3",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xcfff83)
                    embedsnackshop3.add_field(name="추가중...", value="추가중...", inline=False)


        msg = await message.channel.send(embed=embedsnackshop)
        await msg.add_reaction("🍜")
        await msg.add_reaction("🍦")
        await msg.add_reaction("🍕")
        await msg.add_reaction("▶")



        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        def check(reaction, user):
            if user.bot == 1:
                return None
            return user == message.author and str(reaction.emoji) == '🍜' or user == message.author and str(
                reaction.emoji) == '🍦' or user == message.author and str(
                reaction.emoji) == '🍕' or user == message.author and str(reaction.emoji) == '▶'

        try:
            reaction, user = await client.wait_for('reaction_add', timeout=30.0, check=check)
        except asyncio.TimeoutError:
            embedsnackshopn = discord.Embed(title="Cancel", description="",
                                            timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                            color=0x83fff8)
            await msg.delete()
            await message.channel.send(embed=embedsnackshopn)

        else:
            if str(reaction.emoji) == "🍜":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        
                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "라면을 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy ramen"
                            a_211 = "remaining "
                            a_4 = "does not have enough M COIN. Please use `!get money`"

                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "ラーメンを購入"
                            a_211 = "残り "
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买拉面"
                            a_211 = "剩余 "
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "라면을 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        if sum >= 10:
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 10
                            embedsnackshop2 = discord.Embed(title="🍜 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(
                                user.name + "{} 🍜 {} [-10M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                            break
                        else:
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")


            elif str(reaction.emoji) == "🍦":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "아이스크림을 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy ice cream"
                            a_211 = "remaining "
                            a_4 = "does not have enough M COIN. Please use `!get money`"

                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "アイスクリームを購入"
                            a_211 = "残り "
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买冰淇淋"
                            a_211 = "剩余 "
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "아이스크림을 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                        if sum >= 5:
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 5
                            embedsnackshop2 = discord.Embed(title="🍦 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} 🍦 {} [-5M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                            break
                        else:
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")


            elif str(reaction.emoji) == "🍕":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "피자를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy a pizza"
                            a_211 = "remaining "
                            a_4 = "does not have enough M COIN. Please use `!get money`"

                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "ピザを購入"
                            a_211 = "残り "
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买一个披萨"
                            a_211 = "剩余 "
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "피자를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                        if sum >= 70:
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 70
                            embedsnackshop2 = discord.Embed(title="🍕 {}".format(a_1), description="",
                                                            timestamp=datetime.datetime.now(
                                                                pytz.timezone('UTC')), color=0x83fff8)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} 🍕 {} [-70M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                            break
                        else:
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")

            if str(reaction.emoji) == "▶":
                await msg.delete()
                msg = await message.channel.send(embed=embedsnackshop2)
                await msg.add_reaction("🍔")
                await msg.add_reaction("🌭")
                await msg.add_reaction("🥛")
                await msg.add_reaction("▶")

        # p2

        def check(reaction, user):
            if user.bot == 1:  # 봇이면 패스
                return None
            return user == message.author and str(reaction.emoji) == '🍔' or user == message.author and str(reaction.emoji) == '🌭' or user == message.author and str(reaction.emoji) == '🥛' or user == message.author and str(reaction.emoji) == '▶'

        try:
            reaction, user = await client.wait_for('reaction_add', timeout=20.0, check=check)
        except asyncio.TimeoutError:
            embedsnackshopn = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83ffb5)
            await msg.delete()
            await message.channel.send(embed=embedsnackshopn)

        else:

            if str(reaction.emoji) == "🍔":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "햄버거를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy a hamburger"
                            a_211 = "remaining "
                            a_4 = "does not have enough M COIN. Please use `!get money`"

                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "ハンバーガーを購入"
                            a_211 = "残り "
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买一个汉堡"
                            a_211 = "剩余 "
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "햄버거를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"

                        if sum >= 40:
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 40
                            embedsnackshop2 = discord.Embed(title="🍔 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} 🍔 {} [-40M COIN]\n {} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                            break
                        else:
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")

            elif str(reaction.emoji) == "🌭":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "핫도그를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy a hot dog"
                            a_211 = "remaining "
                            a_4 = "does not have enough M COIN. Please use `!get money`"

                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "ホットドッグを購入"
                            a_211 = "残り "
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买热狗"
                            a_211 = "剩余 "
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "핫도그를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"
                            
                        if sum >= 30:
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 30
                            embedsnackshop2 = discord.Embed(title="🌭 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} 🌭 {} [-30M COIN]\n{} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                            break
                        else:
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")

            elif str(reaction.emoji) == "🥛":
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        sum = sheet["C" + str(i)].value
                        if sheet["Y" + str(i)].value == 0:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "우유를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"


                        elif sheet["Y" + str(i)].value == 1:
                            a_1 = "Purchase complete!"
                            a_2 = """'s"""
                            a_21 = "buy milk"
                            a_211 = "remaining "
                            a_4 = "does not have enough M COIN. Please use `!get money`"

                            
                        elif sheet["Y" + str(i)].value == 2:
                            a_1 = "購入完了！"
                            a_2 = "様が"
                            a_21 = "牛乳を購入"
                            a_211 = "残り "
                            a_4 = "様M COINが不足します。 `!お金を受け取る` を使用してください"

                            
                        elif sheet["Y" + str(i)].value == 3:
                            a_1 = "购买完成！"
                            a_2 = "尼姆"
                            a_21 = "买牛奶"
                            a_211 = "剩余 "
                            a_4 = "您没有足够的 M COIN。请用 `!拿钱`"

                            
                        else:
                            a_1 = "구매완료!"
                            a_2 = "님이"
                            a_21 = "우유를 구매"
                            a_211 = "남은"
                            a_4 = "님 M COIN 이 부족합니다. `!돈받기` 를 사용하세요!"
                            
                        if sum >= 30:
                            summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 10
                            embedsnackshop2 = discord.Embed(title="🥛 {}".format(a_1), description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83ffb5)
                            await msg.delete()
                            await message.channel.send(embed=embedsnackshop2)
                            await reaction.message.channel.send(user.name + "{} 🥛 {} [-10M COIN]\n{} M COIN : ".format(a_2, a_21, a_211) + str(summ))
                            break
                        else:
                            await msg.delete()
                            await reaction.message.channel.send(message.author.mention + "{}\n{} M COIN : ".format(a_4, a_211) + str(sum))
                file.save("user.xlsx")

            elif str(reaction.emoji) == "▶":
                await msg.delete()
                msg = await message.channel.send(embed=embedsnackshop3)
                await msg.add_reaction("⛔")

        # p3

        def check(reaction, user):
            if user.bot == 1:  # 봇이면 패스
                return None
            return user == message.author and str(reaction.emoji) == '⛔'

        try:
            reaction, user = await client.wait_for('reaction_add', timeout=10.0, check=check)
        except asyncio.TimeoutError:
            embedsnackshopn = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xcfff83)
            await msg.delete()
            await message.channel.send(embed=embedsnackshopn)

        else:
            if str(reaction.emoji) == "⛔":
                embedsnackshopn = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xcfff83)
                await msg.delete()
                await message.channel.send(embed=embedsnackshopn)
    

# 컴퓨터

    if message.content.startswith("!컴퓨터") or message.content.startswith("!컴") or msg == "!computer" or msg == "!コンピュータ" or msg == "!计算机" or msg == "!电脑":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):  # 유저아이디 확인
                if sheet["Y" + str(i)].value == 0:
                    main = "컴퓨터"
                    val = "님의 컴퓨터"
                    sta = "상태"
                    os = "운영체제"
                    cpu = "프로세서"
                    ram = "메모리"
                    
                    none = "님 컴퓨터를 소지하고 계시지 않습니다"

                elif sheet["Y" + str(i)].value == 1:
                    main = "Computer"
                    val = "'s Computer"
                    sta = "State"
                    os = "Operating System"
                    cpu = "Processor"
                    ram = "Memory"
                    
                    none = "doesn't have a computer `!e-shop`"

                elif sheet["Y" + str(i)].value == 2:
                    main = "コンピュータ"
                    val = "さんのコンピュータ"
                    sta = "状態"
                    os = "オペレーティングシステム"
                    cpu = "プロセッサ"
                    ram = "メモリ"
                    
                    none = "様のコンピュータを所持しておられません `!電子商店`"

                elif sheet["Y" + str(i)].value == 3:
                    main = "电脑"
                    val = "的电脑"
                    sta = "状态"
                    os = "操作系统"
                    cpu = "处理器"
                    ram = "记忆"
                    
                    none = "没有电脑 `!电子商店`"

                else:
                    main = "컴퓨터"
                    val = "님의 컴퓨터"
                    sta = "상태"
                    os = "운영체제"
                    cpu = "프로세서"
                    ram = "메모리"

                    none = "님 컴퓨터를 소지하고 계시지 않습니다"
                    

                if sheet["AA" + str(i)].value == 1:  # 컴퓨터 보유여부 확인
                    if sheet["AC" + str(i)].value == 1:  # os 설치여부 확인
                        if sheet["AD" + str(i)].value == 1: #온라인 확인
                            embed = discord.Embed(title="**{}**".format(main), description="{}{}\n\n🎈{} : 🟢\n⚙{} : M DOW\n<:cpu:851445106649333762>{} : MD M7900S\n<:ram:851445251080060950>{} : 16GB".format(message.author.name, val, sta, os, cpu, ram),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                            embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)

                            msg = await message.channel.send(embed=embed)

                        else:
                            embed = discord.Embed(title="**{}**".format(main), description="{}{}\n\n🎈{} : 🔴\n⚙{} : M DOW\n<:cpu:851445106649333762>{} : MD M7900S\n<:ram:851445251080060950>{} : 16GB".format(message.author.name, val, sta, os, cpu, ram),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                            embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                            msg = await message.channel.send(embed=embed)

                    else:
                        embed = discord.Embed(title="**{}**".format(main), description="{}{}\n\n🎈{} : 🔴\n⚙{} : Not install\n<:cpu:851445106649333762>{} : MD M7900S\n<:ram:851445251080060950>{} : 16GB".format(message.author.name, val, sta, os, cpu, ram),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                        embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                        msg = await message.channel.send(embed=embed)

                else:
                    await message.channel.send(message.author.mention+none)


    if msg == "!부팅" or msg == "!boot" or msg == "!起動" or msg == "!开机":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui): #유저아이디 확인

                if sheet["Y" + str(i)].value == 0:
                    main = "컴퓨터"
                    val = "부팅중..."
                    sta = "전원이 좋료되었습니다."

                    al = "이미 켜져있습니다. `!화면`"
                    nos = "os가 설치되어 있지 않습니다 `!os설치`"
                    
                    none = "님 컴퓨터를 소지하고 계시지 않습니다 `!전자상점`"

                elif sheet["Y" + str(i)].value == 1:
                    main = "Computer"
                    val = "booting up..."
                    sta = "Power has been turned off."

                    al = "It's already on. `!screen`"
                    nos = "os is not installed `!os install`"
                    
                    none = "doesn't have a computer `!e-shop`"

                elif sheet["Y" + str(i)].value == 2:
                    main = "コンピュータ"
                    val = "起動中..."
                    sta = "電源が終了しました。"

                    al = "すでにオンになっています。`!画面`"
                    nos = "osがインストールされていません `!osのインストール`"
                    
                    none = "様のコンピュータを所持しておられません `!電子商店`"

                elif sheet["Y" + str(i)].value == 3:
                    main = "电脑"
                    val = "正在启动..."
                    sta = "电源已关闭。"

                    al = "它已经开始了。`!屏幕`"
                    nos = "没有安装os `!操作系统安装`"
                    
                    none = "没有电脑 `!电子商店`"

                else:
                    main = "컴퓨터"
                    val = "부팅중..."
                    sta = "전원이 종료되었습니다."

                    al = "이미 켜져있습니다. `!화면`"
                    nos = "os가 설치되어 있지 않습니다 `!os설치`"

                    none = "님 컴퓨터를 소지하고 계시지 않습니다 `!전자상점`"
                    

                if sheet["AA" + str(i)].value == 1: #컴퓨터 보유 확인
                    if sheet["AC" + str(i)].value == 1: #os 설치 여부 확인
                        if sheet["AD" + str(i)].value == 0:  #  켜짐 여부 확인
                            embed = discord.Embed(title="**{}**".format(main), description="{}".format(val),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                            embed.set_image(url="https://cdn.discordapp.com/attachments/833181581526433795/851618823144603659/loading.gif")
                            embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)
                            msg = await message.channel.send(embed=embed)

                            sleep(3)
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):  # 유저아이디 확인
                                    sheet["AD" + str(i)].value = 1
                                    file.save("user.xlsx")
                                    break

                            embed= discord.Embed(title="**{}**".format(main),description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xa664ff)
                            embed.set_image(url="https://media.discordapp.net/attachments/833181581526433795/851621108440760320/bde59507170a5ef6.png?width=1067&height=600")
                            embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)


                            await msg.edit(content=message.author.mention, embed=embed)
                            await msg.add_reaction("<:Chrome:851614010431897640>")
                            await msg.add_reaction("<:off:841555156935639080>")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '<:off:841555156935639080>' or user == message.author and str(reaction.emoji) == '<:Chrome:851614010431897640>'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=99999.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                            else:
                                if str(reaction.emoji) == "<:off:841555156935639080>":  #종료
                                    await msg.delete()
                                    await message.channel.send("```{}```".format(sta))
                                    sheet["AD" + str(i)].value = 0

                                if str(reaction.emoji) == "<:Chrome:851614010431897640>":  #크롬
                                    await message.channel.send("Developing")


                        else:
                            await message.channel.send(message.author.mention+al)

                    else:
                        await message.channel.send(message.author.mention + nos)

                else:
                    await message.channel.send(message.author.mention+none)

        file.save("user.xlsx")


    if message.content.startswith("!os설치") or message.content.startswith("!Os설치") or message.content.startswith("!oS설치") or message.content.startswith("!OS설치") or message.content.startswith("!os 설치") or message.content.startswith("!Os 설치") or message.content.startswith("!oS 설치") or message.content.startswith("!OS 설치") or msg == "!os install" or msg == "!osのインストール" or msg == "!操作系统安装":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui): #유저아이디 확인

                if sheet["Y" + str(i)].value == 0:
                    complete = "님 os설치가 끝났습니다 `!컴퓨터`"
                    already = "님 이미 os가 설치되었습니다 `!컴퓨터`"
                    none = "님 os가 없습니다 `!전자상점`"

                elif sheet["Y" + str(i)].value == 1:
                    complete = "'s os installation is complete `!computer`"
                    already = "os already installed `!computer`"
                    none = "no have os `!e-shop`"


                elif sheet["Y" + str(i)].value == 2:
                    complete = "さんosインストールが完了しました `!コンピュータ`"
                    already = "さんはosがインストールされてい `!计算机`"
                    none = "様osはありません `!電子商店`"
        

                elif sheet["Y" + str(i)].value == 3:
                    complete = "的os安装完成 `!计算机`"
                    already = "操作系统已经安装 `!计算机`"
                    none = "没有操作系统 `!电子商店`"
        

                else:
                    complete = "님 os설치가 끝났습니다 `!컴퓨터`"
                    already = "님 이미 os가 설치되었습니다 `!컴퓨터`"
                    none = "님 os가 없습니다 `!전자상점`"
        


                if sheet["AB" + str(i)].value == 1: #os보유 확인
                    if sheet["AC" + str(i)].value == 0: #os 설치확인

                        embed = discord.Embed(title="**[M DOW Install]**", description="Click ⭕ or ❌",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x7289da)
                        embed.set_image(url="https://images-ext-1.discordapp.net/external/FCpuyyygXvOkJ8icKeaMp77r0MAGX8q9QizhIB2oyPA/%3Fwidth%3D1070%26height%3D602/https/media.discordapp.net/attachments/832849028873191516/841985028497866752/27660551273523e1.png")
                        msg = await message.channel.send(embed=embed)
                        await msg.add_reaction("⭕")
                        await msg.add_reaction("❌")
                        def check(reaction, user):
                            if user.bot == 1:  # 봇이면 패스
                                return None
                            return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(
                                reaction.emoji) == '❌'
    
                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=30.0, check=check)
                        except asyncio.TimeoutError:
                            embed = discord.Embed(title="Cancel!", description="",
                                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

                        else:
                            if str(reaction.emoji) == "❌":
                                embed = discord.Embed(title="Cancel!", description="",
                                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                                await msg.delete()
                                await message.channel.send(embed=embed)
                
                            elif str(reaction.emoji) == "⭕": #설치
                                await msg.delete()
                                embed = discord.Embed(title="**[M DOW Install]**", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x7289da)
                                embed.set_image(url="https://cdn.discordapp.com/attachments/859954531692576778/869423086217334824/d23f4fb4d8eca6a7.gif")
                                msg = await message.channel.send(embed=embed)
                                sheet["AC" + str(i)].value = 1
                                sleep(18)
                                await msg.delete()
                                await message.channel.send("{}{}".format(user.name, complete))
                    
                    else:
                        await message.channel.send(message.author.mention + already)
                        break

                else:
                    await message.channel.send(message.author.mention + none)
                    break
        file.save("user.xlsx")
    

    if msg == "!화면":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui): #유저아이디 확인

                if sheet["Y" + str(i)].value == 0:
                    main = "컴퓨터"
                    val = "부팅중..."
                    sta = "전원이 좋료되었습니다."

                    al = "전원이 종료되어있습니다. `!부팅`"
                    nos = "os가 설치되어 있지 않습니다 `!os설치`"
                    
                    none = "님 컴퓨터를 소지하고 계시지 않습니다 `!전자상점`"

                elif sheet["Y" + str(i)].value == 1:
                    main = "Computer"
                    val = "booting up..."
                    sta = "Power has been turned off."

                    al = "Power is off. `!boot`"
                    nos = "os is not installed `!os install`"
                    
                    none = "doesn't have a computer `!e-shop`"

                elif sheet["Y" + str(i)].value == 2:
                    main = "コンピュータ"
                    val = "起動中..."
                    sta = "電源が終了しました。"

                    al = "電源が終了しています。 `!起動`"
                    nos = "osがインストールされていません `!osのインストール`"
                    
                    none = "様のコンピュータを所持しておられません `!電子商店`"

                elif sheet["Y" + str(i)].value == 3:
                    main = "电脑"
                    val = "正在启动..."
                    sta = "电源已关闭。"

                    al = "电源关闭。 `!开机`"
                    nos = "没有安装os `!操作系统安装`"
                    
                    none = "没有电脑 `!电子商店`"

                else:
                    main = "컴퓨터"
                    val = "부팅중..."
                    sta = "전원이 종료되었습니다."

                    al = "전원이 종료되어있습니다. `!부팅`"
                    nos = "os가 설치되어 있지 않습니다 `!os설치`"

                    none = "님 컴퓨터를 소지하고 계시지 않습니다 `!전자상점`"

                if sheet["AA" + str(i)].value == 1: #컴퓨터 보유 확인
                    if sheet["AC" + str(i)].value == 1: #os 설치 여부 확인
                        if sheet["AD" + str(i)].value == 1:  #  켜짐 여부 확인

                            embed= discord.Embed(title="**Computer**",description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xa664ff)
                            embed.set_image(url="https://media.discordapp.net/attachments/833181581526433795/851621108440760320/bde59507170a5ef6.png?width=1067&height=600")
                            embed.set_footer(text="{}".format(message.author), icon_url=message.author.avatar_url)


                            msg = await message.channel.send(message.author.mention, embed=embed)
                            await msg.add_reaction("<:Chrome:851614010431897640>")
                            await msg.add_reaction("<:off:841555156935639080>")

                            def check(reaction, user):
                                if user.bot == 1:  # 봇이면 패스
                                    return None
                                return user == message.author and str(reaction.emoji) == '<:off:841555156935639080>' or user == message.author and str(reaction.emoji) == '<:Chrome:851614010431897640>'

                            try:
                                reaction, user = await client.wait_for('reaction_add', timeout=99999.0, check=check)
                            except asyncio.TimeoutError:
                                await msg.delete()
                            else:
                                if str(reaction.emoji) == "<:off:841555156935639080>":  #종료
                                    await msg.delete()
                                    await message.channel.send("```{}```".format(sta))
                                    sheet["AD" + str(i)].value = 0

                                if str(reaction.emoji) == "<:Chrome:851614010431897640>":  #크롬
                                    await message.channel.send("developing")


                        else:
                            await message.channel.send(message.author.mention + al)

                    else:
                        await message.channel.send(message.author.mention + nos)

                else:
                    await message.channel.send(message.author.mention + none)

        file.save("user.xlsx")
        

# 공지-----------------------

    if message.content.startswith("!공지") and not msg == "!공지":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content[4:]
            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**공지사항**",description="공지사항 내용은 항상 숙지 해주시기 바랍니다\n――――――――――――――――――――――――――――\n\n{}\n\n――――――――――――――――――――――――――――\n확인후 아래 체크표시를 눌러주세요!".format(notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="공지 발신자 : {}".format(message.author), icon_url=message.author.avatar_url)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("||@everyone||", embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            embed1 = discord.Embed(title="**[ M BOT 공지 알림 ]**\n정상적으로 공지가 채널에 작성이완료되었습니다 : )",description="\n\n[ 작성 서버 ] : {}\n[ 서버 ID ] : {}\n[ 작성 채널 ] : {}\n[ 채널 ID ] : {}\n[ 공지 발신자 ] : {}\n[ 발신자 ID ] : {}\n[ 작성 시간 ] : {}\n\n[ 내용 ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | 담당 관리자 : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)

        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : 권한부족", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


    if message.content.startswith("!notice") and not msg == "!notice":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content[4:]
            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**Notice**",description="Please be aware of the notices at all times.\n――――――――――――――――――――――――――――\n\n{}\n\n――――――――――――――――――――――――――――\nAfter confirming, please click the checkmark below!".format(notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="notice sender : {}".format(message.author), icon_url=message.author.avatar_url)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("||@everyone||", embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            embed1 = discord.Embed(title="**[ M BOT notice notice ]**\nNotifications have been successfully written to the channel : )",description="\n\n[ compose server ] : {}\n[ Server ID ] : {}\n[ writing channel ] : {}\n[ Channel ID ] : {}\n[ notice sender ] : {}\n[ Caller ID ] : {}\n[ writing time ] : {}\n\n[ Contents ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | responsible manager : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)

        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : lack of authority", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


    if message.content.startswith("!お知らせ") and not msg == "!お知らせ":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content[6:]
            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**お知らせ**",description="お知らせの内容は、常に熟知してください\n――――――――――――――――――――――――――――\n\n{}\n\n――――――――――――――――――――――――――――\n確認後、以下のチェックマークを押してください！".format(notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="お知らせ発信者 : {}".format(message.author), icon_url=message.author.avatar_url)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("||@everyone||", embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            embed1 = discord.Embed(title="**[ M BOT お知らせ 通知 ]**\n通常公知のチャンネルに作成が完了しました : )",description="\n\n[ 作成サーバー ] : {}\n[ サーバID ] : {}\n[ 作成チャンネル ] : {}\n[ チャネルID ] : {}\n[ お知らせ発信者 ] : {}\n[ 発信者ID ] : {}\n[ 作成時間 ] : {}\n\n[ 内容 ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | 担当者 : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)

        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="エラー : 権限不足", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


    if message.content.startswith("!注意") and not msg == "!注意":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content[4:]
            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**注意**",description="请随时注意通知。\n――――――――――――――――――――――――――――\n\n{}\n\n――――――――――――――――――――――――――――\n确认无误后，请点击下方的勾号！".format(notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="通知发件人 : {}".format(message.author), icon_url=message.author.avatar_url)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("||@everyone||", embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            embed1 = discord.Embed(title="**[ M BOT 通知通知 ]**\n通知已成功写入频道 : )",description="\n\n[ 组合服务器 ] : {}\n[ 服务器 ID ] : {}\n[ 写作频道 ] : {}\n[ 渠道 ID ] : {}\n[ 通知发件人 ] : {}\n[ 来电显示 ] : {}\n[ 写作时间 ] : {}\n\n[ 内容 ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | 负责人 : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)

        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : 缺乏权威", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


# 투표 

    if message.content.startswith("!투표"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            await message.channel.purge(limit=1)
            i = (message.author.guild_permissions.administrator)
            mi = 789670002163974145
            if i is True or message.author.id == mi:
                notice = message.content.split("/")
                #notice = message.content[4:]
                print(notice)
                no_1 = notice[0]
                no_1 = str(no_1[4:])
                no_2 = notice[1]

                c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
                rc = str(random.sample(c, 1))
                rct = rc[1:8]
                rct = int(rct)
                embed = discord.Embed(title="**투표**",description="투표내용\n――――――――――――――――――――――――――――\n\n⭕ {}\n❌ {}\n\n――――――――――――――――――――――――――――\n확인후 아래에 투표해 주세요!".format(no_1,no_2), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed.set_footer(text="담당 관리자 : {}".format(message.author), icon_url=message.author.avatar_url)
                #embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
                msg = await message.channel.send("@here", embed=embed)
                await msg.add_reaction("⭕")
                await msg.add_reaction("❌")
                embed1 = discord.Embed(title="**[ M BOT 투표 알림 ]**\n정상적으로 투표가 시작되었습니다 : )",description="\n\n[ 작성 서버 ] : {}\n[ 서버 ID ] : {}\n[ 작성 채널 ] : {}\n[ 채널 ID ] : {}\n[ 공지 발신자 ] : {}\n[ 발신자 ID ] : {}\n[ 작성 시간 ] : {}\n\n[ 내용 ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
                embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | 담당 관리자 : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
                await message.author.send(embed=embed1)
        except:
            await message.channel.send("사용법 : !투표 <내용1>/<내용2>")


        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : 권한부족", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


    if message.content.startswith("!vote"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content.split("/")
            #notice = message.content[4:]
            print(notice)
            no_1 = notice[0]
            no_1 = str(no_1[6:])
            no_2 = notice[1]

            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**Vote**",description="Voting contents\n――――――――――――――――――――――――――――\n\n⭕ {}\n❌ {}\n\n――――――――――――――――――――――――――――\nPlease check and vote below!".format(no_1,no_2), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="Voter : {}".format(message.author), icon_url=message.author.avatar_url)
            #embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("@here", embed=embed)
            await msg.add_reaction("⭕")
            await msg.add_reaction("❌")
            embed1 = discord.Embed(title="**[ M BOT Voting Notification ]**\nVoting has started normally : )",description="\n\n[ compose server ] : {}\n[ Server ID ] : {}\n[ writing channel ] : {}\n[ Channel ID ] : {}\n[ notice sender ] : {}\n[ Caller ID ] : {}\n[ writing time ] : {}\n\n[ Contents ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | responsible manager : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)


        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : lack of authority", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


    if message.content.startswith("!投票"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content.split("/")
            #notice = message.content[4:]
            print(notice)
            no_1 = notice[0]
            no_1 = str(no_1[4:])
            no_2 = notice[1]

            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**投票**",description="投票内容\n――――――――――――――――――――――――――――\n\n⭕ {}\n❌ {}\n\n――――――――――――――――――――――――――――\n確認後、下に投票してください！".format(no_1,no_2), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="投票発信者: {}".format(message.author), icon_url=message.author.avatar_url)
            #embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("@here", embed=embed)
            await msg.add_reaction("⭕")
            await msg.add_reaction("❌")
            embed1 = discord.Embed(title="**[ M BOT 投票の通知 ]**\n通常の投票が開始されました : )",description="\n\n[ 作成サーバー ] : {}\n[ サーバID ] : {}\n[ 作成チャンネル ] : {}\n[ チャネルID ] : {}\n[ お知らせ発信者 ] : {}\n[ 発信者ID ] : {}\n[ 作成時間 ] : {}\n\n[ 内容 ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | 担当者 : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)


        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : 権限不足", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


    if message.content.startswith("!投票"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        await message.channel.purge(limit=1)
        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        if i is True or message.author.id == mi:
            notice = message.content.split("/")
            #notice = message.content[4:]
            print(notice)
            no_1 = notice[0]
            no_1 = str(no_1[4:])
            no_2 = notice[1]

            c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
            rc = str(random.sample(c, 1))
            rct = rc[1:8]
            rct = int(rct)
            embed = discord.Embed(title="**投票**",description="投票内容\n――――――――――――――――――――――――――――\n\n⭕ {}\n❌ {}\n\n――――――――――――――――――――――――――――\n请在下方查看并投票！".format(no_1,no_2), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed.set_footer(text="담당 관리자 : {}".format(message.author), icon_url=message.author.avatar_url)
            #embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/832849028873191516/848044163802595339/172a6d98e7cca03e.png")
            msg = await message.channel.send("@here", embed=embed)
            await msg.add_reaction("⭕")
            await msg.add_reaction("❌")
            embed1 = discord.Embed(title="**[ M BOT 投票通知 ]**\n投票已正常开始 : )",description="\n\n[ 组合服务器 ] : {}\n[ 服务器 ID ] : {}\n[ 写作频道 ] : {}\n[ 渠道 ID ] : {}\n[ 通知发件人 ] : {}\n[ 来电显示 ] : {}\n[ 写作时间 ] : {}\n\n[ 内容 ]\n{}".format(message.guild,message.guild.id,message.channel,message.channel.id,message.author,message.author.id,datetime.datetime.today(),notice), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=rct)
            embed1.set_footer(text="Bot Made by. 마인잡지 #0001 | 负责人 : 마인잡지 #0001",icon_url="https://cdn.discordapp.com/avatars/789670002163974145/42158b206a17adc8ca016d97ffd8de21.webp?size=1024")
            await message.author.send(embed=embed1)


        if i is False:
            embed = discord.Embed(title="Error : Insufficient authority", description="Error : 缺乏权威", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.author.send(embed=embed)


# 청소

    if message.content.startswith("!청소") or message.content.startswith("!삭제") and not msg == "!청소" and not msg == "!삭제":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        amount = message.content[4:]
        if int(amount) <= 1000:
            if i is True or message.author.id == mi:
    
                await message.channel.purge(limit=1)
                await message.channel.purge(limit=int(amount))
                c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
                rc = str(random.sample(c, 1))
                rct = rc[1:8]
                rct = int(rct)
                embed = discord.Embed(title="**메시지 삭제 알림**",description="메세지 {}개 삭제 완료!".format(amount), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=rct)
                await message.author.send(embed=embed)

    
            if i is False:
                embed = discord.Embed(title="Error : Insufficient authority", description="Error : 권한부족", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                await message.author.send(embed=embed)
        else:
            embed = discord.Embed(title="Error", description="메세지는 1회 최대 1000개만 삭제할수 있습니다",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.channel.send(embed=embed)


    if message.content.startswith("!delete") and not msg == "!delete":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        amount = message.content[8:]
        if int(amount) <= 1000:
            if i is True or message.author.id == mi:
    
                await message.channel.purge(limit=1)
                await message.channel.purge(limit=int(amount))
                c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
                rc = str(random.sample(c, 1))
                rct = rc[1:8]
                rct = int(rct)
                embed = discord.Embed(title="**Message deletion notification**",description="message {}deleted!".format(amount), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=rct)
                await message.author.send(embed=embed)

    
            if i is False:
                embed = discord.Embed(title="Error : Insufficient authority", description="Error : lack of authority", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                await message.author.send(embed=embed)
        else:
            embed = discord.Embed(title="Error", description="You can only delete up to 1000 messages at a time.",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.channel.send(embed=embed)


    if message.content.startswith("!削除") and not msg == "!削除":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        amount = message.content[4:]
        if int(amount) <= 1000:
            if i is True or message.author.id == mi:
    
                await message.channel.purge(limit=1)
                await message.channel.purge(limit=int(amount))
                c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
                rc = str(random.sample(c, 1))
                rct = rc[1:8]
                rct = int(rct)
                embed = discord.Embed(title="**メッセージの削除通知**",description="メッセージ {}個、削除完了！".format(amount), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=rct)
                await message.author.send(embed=embed)

    
            if i is False:
                embed = discord.Embed(title="Error : Insufficient authority", description="Error : 権限不足", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                await message.author.send(embed=embed)
        else:
            embed = discord.Embed(title="Error", description="メッセージは、1回の最大1000個だけ削除することができます。",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.channel.send(embed=embed)


    if message.content.startswith("!删除") and not msg == "!删除":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        i = (message.author.guild_permissions.administrator)
        mi = 789670002163974145
        amount = message.content[4:]
        if int(amount) <= 1000:
            if i is True or message.author.id == mi:
    
                await message.channel.purge(limit=1)
                await message.channel.purge(limit=int(amount))
                c = (0xffd630, 0xbfff63, 0x63ff72, 0x63ffd1, 0x63b0ff, 0x6e63ff, 0x9363ff, 0xeb63ff, 0xff63b7, 0xff6363,0xff9663)
                rc = str(random.sample(c, 1))
                rct = rc[1:8]
                rct = int(rct)
                embed = discord.Embed(title="**消息删除通知**",description="删除了 {} 条消息！".format(amount), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=rct)
                await message.author.send(embed=embed)

    
            if i is False:
                embed = discord.Embed(title="Error : Insufficient authority", description="Error : 缺乏权威", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                await message.author.send(embed=embed)
        else:
            embed = discord.Embed(title="Error", description="您一次最多只能删除 1000 条消息。",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
            await message.channel.send(embed=embed)


# 게임 & 은행

    if message.content.startswith("!뽑기") or msg == "draw" or msg == "抜く" or msg == "画":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    a = "뽑기"
                    aa = "뽑기를 하시겠습니까?"
                    aaa = "1회당 300 M COIN"
                    b = "뽑기를 하시려면"
                    bb = "취소하시려면"
                    
                    none_a = "님 소지하고 계신 M COIN 이 부족합니다"

                    zero = "꽝!\n다음기회에~"
                    p3 = "3%확율을 뚫다니 축하드려요!!🎉"
                    p2 = "2%확율을 뚫다니 축하드려요!!🎉"
                    p1 = "1%확율을 뚫다니 축하드려요!!🎉"

    
                elif sheet["Y" + str(i)].value == 1:
                    a = "draw"
                    aa = "Do you want to draw?"
                    aaa = "300 M COIN per one time"
                    b = "뽑기를 하시려면"
                    bb = "To cancel"
                    
                    none_a = "has insufficient M COIN"

                    zero = "bomb"
                    p3 = "Congrats on breaking the 3% chance!!🎉"
                    p2 = "Congrats on breaking the 2% chance!!🎉"
                    p1 = "Congrats on breaking the 1% chance!!🎉"
    
                elif sheet["Y" + str(i)].value == 2:
                    a = "抜く"
                    aa = "抜きをしますか？"
                    aaa = "1回あたり300 M COIN"
                    b = "뽑기를 하시려면"
                    bb = "キャンセルする場合は、"

                    none_a = "様所持しているM COINが不足して"

                    zero = "ブーム"
                    p3 = "3％の確率を開けなんておめでとうございます!!🎉"
                    p2 = "2％の確率を開けなんておめでとうございます!!🎉"
                    p1 = "1％の確率を開けなんておめでとうございます!!🎉"
    
                elif sheet["Y" + str(i)].value == 3:
                    a = "画"
                    aa = "你想画画吗？"
                    aaa = "1회당 300 M COIN"
                    b = "每次 300 M COIN"
                    bb = "取消"

                    none_a = "你没有足够的 M COIN"

                    zero = "砰"
                    p3 = "恭喜你打破了 3% 的机会！！🎉"
                    p2 = "恭喜你打破了 2% 的机会！！🎉"
                    p1 = "恭喜你打破了 1% 的机会！！🎉"

                else:
                    a = "뽑기"
                    aa = "뽑기를 하시겠습니까?"
                    aaa = "1회당 300 M COIN"
                    b = "뽑기를 하시려면"
                    bb = "취소하시려면"

                    none_a = "님 소지하고 계신 M COIN 이 부족합니다"

                    zero = "꽝!\n다음기회에~"
                    p3 = "3%확율을 뚫다니 축하드려요!!🎉"
                    p2 = "2%확율을 뚫다니 축하드려요!!🎉"
                    p1 = "1%확율을 뚫다니 축하드려요!!🎉"

                if sheet["C" + str(i)].value >= 150:
                    embed = discord.Embed(title=a, description="{}\n{}".format(aa,aaa),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x66c072)
                    embed.add_field(name="{} ┃ ".format(b), value="⭕", inline=True)
                    embed.add_field(name="{} ┃ ".format(bb), value="❌", inline=True)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("⭕")
                    await msg.add_reaction("❌")

                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(reaction.emoji) == '❌'

                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=30.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="Cancel!", description="",
                                              timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)


                    else:
                        if str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

                        elif str(reaction.emoji) == "⭕":
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - 300
                                    r = (1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,5,5,5,5,5,5,5,5,5,5,6,6,6,6,6,7,7,7,8,8,9)
                                    rr = str(random.sample(r, 1))
                                    rr = rr[1]
                                    rr = int(rr)
                                    if rr == 1:
                                        embed = discord.Embed(title=a, description="```{}```".format(zero),timestamp=datetime.datetime.now(pytz.timezone('UTC')))
                                        await message.channel.send(message.author.mention,embed=embed)
                                        break


                                    if rr == 2:
                                        embed = discord.Embed(title=a, description="+100 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xff8383)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 100
                                        break


                                    elif rr == 3:
                                        embed = discord.Embed(title=a, description="+200 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xffb583)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 200
                                        break


                                    elif rr == 4:
                                        embed = discord.Embed(title=a, description="+300 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xffe983)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 300
                                        break


                                    elif rr == 5:
                                        embed = discord.Embed(title=a, description="+500 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xbdff83)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 500
                                        break


                                    elif rr == 6:
                                        embed = discord.Embed(title=a, description="+600 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83ffa3)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 600
                                        break


                                    elif rr == 7:
                                        embed = discord.Embed(title=a, description="+700 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83fffb)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        await message.channel.send(p3)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 700
                                        break


                                    elif rr == 8:
                                        embed = discord.Embed(title=a, description="+800 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x83c6ff)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        await message.channel.send(p2)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 800
                                        break


                                    elif rr == 9:
                                        embed = discord.Embed(title=a, description="+1000 M COIN",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xc083ff)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        await message.channel.send(p1)
                                        for i in range(1, 1001):
                                            if sheet["B" + str(i)].value == str(ui):
                                                sum = sheet["C" + str(i)].value
                                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1000
                                        break


                            file.save("user.xlsx")

                else:
                    await message.channel.send(message.author.mention+none_a)


    if message.content.startswith("!도박") or message.content.startswith("!배팅") or message.content.startswith("!赌博") and not msg == "赌博" and not msg == "!도박" and not msg == "!도박 " and not msg == "!배팅" and not msg == "!배팅 ":
        try:
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

            amount = message.content[4:]
            amount = int(amount)
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    if sheet["Y" + str(i)].value == 0:
                        w = "도박은 최소 100 M COIN 부터 가능합니다."
                        gamble = "도박"

                        nextg = "다음기회에~"
                        con = "축하드려요"
                        
                        n = "님 소지하고 계신 M COIN 이 부족합니다"
                        
                    elif sheet["Y" + str(i)].value == 1:
                        w = "Gambling is possible from a minimum of 100 M COIN"
                        gamble = "gamble"

                        nextg = "Next time~"
                        con = "congratulations"
                        
                        n = "You do not have enough M COIN"
                        
                    elif sheet["Y" + str(i)].value == 2:
                        w = "賭博は少なくとも100 M COINから可能です"
                        gamble = "ギャンブル"

                        nextg = "次の機会に〜"
                        con = "おめでとうございます"
                        
                        n = "様所持しているM COINが不足して"
                        
                    elif sheet["Y" + str(i)].value == 3:
                        w = "至少 100 M COIN 即可进行赌博"
                        gamble = "赌博"

                        nextg = "下次吧~"
                        con = "祝贺"
                        
                        n = "你没有足够的 M COIN"
                        
                    else:
                        w = "도박은 최소 100 M COIN 부터 가능합니다."
                        gamble = "도박"

                        nextg = "다음기회에~"
                        con = "축하드려요"

                        n = "님 소지하고 계신 M COIN 이 부족합니다"
                        


                    if amount < 100:
                        await message.channel.send(message.author.mention + w)
                    else:
                        if sheet["C" + str(i)].value >= amount:
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - amount
                                    sheet["N" + str(i)].value = sheet["N" + str(i)].value + 1
                                    r = (1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,5,5,5,5,5,5,5,5,5,6,6,6,6,6,7)
                                    rr = str(random.sample(r, 1))
                                    rr = rr[1]
                                    rr = int(rr)
                                    
                                    if sheet["N" + str(i)].value == 100:
                                        embed = discord.Embed(title="**《업적달성》**",description="『도박중독』 - 도박 100번 하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『도박중독』 "
                                    if rr == 1:
                                        embed = discord.Embed(title=gamble, description="```0.25x\n{}```".format(nextg),timesamp=datetime.datetime.now(pytz.timezone('UTC')))
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 0.25)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 2:
                                        embed = discord.Embed(title=gamble,description="```0.5x\n{}```".format(nextg),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xa0a0a0)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 0.5)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 3:
                                        embed = discord.Embed(title=gamble, description="```1x...```",timesamp=datetime.datetime.now(pytz.timezone('UTC')))
                                        await message.channel.send(message.author.mention, embed=embed)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + amount
                                        break
                                    if rr == 4:
                                        embed = discord.Embed(title=gamble, description="```2x! \n{}!```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 2)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 5:
                                        embed = discord.Embed(title=gamble, description="```3x! {}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff2)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 3)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 6:
                                        embed = discord.Embed(title=gamble, description="```5x! {}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 5)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 7:
                                        embed = discord.Embed(title=gamble,description="```10x! 🎉🎉!!!~{}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 10)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break

                            file.save("user.xlsx")
                        else:
                            await message.channel.send(message.author.mention + " {}".format(n))
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


    if message.content.startswith("!gambling") and not msg == "!gambling":
        try:
            print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

            amount = message.content[10:]
            amount = int(amount)
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    if sheet["Y" + str(i)].value == 0:
                        w = "도박은 최소 100 M COIN 부터 가능합니다."
                        gamble = "도박"

                        nextg = "다음기회에~"
                        con = "축하드려요"
                        
                        n = "님 소지하고 계신 M COIN 이 부족합니다"
                        
                    elif sheet["Y" + str(i)].value == 1:
                        w = "Gambling is possible from a minimum of 100 M COIN"
                        gamble = "gamble"

                        nextg = "Next time~"
                        con = "congratulations"
                        
                        n = "You do not have enough M COIN"
                        
                    elif sheet["Y" + str(i)].value == 2:
                        w = "賭博は少なくとも100 M COINから可能です"
                        gamble = "ギャンブル"

                        nextg = "次の機会に〜"
                        con = "おめでとうございます"
                        
                        n = "様所持しているM COINが不足して"
                        
                    elif sheet["Y" + str(i)].value == 3:
                        w = "至少 100 M COIN 即可进行赌博"
                        gamble = "赌博"

                        nextg = "下次吧~"
                        con = "祝贺"
                        
                        n = "你没有足够的 M COIN"
                        
                    else:
                        w = "도박은 최소 100 M COIN 부터 가능합니다."
                        gamble = "도박"

                        nextg = "다음기회에~"
                        con = "축하드려요"

                        n = "님 소지하고 계신 M COIN 이 부족합니다"
                        


                    if amount < 100:
                        await message.channel.send(message.author.mention + w)
                    else:
                        if sheet["C" + str(i)].value >= amount:
                            for i in range(1, 1001):
                                if sheet["B" + str(i)].value == str(ui):
                                    sum = sheet["C" + str(i)].value
                                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value - amount
                                    sheet["N" + str(i)].value = sheet["N" + str(i)].value + 1
                                    r = (1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,5,5,5,5,5,5,5,5,5,6,6,6,6,6,7)
                                    rr = str(random.sample(r, 1))
                                    rr = rr[1]
                                    rr = int(rr)
                                    
                                    if sheet["N" + str(i)].value == 100:
                                        embed = discord.Embed(title="**《업적달성》**",description="『도박중독』 - 도박 100번 하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention,embed=embed)
                                        sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『도박중독』 "
                                    if rr == 1:
                                        embed = discord.Embed(title=gamble, description="```0.25x\n{}```".format(nextg),timesamp=datetime.datetime.now(pytz.timezone('UTC')))
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 0.25)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 2:
                                        embed = discord.Embed(title=gamble,description="```0.5x\n{}```".format(nextg),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xa0a0a0)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 0.5)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 3:
                                        embed = discord.Embed(title=gamble, description="```1x...```",timesamp=datetime.datetime.now(pytz.timezone('UTC')))
                                        await message.channel.send(message.author.mention, embed=embed)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + amount
                                        break
                                    if rr == 4:
                                        embed = discord.Embed(title=gamble, description="```2x! \n{}!```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 2)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 5:
                                        embed = discord.Embed(title=gamble, description="```3x! {}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff2)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 3)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 6:
                                        embed = discord.Embed(title=gamble, description="```5x! {}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 5)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break
                                    if rr == 7:
                                        embed = discord.Embed(title=gamble,description="```10x! 🎉🎉!!!~{}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        a = round(amount * 10)
                                        summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                                        break

                            file.save("user.xlsx")
                        else:
                            await message.channel.send(message.author.mention + " {}".format(n))
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


    if msg == "!올인" or msg == "!all" or msg == "!all in" or msg == "!全在" or msg == "!オールイン":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    gamble = "도박"

                    nextg = "다음기회에~"
                    con = "축하드려요"
                    
                    n = "님 소지하고 계신 M COIN 이 부족합니다"
                    
                elif sheet["Y" + str(i)].value == 1:
                    gamble = "gamble"

                    nextg = "Next time~"
                    con = "congratulations"
                    
                    n = "You do not have enough M COIN"
                    
                elif sheet["Y" + str(i)].value == 2:
                    gamble = "ギャンブル"

                    nextg = "次の機会に〜"
                    con = "おめでとうございます"
                    
                    n = "様所持しているM COINが不足して"
                    
                elif sheet["Y" + str(i)].value == 3:
                    gamble = "赌博"

                    nextg = "下次吧~"
                    con = "祝贺"
                    
                    n = "你没有足够的 M COIN"
                    
                else:
                    gamble = "도박"

                    nextg = "다음기회에~"
                    con = "축하드려요"

                    n = "님 소지하고 계신 M COIN 이 부족합니다"

                amount = sheet["C" + str(i)].value
                summ = sheet["C" + str(i)].value - amount
                sheet["N" + str(i)].value = sheet["N" + str(i)].value + 1

                r = (1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,5,5,5,5,5,5,5,5,5,6,6,6,6,6,7)
                rr = str(random.sample(r, 1))
                rr = rr[1]
                rr = int(rr)
                
                if sheet["N" + str(i)].value == 100:
                    embed = discord.Embed(title="**《업적달성》**",description="『도박중독』 - 도박 100번 하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    await message.channel.send(message.author.mention,embed=embed)
                    sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『도박중독』 "
                if rr == 1:
                    embed = discord.Embed(title=gamble, description="```0.25x\n{}```".format(nextg),timesamp=datetime.datetime.now(pytz.timezone('UTC')))
                    await message.channel.send(message.author.mention, embed=embed)
                    a = round(amount * 0.25)
                    sheet["C" + str(i)].value = a
                    break
                if rr == 2:
                    embed = discord.Embed(title=gamble,description="```0.5x\n{}```".format(nextg),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xa0a0a0)
                    await message.channel.send(message.author.mention, embed=embed)
                    a = round(amount * 0.5)
                    sheet["C" + str(i)].value = a
                    break
                if rr == 3:
                    embed = discord.Embed(title=gamble, description="```1x...```",timesamp=datetime.datetime.now(pytz.timezone('UTC')))
                    await message.channel.send(message.author.mention, embed=embed)
                    sheet["C" + str(i)].value = a
                    break
                if rr == 4:
                    embed = discord.Embed(title=gamble, description="```2x! \n{}!```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                    await message.channel.send(message.author.mention, embed=embed)
                    a = round(amount * 2)
                    sheet["C" + str(i)].value = a
                    break
                if rr == 5:
                    embed = discord.Embed(title=gamble, description="```3x! {}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff2)
                    await message.channel.send(message.author.mention, embed=embed)
                    a = round(amount * 3)
                    sheet["C" + str(i)].value = a
                    break
                if rr == 6:
                    embed = discord.Embed(title=gamble, description="```5x! {}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                    await message.channel.send(message.author.mention, embed=embed)
                    a = round(amount * 5)
                    sheet["C" + str(i)].value = a
                    break
                if rr == 7:
                    embed = discord.Embed(title=gamble,description="```10x! 🎉🎉!!!~{}!!!~🎉🎉```".format(con),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xc083ff)
                    await message.channel.send(message.author.mention, embed=embed)
                    a = round(amount * 10)
                    sheet["C" + str(i)].value = a
                    break
        file.save("user.xlsx")


    if message.content.startswith("!계좌개설") or msg == "!개설" or msg == "!account opening" or msg == "!accountopening" or msg == "!Accountopening" or msg == "!开户" or msg == "!口座開設":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):

                if sheet["Y" + str(i)].value == 0:
                    al = "님 이미 계좌가 있습니다\n`!은행`"
                    qo = "계좌를 개설하시겠습니까?"
                    baut = "은행 자동알림"
                    baut_1 = "정상적으로 계좌가 개설되었습니다"

                    a_1 = "개설은행"
                    a_2 = "개설자"
                    a_3 = "계좌번호"
                    a_4 = "개설일"
                    a_5 = "`!은행`을 사용해 주세요!"

                elif sheet["Y" + str(i)].value == 1:
                    al = "you already have an account\n`!bank`"
                    qo = "Would you like to open an account?"
                    baut = "Bank automatic notification"
                    baut_1 = "Account has been successfully opened"

                    a_1 = "issuing bank"
                    a_2 = "Founder"
                    a_3 = "account number"
                    a_4 = "Opening date"
                    a_5 = "Please use `!bank`"

                elif sheet["Y" + str(i)].value == 2:
                    al = "様の口座が既に存在します\n`!銀行`"
                    qo = "口座を開設しますか？"
                    baut = "銀行の自動通知"
                    baut_1 = "通常口座が開設されました"

                    a_1 = "開設銀行"
                    a_2 = "開設者"
                    a_3 = "介さ番号"
                    a_4 = "開設"
                    a_5 = "`!銀行` を使用してください！"

                elif sheet["Y" + str(i)].value == 3:
                    al = "你已经有一个账户了吗\n`!银行`"
                    qo = "您想开户吗？"
                    baut = "银行自动通知"
                    baut_1 = "账户已成功开立"

                    a_1 = "开证银行"
                    a_2 = "创始人"
                    a_3 = "账号"
                    a_4 = "开幕日期"
                    a_5 = "请用 `!银行`"

                else:
                    al = "님 이미 계좌가 있습니다\n`!은행`"
                    qo = "계좌를 개설하시겠습니까?"
                    baut = "은행 자동알림"
                    baut_1 = "정상적으로 계좌가 개설되었습니다"

                    a_1 = "개설은행"
                    a_2 = "개설자"
                    a_3 = "계좌번호"
                    a_4 = "개설일"
                    a_5 = "`!은행`을 사용해 주세요!"



                if sheet["O" + str(i)].value == 1:
                    await message.channel.send(message.author.mention+al)

                if sheet["O" + str(i)].value == 0:
                    embed = discord.Embed(title="💰BANK🏧", description=qo,timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff117)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("⭕")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '⭕' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=15.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        embed.set_footer(text=f'{(message.author.name)}', icon_url=message.author.avatar_url)
                        await msg.delete()
                        await message.channel.send(message.author.mention, embed=embed)
                    else:
                        if str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="Cancel!", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            embed.set_footer(text=f'{(message.author.name)}', icon_url=message.author.avatar_url)
                            await msg.delete()
                            await message.channel.send(message.author.mention, embed=embed)

                        elif str(reaction.emoji) == "⭕":
                            embed = discord.Embed(title="**[ {} ]**\n{} : )".format(baut,baut_1),description="{} : M 은행\n{} : {}\n{} : 1515-{}\n{} : {}\n{}".format(a_1, a_2, message.author, a_3, ui, a_4, datetime.datetime.today(), a_5),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff117)
                            embed.set_footer(text=f'{(message.author.name)}', icon_url=message.author.avatar_url)

                            await message.author.send(embed=embed)
                            embed1 = discord.Embed(title="**[ {} ]**\n{} : )".format(baut,baut_1),description="{} : M 은행\n{} : {}\n{} : 1515-{}\n{} : {}\n{}".format(a_1, a_2, message.author, a_3, ui, a_4, datetime.datetime.today(), a_5),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff117)
                            embed1.set_footer(text=f'{(message.author.name)}', icon_url=message.author.avatar_url)

                            await msg.delete()
                            await message.channel.send(embed=embed1)
                            now = datetime.datetime.now()
                            ny = now.year
                            nm = now.month
                            nd = now.day
                            sheet["O" + str(i)].value = 1  #계좌 true or false
                            sheet["P" + str(i)].value = 0 #돈
                            sheet["Q" + str(i)].value = datetime.datetime.today() #날짜
                            sheet["R" + str(i)].value = str(1515)+"-"+str(ui) #계좌

                            sheet["S" + str(i)].value = ny #년
                            sheet["T" + str(i)].value = nm #월
                            sheet["U" + str(i)].value = nd #일
        file.save("user.xlsx")


    if message.content.startswith("!계좌") and not msg == "!계좌개설" or msg == "!account" or msg == "!Account" or msg == "!口座" or msg == "!帐户":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):

                if sheet["Y" + str(i)].value == 0:
                    a_1 = "**[ 은행 ]**\n계좌 내역 : )"
                    a_2 = "은행"
                    a_3 = "님"
                    a_4 = "계좌번호"
                    a_5 = "잔액"
                    a_6 = "개설일"

                    b_1 = "계좌정보가 DM으로 전송되었습니다!"
                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"

                    
                elif sheet["Y" + str(i)].value == 1:
                    a_1 = "**[ Bank ]**\nAccount Details : )"
                    a_2 = "Bank"
                    a_3 = " "
                    a_4 = "Account Number"
                    a_5 = "Balance"
                    a_6 = "Opening date"

                    b_1 = "Your account information has been sent via DM!"
                    b_2 = "**You do not have an account**\nPlease use `!accountopening`****"

                    
                elif sheet["Y" + str(i)].value == 2:
                    a_1 = "**[ 銀行 ]**\n口座内訳：）"
                    a_2 = "銀行"
                    a_3 = "様"
                    a_4 = "口座番号"
                    a_5 = "残高"
                    a_6 = "開設"

                    b_1 = "口座情報がDMに送信されました！"
                    b_2 = "様**口座はありません**\n`！口座開設` **を使用してください**"

                    
                elif sheet["Y" + str(i)].value == 3:
                    a_1 = "**[ 银行 ]**\n账户详情 : )"
                    a_2 = "银行"
                    a_3 = "先生"
                    a_4 = "账号"
                    a_5 = "平衡"
                    a_6 = "开幕日期"

                    b_1 = "您的帐户信息已通过 DM 发送！"
                    b_2 = "**您还没有账户**\n**请使用** `!开户`"


                else:
                    a_1 = "**[ 은행 ]**\n계좌 내역 : )"
                    a_2 = "은행"
                    a_3 = "님"
                    a_4 = "계좌번호"
                    a_5 = "잔액"
                    a_6 = "개설일"

                    b_1 = "계좌정보가 DM으로 전송되었습니다!"
                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"

                if sheet["O" + str(i)].value == 1:
                    r = sheet["R" + str(i)].value
                    m = sheet["P" + str(i)].value
                    t = sheet["Q" + str(i)].value
                    embed = discord.Embed(title=a_1,description="{} : M 은행\n{} {}\n{} : {}\n{} : {}\n{} : {}".format(a_2, message.author, a_3, a_4, r, a_5, m, a_6, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff117)
                    embed.set_footer(text=f'{(message.author)}', icon_url=message.author.avatar_url)
                    await message.author.send(embed=embed)
                    await message.channel.send(message.author.mention + b_1)


                if sheet["O" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + b_2)


    if message.content.startswith("!은행") or msg == "!Bank" or msg == "!bank" or msg == "!銀行" or msg == "!银行":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):


                if sheet["Y" + str(i)].value == 0:
                    a_1 = "**[ 은행 ]**"
                    a_2 = "은행 입니다\n무었을 도와드릴까요?\n\n입금 : 1️⃣\n출금 : 2️⃣\n잔액 : 3️⃣\n입니다"
                    a_3 = "사용법 :  !입금 [입금액]"
                    a_4 = "사용법 :  !출금 [출금액]"
                    a_5 = "님의 잔액"

                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"

                    
                elif sheet["Y" + str(i)].value == 1:
                    a_1 = "**[ Bank ]**"
                    a_2 = "is a bank\nWhat can I help you with?\n\nDeposit: 1️⃣\nWithdrawal: 2️⃣\nBalance: 3️⃣"
                    a_3 = "How to use :  !deposit [deposit amount]"
                    a_4 = "How to use :  !withdraw [Withdrawal amount]"
                    a_5 = "'s balance"

                    b_2 = "**You do not have an account**\nPlease use `!accountopening`****"

                    
                elif sheet["Y" + str(i)].value == 2:
                    a_1 = "**[ 銀行 ]**"
                    a_2 = "銀行です\nなにを手伝ってくれる？\n\n入金：1️⃣\n出金：2️⃣\n残高：3️⃣\nです"
                    a_3 = "使い方 :  !入金 [デポジット]"
                    a_4 = "使い方 :  !出金 [チュルグムエク]"
                    a_5 = "さんの残高"

                    b_2 = "様**口座はありません**\n`！口座開設` **を使用してください**"

                    
                elif sheet["Y" + str(i)].value == 3:
                    a_1 = "**[ 银行 ]**"
                    a_2 = "是银行\n有什么可以帮您？\n\n入金：1️⃣\n出金：2️⃣\n余额：3️⃣"
                    a_3 = "如何使用 :  !订金 [存款金额]"
                    a_4 = "如何使用 :  !提取 [提款金额]"
                    a_5 = "余额"

                    b_2 = "**您还没有账户**\n**请使用** `!开户`"


                else:
                    a_1 = "**[ 은행 ]**"
                    a_2 = "은행 입니다\n무었을 도와드릴까요?\n\n입금 : 1️⃣\n출금 : 2️⃣\n잔액 : 3️⃣\n입니다"
                    a_3 = "사용법 :  !입금 [입금액]"
                    a_4 = "사용법 :  !출금 [출금액]"
                    a_5 = "님의 잔액"

                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"


                if sheet["O" + str(i)].value == 1:
                    embed = discord.Embed(title=a_1, description=a_2,timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff117)
                    embed.set_footer(text=f'{(message.author)}', icon_url=message.author.avatar_url)
                    msg = await message.channel.send(embed=embed)
                    await msg.add_reaction("1️⃣")
                    await msg.add_reaction("2️⃣")
                    await msg.add_reaction("3️⃣")

                if sheet["O" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + b_2)

                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '1️⃣' or user == message.author and str(reaction.emoji) == '2️⃣' or user == message.author and str(reaction.emoji) == '3️⃣'

                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=30.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="Cancel!", description="", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                    embed.set_footer(text=f'{(message.author)}', icon_url=message.author.avatar_url)

                    await msg.delete()
                    await message.channel.send(embed=embed)

                else:
                    if str(reaction.emoji) == "1️⃣":
                        await msg.delete()
                        embed = discord.Embed(title=a_1, description=a_3,timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                        embed.set_footer(text=f'{(message.author)}', icon_url=message.author.avatar_url)
                        await message.channel.send(embed=embed)

                    elif str(reaction.emoji) == "2️⃣":
                        await msg.delete()
                        embed = discord.Embed(title=a_1, description=a_4,timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                        await message.channel.send(embed=embed)

                    elif str(reaction.emoji) == "3️⃣":
                        await msg.delete()
                        embed = discord.Embed(title=a_1, description="{}{} : {}".format(message.author.name, a_5, sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x83fff8)
                        embed.set_footer(text=f'{(message.author)}', icon_url=message.author.avatar_url)
                        await message.channel.send(embed=embed)


    if message.content.startswith("!입금") and not msg == "!입금" and not msg == "!입금 " or msg == "!deposit" or msg == "!Deposit" or msg == "!入金" or msg == "!订金":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        amount = message.content[4:]
        amount = int(amount)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    a_1 = " 님 소지하고 계신 M COIN 이 부족합니다"
                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"
                    
                elif sheet["Y" + str(i)].value == 1:
                    a_1 = "You do not have enough M COIN"
                    b_2 = "**You do not have an account**\nPlease use `!accountopening`****"
                    
                elif sheet["Y" + str(i)].value == 2:
                    a_1 = "様所持しているM COINが不足して"
                    b_2 = "様**口座はありません**\n`！口座開設` **を使用してください**"
                    
                elif sheet["Y" + str(i)].value == 3:
                    a_1 = "你没有足够的 M COIN"
                    b_2 = "**您还没有账户**\n**请使用** `!开户`"
                    
                else:
                    a_1 = " 님 소지하고 계신 M COIN 이 부족합니다"
                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"
                    
                if sheet["O" + str(i)].value == 1:
                        if sheet["C" + str(i)].value >= amount:
                            sheet["C" + str(i)].value = sheet["C" + str(i)].value - amount
                            sheet["P" + str(i)].value = sheet["P" + str(i)].value + amount

                            if sheet["Y" + str(i)].value == 0:
                                embed = discord.Embed(title="**[은행]**", description="{}님 {} M COIN이 입금이 완료되었습니다\n`!계좌` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                            elif sheet["Y" + str(i)].value == 1:
                                embed = discord.Embed(title="**[Bank]**", description="{}, {} M COIN has been deposited.\n`!account` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                            elif sheet["Y" + str(i)].value == 2:
                                embed = discord.Embed(title="**[銀行]**", description="{}님 {} M COINが入金が完了しました。\n`!口座` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                            elif sheet["Y" + str(i)].value == 3:
                                embed = discord.Embed(title="**[银行]**", description="{}님 {} M COIN 已存入。\n`!帐户` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                            else:
                                embed = discord.Embed(title="**[은행]**", description="{}님 {} M COIN이 입금이 완료되었습니다\n`!계좌` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                            await message.channel.send(embed=embed)
                            u = sheet["W" + str(i)].value
                            if "『저축』" in u:
                                pass
                            else:
                                embeda = discord.Embed(title="**《업적달성》**",description="『저축』 - 입금하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention,embed=embeda)
                                sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『저축』 "

                            break
                        else:
                            await message.channel.send(message.author.mention + a_1)

                if sheet["O" + str(i)].value == 0:
                    await message.channel.send(message.author.mention + b_2)
        file.save("user.xlsx")


    if message.content.startswith("!출금") and not msg == "!출금" and not msg == "!출금" or msg == "!withdraw" or msg == "!Withdraw" or msg == "!出金" or msg == "!提取":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        amount = message.content[4:]
        amount = int(amount)
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):

                if sheet["Y" + str(i)].value == 0:
                    a_1 = " 님 소지하고 계신 M COIN 이 부족합니다"
                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"
                    
                elif sheet["Y" + str(i)].value == 1:
                    a_1 = "You do not have enough M COIN"
                    b_2 = "**You do not have an account**\nPlease use `!accountopening`****"
                    
                elif sheet["Y" + str(i)].value == 2:
                    a_1 = "様所持しているM COINが不足して"
                    b_2 = "様**口座はありません**\n`！口座開設` **を使用してください**"
                    
                elif sheet["Y" + str(i)].value == 3:
                    a_1 = "你没有足够的 M COIN"
                    b_2 = "**您还没有账户**\n**请使用** `!开户`"
                    
                else:
                    a_1 = " 님 소지하고 계신 M COIN 이 부족합니다"
                    b_2 = "님 **계좌가 없습니다**\n`!계좌개설`**을 사용해 주세요**"


                if sheet["O" + str(i)].value == 1: # 계좌 존재 확인
                    if sheet["P" + str(i)].value >= amount: # 계좌돈이 출금할 돈보다 작거나 같다
                        sheet["C" + str(i)].value = sheet["C" + str(i)].value + amount
                        sheet["P" + str(i)].value = sheet["P" + str(i)].value - amount

                        if sheet["Y" + str(i)].value == 0:
                            embed = discord.Embed(title="**[은행]**", description="{}님 {} M COIN이 출금이 완료되었습니다\n`!계좌` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                        elif sheet["Y" + str(i)].value == 1:
                            embed = discord.Embed(title="**[Bank]**", description="{}, {} Withdrawal of M COIN has been completed.\n`!account` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                        elif sheet["Y" + str(i)].value == 2:
                            embed = discord.Embed(title="**[銀行]**", description="{}님 {} M COINが出金が完了しました。\n`!口座` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                        elif sheet["Y" + str(i)].value == 3:
                            embed = discord.Embed(title="**[银行]**", description="{}님 {} M COIN提现完成。\n`!帐户` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)

                        else:
                            embed = discord.Embed(title="**[은행]**", description="{}님 {} M COIN이 출금이 완료되었습니다\n`!계좌` : {} M COIN".format(message.author.name,amount,sheet["P" + str(i)].value),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xfff117)
                        
                        await message.channel.send(embed=embed)
                        file.save("user.xlsx")
                        break

                    else:
                        await message.channel.send(message.author.mention + a_1)

                if sheet["O" + str(i)].value == 0:
                    await message.channel.send(message.author.mention+b_2)


# 타이머

    if message.content.startswith("!타이머") or message.content.startswith("!计时器") and not msg == "!타이머" and not msg == "!타이머" and not msg == "!计时器" and not msg == "!计时器 ":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sc = sheet["C" + str(i)].value
                    bc = sheet["P" + str(i)].value

                    if sheet["Y" + str(i)].value == 0:
                        w = "초 타이머를 시작합니다\n남은시간"
                        a = "초 타이머\n남은시간"
                        endt = "님 타이머가 종료되었습니다!"
                        upl = "타이머는 양수만 가능합니다!"
                        limit = "한번에 최대 100초까지 설정 가능합니다!"

                    elif sheet["Y" + str(i)].value == 1:
                        w = " second timer\nTime remaining"
                        a = " second timer\nTime remaining"
                        endt = "Your timer has expired!"
                        upl = "Timer can only be positive!"
                        limit = "You can set up to 100 seconds at a time!"

                    elif sheet["Y" + str(i)].value == 2:
                        w = " 秒のタイマーを開始します\n残り時間"
                        a = " 秒のタイマー\n残り時間"
                        endt = "様タイマーが終了しました！"
                        upl = "タイマーは、正のみ可能です！"
                        limit = "一度に最大100秒まで設定可能です！"

                    elif sheet["Y" + str(i)].value == 3:
                        w = " 秒计时器\n剩余时间"
                        a = " 秒计时器\n剩余时间"
                        endt = "您的计时器已过期！"
                        upl = "定时器只能是正的！"
                        limit = "您一次最多可以设置 100 秒！"




            t = message.content[5:]
            t = int(t)
            try:
                if t <= 100:
                    if t > 0:
                        tt = t
                        embed = discord.Embed(title="**[Timer]**", description="{}{} : {}".format(tt, w, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6699ff)
                        msg = await message.channel.send(embed=embed)

                        while True:
                            if t != 0:
                                t = t - 1
                                embed = discord.Embed(title="**[Timer]**", description="{}{} : {}".format(tt, a, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6699ff)
                                await msg.edit(embed=embed)
                                await asyncio.sleep(1)

                            elif t == 0:
                                await msg.delete()
                                await message.channel.send(message.author.mention + endt)
                                break
                    else:
                        await message.channel.send(message.author.mention + upl)
                else:
                    await message.channel.send(message.author.mention + limit)
            
            except:
                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                un = message.author.name
                ui = message.author.id
                us = message.author.avatar_url
                now = datetime.datetime.now()
                ny = now.year
                nm = now.month
                nd = now.day
                nh = now.hour
                nmm = now.minute
                embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                msg = await message.channel.send(message.author.mention,embed=embed)
                await msg.add_reaction("<a:check:848042146044575767>")
                await msg.add_reaction("❌")
                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)
                else:
                    if str(reaction.emoji) == "<a:check:848042146044575767>":
                        for i in range(1, 1001):
                            if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                sheet["A" + str(i)].value = str(un)
                                sheet["B" + str(i)].value = str(ui)
                                sheet["C" + str(i)].value = 5000
                                sheet["D" + str(i)].value = "안녕하세요"
                                sheet["E" + str(i)].value = str(us)
                                sheet["F" + str(i)].value = ny
                                sheet["G" + str(i)].value = nm
                                sheet["H" + str(i)].value = nd
                                sheet["I" + str(i)].value = ny
                                sheet["J" + str(i)].value = nm
                                sheet["K" + str(i)].value = nd
                                sheet["L" + str(i)].value = nh
                                sheet["M" + str(i)].value = nmm
                                sheet["N" + str(i)].value = 0
                                sheet["O" + str(i)].value = 0
                                sheet["AA" + str(i)].value = 0
                                sheet["AB" + str(i)].value = 0
                                sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                file.save("user.xlsx")
                                break
                    elif str(reaction.emoji) == "❌":
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)

        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return

    if message.content.startswith("!timer") and not msg == "!timer":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sc = sheet["C" + str(i)].value
                    bc = sheet["P" + str(i)].value

                    if sheet["Y" + str(i)].value == 0:
                        w = "초 타이머를 시작합니다\n남은시간"
                        a = "초 타이머\n남은시간"
                        endt = "님 타이머가 종료되었습니다!"
                        upl = "타이머는 양수만 가능합니다!"
                        limit = "한번에 최대 100초까지 설정 가능합니다!"

                    elif sheet["Y" + str(i)].value == 1:
                        w = " second timer\nTime remaining"
                        a = " second timer\nTime remaining"
                        endt = "Your timer has expired!"
                        upl = "Timer can only be positive!"
                        limit = "You can set up to 100 seconds at a time!"

                    elif sheet["Y" + str(i)].value == 2:
                        w = " 秒のタイマーを開始します\n残り時間"
                        a = " 秒のタイマー\n残り時間"
                        endt = "様タイマーが終了しました！"
                        upl = "タイマーは、正のみ可能です！"
                        limit = "一度に最大100秒まで設定可能です！"

                    elif sheet["Y" + str(i)].value == 3:
                        w = " 秒计时器\n剩余时间"
                        a = " 秒计时器\n剩余时间"
                        endt = "您的计时器已过期！"
                        upl = "定时器只能是正的！"
                        limit = "您一次最多可以设置 100 秒！"




            t = message.content[7:]
            t = int(t)
            try:
                if t <= 100:
                    if t > 0:
                        tt = t
                        embed = discord.Embed(title="**[Timer]**", description="{}{} : {}".format(tt, w, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6699ff)
                        msg = await message.channel.send(embed=embed)

                        while True:
                            if t != 0:
                                t = t - 1
                                embed = discord.Embed(title="**[Timer]**", description="{}{} : {}".format(tt, a, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6699ff)
                                await msg.edit(embed=embed)
                                await asyncio.sleep(1)

                            elif t == 0:
                                await msg.delete()
                                await message.channel.send(message.author.mention + endt)
                                break
                    else:
                        await message.channel.send(message.author.mention + upl)
                else:
                    await message.channel.send(message.author.mention + limit)
            
            except:
                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                un = message.author.name
                ui = message.author.id
                us = message.author.avatar_url
                now = datetime.datetime.now()
                ny = now.year
                nm = now.month
                nd = now.day
                nh = now.hour
                nmm = now.minute
                embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                msg = await message.channel.send(message.author.mention,embed=embed)
                await msg.add_reaction("<a:check:848042146044575767>")
                await msg.add_reaction("❌")
                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)
                else:
                    if str(reaction.emoji) == "<a:check:848042146044575767>":
                        for i in range(1, 1001):
                            if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                sheet["A" + str(i)].value = str(un)
                                sheet["B" + str(i)].value = str(ui)
                                sheet["C" + str(i)].value = 5000
                                sheet["D" + str(i)].value = "안녕하세요"
                                sheet["E" + str(i)].value = str(us)
                                sheet["F" + str(i)].value = ny
                                sheet["G" + str(i)].value = nm
                                sheet["H" + str(i)].value = nd
                                sheet["I" + str(i)].value = ny
                                sheet["J" + str(i)].value = nm
                                sheet["K" + str(i)].value = nd
                                sheet["L" + str(i)].value = nh
                                sheet["M" + str(i)].value = nmm
                                sheet["N" + str(i)].value = 0
                                sheet["O" + str(i)].value = 0
                                sheet["AA" + str(i)].value = 0
                                sheet["AB" + str(i)].value = 0
                                sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                file.save("user.xlsx")
                                break
                    elif str(reaction.emoji) == "❌":
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)

        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


    if message.content.startswith("!タイマー") and not msg == "!タイマー":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sc = sheet["C" + str(i)].value
                    bc = sheet["P" + str(i)].value

                    if sheet["Y" + str(i)].value == 0:
                        w = "초 타이머를 시작합니다\n남은시간"
                        a = "초 타이머\n남은시간"
                        endt = "님 타이머가 종료되었습니다!"
                        upl = "타이머는 양수만 가능합니다!"
                        limit = "한번에 최대 100초까지 설정 가능합니다!"

                    elif sheet["Y" + str(i)].value == 1:
                        w = " second timer\nTime remaining"
                        a = " second timer\nTime remaining"
                        endt = "Your timer has expired!"
                        upl = "Timer can only be positive!"
                        limit = "You can set up to 100 seconds at a time!"

                    elif sheet["Y" + str(i)].value == 2:
                        w = " 秒のタイマーを開始します\n残り時間"
                        a = " 秒のタイマー\n残り時間"
                        endt = "様タイマーが終了しました！"
                        upl = "タイマーは、正のみ可能です！"
                        limit = "一度に最大100秒まで設定可能です！"

                    elif sheet["Y" + str(i)].value == 3:
                        w = " 秒计时器\n剩余时间"
                        a = " 秒计时器\n剩余时间"
                        endt = "您的计时器已过期！"
                        upl = "定时器只能是正的！"
                        limit = "您一次最多可以设置 100 秒！"


            t = message.content[6:]
            t = int(t)
            try:
                if t <= 100:
                    if t > 0:
                        tt = t
                        embed = discord.Embed(title="**[Timer]**", description="{}{} : {}".format(tt, w, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6699ff)
                        msg = await message.channel.send(embed=embed)

                        while True:
                            if t != 0:
                                t = t - 1
                                embed = discord.Embed(title="**[Timer]**", description="{}{} : {}".format(tt, a, t),timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6699ff)
                                await msg.edit(embed=embed)
                                await asyncio.sleep(1)

                            elif t == 0:
                                await msg.delete()
                                await message.channel.send(message.author.mention + endt)
                                break
                    else:
                        await message.channel.send(message.author.mention + upl)
                else:
                    await message.channel.send(message.author.mention + limit)
            
            except:
                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                un = message.author.name
                ui = message.author.id
                us = message.author.avatar_url
                now = datetime.datetime.now()
                ny = now.year
                nm = now.month
                nd = now.day
                nh = now.hour
                nmm = now.minute
                embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                msg = await message.channel.send(message.author.mention,embed=embed)
                await msg.add_reaction("<a:check:848042146044575767>")
                await msg.add_reaction("❌")
                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)
                else:
                    if str(reaction.emoji) == "<a:check:848042146044575767>":
                        for i in range(1, 1001):
                            if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                sheet["A" + str(i)].value = str(un)
                                sheet["B" + str(i)].value = str(ui)
                                sheet["C" + str(i)].value = 5000
                                sheet["D" + str(i)].value = "안녕하세요"
                                sheet["E" + str(i)].value = str(us)
                                sheet["F" + str(i)].value = ny
                                sheet["G" + str(i)].value = nm
                                sheet["H" + str(i)].value = nd
                                sheet["I" + str(i)].value = ny
                                sheet["J" + str(i)].value = nm
                                sheet["K" + str(i)].value = nd
                                sheet["L" + str(i)].value = nh
                                sheet["M" + str(i)].value = nmm
                                sheet["N" + str(i)].value = 0
                                sheet["O" + str(i)].value = 0
                                sheet["AA" + str(i)].value = 0
                                sheet["AB" + str(i)].value = 0
                                sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                file.save("user.xlsx")
                                break
                    elif str(reaction.emoji) == "❌":
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)

        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


# 관리
    #_____

    if message.content.startswith("!슬로우") and not msg == "!슬로우" and not msg == "!슬로우 " or message.content.startswith("!slow") or message.content.startswith("!Slow") and not msg == "!slow" and not msg == "!Slow"  and not msg == "!slow " and not msg == "!Slow " or message.content.startswith("!スロー") and not msg == "!スロー" and not msg == "!スロー " or message.content.startswith("!慢") and not msg == "!慢":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            m = message.content.split(" ")
            m = int(m[1])
            i = (message.author.guild_permissions.administrator)
            mi = 789670002163974145

            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    if sheet["Y" + str(i)].value == 0:
                        w = "채널에 슬로우 모드"
                        w_1 = "초를 적용하였습니다"

                        e = "권한부족"

                    elif sheet["Y" + str(i)].value == 1:
                        w = "Channel applied in slow mode"
                        w_1 = " seconds"

                        e = "lack of authority"

                    elif sheet["Y" + str(i)].value == 2:
                        w = "チャンネルにスローモード"
                        w_1 = " 秒を適用しました"

                        e = "権限不足"

                    elif sheet["Y" + str(i)].value == 3:
                        w = "通道应用慢速模式"
                        w_1 = " 秒"

                        e = "缺乏权威"

            try:
                if i is True or message.author.id == mi:
                    await message.channel.send(message.author.mention+" {}{} {}{}.".format(message.channel.mention, w, m, w_1))
                    await message.channel.edit(slowmode_delay=m)
            except:
                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                un = message.author.name
                ui = message.author.id
                us = message.author.avatar_url
                now = datetime.datetime.now()
                ny = now.year
                nm = now.month
                nd = now.day
                nh = now.hour
                nmm = now.minute
                embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                msg = await message.channel.send(message.author.mention,embed=embed)
                await msg.add_reaction("<a:check:848042146044575767>")
                await msg.add_reaction("❌")
                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)
                else:
                    if str(reaction.emoji) == "<a:check:848042146044575767>":
                        for i in range(1, 1001):
                            if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                sheet["A" + str(i)].value = str(un)
                                sheet["B" + str(i)].value = str(ui)
                                sheet["C" + str(i)].value = 5000
                                sheet["D" + str(i)].value = "안녕하세요"
                                sheet["E" + str(i)].value = str(us)
                                sheet["F" + str(i)].value = ny
                                sheet["G" + str(i)].value = nm
                                sheet["H" + str(i)].value = nd
                                sheet["I" + str(i)].value = ny
                                sheet["J" + str(i)].value = nm
                                sheet["K" + str(i)].value = nd
                                sheet["L" + str(i)].value = nh
                                sheet["M" + str(i)].value = nmm
                                sheet["N" + str(i)].value = 0
                                sheet["O" + str(i)].value = 0
                                sheet["AA" + str(i)].value = 0
                                sheet["AB" + str(i)].value = 0
                                sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                file.save("user.xlsx")
                                break
                    elif str(reaction.emoji) == "❌":
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)


            if i is False:
                embed = discord.Embed(title="Error : Insufficient authority", description="Error : {}".format(e), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                await message.author.send(embed=embed)

        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


    if msg == "!슬로우" or msg == "!slow" or msg == "!Slow" or msg == "!スロー" or msg == "!慢":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            i = (message.author.guild_permissions.administrator)
            mi = 789670002163974145
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    if sheet["Y" + str(i)].value == 0:
                        w = "채널에 슬로우 모드를 초기화 했습니다."
                        e = "권한부족"

                    elif sheet["Y" + str(i)].value == 1:
                        w = "Initialized slow mode on the channel."
                        e = "lack of authority"

                    elif sheet["Y" + str(i)].value == 2:
                        w = "チャンネルにスローモードを初期化しました。"
                        e = "権限不足"

                    elif sheet["Y" + str(i)].value == 3:
                        w = "在通道上初始化慢模式。"
                        e = "缺乏权威"

     

                if i is True or message.author.id == mi:
                    try:     
                        await message.channel.edit(slowmode_delay=0)
                        await message.channel.send(message.author.mention+" {}{}".format(message.channel.mention,w))
                    except:
                        file = openpyxl.load_workbook("user.xlsx")
                        sheet = file.active
                        un = message.author.name
                        ui = message.author.id
                        us = message.author.avatar_url
                        now = datetime.datetime.now()
                        ny = now.year
                        nm = now.month
                        nd = now.day
                        nh = now.hour
                        nmm = now.minute
                        embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        msg = await message.channel.send(message.author.mention,embed=embed)
                        await msg.add_reaction("<a:check:848042146044575767>")
                        await msg.add_reaction("❌")
                        def check(reaction, user):
                            if user.bot == 1:  # 봇이면 패스
                                return None
                            return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                        except asyncio.TimeoutError:
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)
                        else:
                            if str(reaction.emoji) == "<a:check:848042146044575767>":
                                for i in range(1, 1001):
                                    if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                        sheet["A" + str(i)].value = str(un)
                                        sheet["B" + str(i)].value = str(ui)
                                        sheet["C" + str(i)].value = 5000
                                        sheet["D" + str(i)].value = "안녕하세요"
                                        sheet["E" + str(i)].value = str(us)
                                        sheet["F" + str(i)].value = ny
                                        sheet["G" + str(i)].value = nm
                                        sheet["H" + str(i)].value = nd
                                        sheet["I" + str(i)].value = ny
                                        sheet["J" + str(i)].value = nm
                                        sheet["K" + str(i)].value = nd
                                        sheet["L" + str(i)].value = nh
                                        sheet["M" + str(i)].value = nmm
                                        sheet["N" + str(i)].value = 0
                                        sheet["O" + str(i)].value = 0
                                        sheet["AA" + str(i)].value = 0
                                        sheet["AB" + str(i)].value = 0
                                        sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                        embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                        file.save("user.xlsx")
                                        break
                            elif str(reaction.emoji) == "❌":
                                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                                await msg.delete()
                                await message.channel.send(embed=embed)
                                break

            if i is False:
                embed = discord.Embed(title="Error : Insufficient authority", description="Error : {}".format(e), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff3737)
                await message.author.send(embed=embed)
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


# 롤 전적

    if message.content.startswith("!롤") or message.content.startswith("!LeagueofLegends") or message.content.startswith("!LoL") or message.content.startswith("!lol") or message.content.startswith("!Lol") or message.content.startswith("!リーグ・オブ・レジェンド") or message.content.startswith("!LOL") or message.content.startswith("!lOl"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id

            Name = message.content.split(" ")
            Name = Name[1]
            SummonerName = "" # 소환사 이름 
            Ranking = ""  # 랭킹
            TierUnranked = ""  # 언랭 판별룡 티어
            LeagueType = [] # 리그타입
            Tier = []  # 티어
            LP = []  # LP (리그 포인트)
            Wins = []  # 승리 판수
            Losses = [] # 패배 판수
            Ratio = [] #승률

            try:
                url='https://www.op.gg/summoner/userName=' + Name 
                hdr = {'Accept-Language': 'ko_KR,en;q=0.8', 'User-Agent': ('Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.70 Mobile Safari/537.36')} 
                req = requests.get(url , headers=hdr) 
                html = req.text 
                soup = BeautifulSoup(html, 'html.parser')

                # 프사, 티어
                rt = requests.get(url)
                st = BeautifulSoup(rt.text, "html.parser")

                # 프사

                p = st.find("div", class_="ProfileIcon").find_all("img",class_="ProfileImage")
                p = str(p)
                p = p[32:]
                p = p[:-4]
                p = "https:"+p

                #레벨
                level = st.find("div", class_="ProfileIcon").find_all("span")
                level = str(level)
                level = level[39:]
                level = level[:-8]

                # 티어
                t = st.find("div", class_="SummonerRatingMedium").find_all("img",class_="Image")
                t = str(t)
                t = t[25:]
                t = t[:-4]
                t = "https:"+t

            # 최근게임

                # 전

                cm = st.find("div", class_="WinRatioTitle").find_all("span",class_="total") 
                cm = str(cm)
                cm = cm[21:]
                cm = cm[:-8]

                # 승

                cw = st.find("div", class_="WinRatioTitle").find_all("span",class_="win") 
                cw = str(cw)
                cw = cw[19:]
                cw = cw[:-8]

                # 패

                cl = st.find("div", class_="WinRatioTitle").find_all("span",class_="lose") 
                cl = str(cl)
                cl = cl[20:]
                cl = cl[:-8]

            # 최근게임 - KDA
                # kill
                kill = st.find("td", class_="KDA").find_all("span",class_="Kill") 
                kill = str(kill)
                kill = kill[20:]
                kill = kill[:-8]


                #Death
                death = st.find("td", class_="KDA").find_all("span",class_="Death") 
                death = str(death)
                death = death[21:]
                death = death[:-8]

                #assist
                assist = st.find("td", class_="KDA").find_all("span",class_="Assist") 
                assist = str(assist)
                assist = assist[22:]
                assist = assist[:-8]

                #KDARatio
                KDARatio = st.find("div", class_="KDARatio").find_all("span",class_="KDARatio") 
                KDARatio = str(KDARatio)
                KDARatio = KDARatio[24:]
                KDARatio = KDARatio[:-8]

                #KDARatio %
                KDARatiop = st.find("div", class_="KDARatio").find_all("span") 
                KDARatiop = str(KDARatiop)
                KDARatiop = KDARatiop[124:]
                KDARatiop = KDARatiop[:-8]




                # 소환사 이름 크롤링 
                for i in soup.select('div[class=SummonerName]'): 
                    SummonerName = i.text 

                # 랭킹 크롤링 
                for i in soup.select('span[class=ranking]'): 
                    Ranking = i.text 

                # 언랭 판별용 티어 크롤링 
                TierUnranked = soup.select('div.Cell') 

                # 리그 타입 크롤링 
                for i in soup.select('div[class=LeagueType]'): 
                    LeagueType.append(i.text) 

                # 티어 크롤링 
                for i in soup.select('div[class=Tier]'): 
                    Tier.append(i.text) 

                # 리그 포인트 크롤링 
                for i in soup.select('div[class=LP]'): 
                    LP.append(i.text) 

                # 승리 패배 판수 크롤링 
                for i in soup.select('span[class=Wins]'): 
                    Wins.append(i.text) 

                for i in soup.select('span[class=Losses]'): 
                    Losses.append(i.text) 

                # 승률 크롤링 
                for i in soup.select('span[class=Ratio]'): 
                    Ratio.append(i.text)

                #닉네임 합치기
                Namea = Name.replace(" ", "+")
                urla='https://www.op.gg/summoner/userName=' + Namea
            except:
                for i in range(1, 1001):
                    if sheet["B" + str(i)].value == str(ui):
                        if sheet["Y" + str(i)].value == 0:
                            slev = "소환사 레벨"
                            srank = "랭크"
                            stier = "티어"
                            ssch = "전적 검색 결과 입니다!"
                            sdeta = "자세히보기"

                        elif sheet["Y" + str(i)].value == 1:
                            slev = "Summoner Level"
                            srank = "Rank"
                            stier = "Tear"
                            ssch = "Player information search results!"
                            sdeta = "Read more"

                        elif sheet["Y" + str(i)].value == 2:
                            slev = "召喚師のレベル"
                            srank = "ランク"
                            stier = "ティア"
                            ssch = "プレイヤー情報検索結果です！"
                            sdeta = "続きを読む"

                        elif sheet["Y" + str(i)].value == 3:
                            slev = "召唤师等级"
                            srank = "秩"
                            stier = "眼泪"
                            ssch = "玩家信息搜索结果！"
                            sdeta = "阅读更多"

                try:
                    embed = discord.Embed(title=Name, description="\n{}: {}".format(slev, level), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xededed)
                    embed.add_field(name=srank, value="{}: Unranked".format(srank), inline=False)
                    embed.set_thumbnail(url=p) # 프사
                    embed.set_author(name="OP.GG League of Legends {}".format(ssch), icon_url=t) # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embed.add_field(name="ㅤ",value="[{}](<{}>)".format(sdeta,urla),inline=True)
                    embed.add_field(name="ㅤ",value="[OP.GG](<https://www.op.gg/>)",inline=True)
                    embed.add_field(name="ㅤ",value="[League of Legends](<https://na.leagueoflegends.com/ko-kr/>)",inline=True)
                    await message.channel.send(embed=embed)
                except:
                    await message.channel.send(embed=discord.Embed(title="Error", description = "검색결과가 없거나 `!가입`을 사용해주세요!\nIf there are no search results, please use '!Join'",color=0xff0000))


            if SummonerName != "": 
                if 'Unranked' in str(TierUnranked[0]): 

                    for i in range(1, 1001):
                        if sheet["B" + str(i)].value == str(ui):
                            if sheet["Y" + str(i)].value == 0:
                                slev = "소환사 레벨"
                                srank = "랭크"
                                stier = "티어"
                                ssch = "전적 검색 결과 입니다!"
                                sdeta = "자세히보기"

                            elif sheet["Y" + str(i)].value == 1:
                                slev = "Summoner Level"
                                srank = "Rank"
                                stier = "Tear"
                                ssch = "Player information search results!"
                                sdeta = "Read more"

                            elif sheet["Y" + str(i)].value == 2:
                                slev = "召喚師のレベル"
                                srank = "ランク"
                                stier = "ティア"
                                ssch = "プレイヤー情報検索結果です！"
                                sdeta = "続きを読む"

                            elif sheet["Y" + str(i)].value == 3:
                                slev = "召唤师等级"
                                srank = "秩"
                                stier = "眼泪"
                                ssch = "玩家信息搜索结果！"
                                sdeta = "阅读更多"

                    try:
                        embed = discord.Embed(title=Name, description="\n{}: {}".format(slev, level), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xededed)
                        embed.add_field(name=srank, value="{}: Unranked".format(srank), inline=False)
                        embed.set_thumbnail(url=p) # 프사
                        embed.set_author(name="OP.GG League of Legends {}".format(ssch), icon_url=t) # 맨위
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        embed.add_field(name="ㅤ",value="[{}](<{}>)".format(sdeta,urla),inline=True)
                        embed.add_field(name="ㅤ",value="[OP.GG](<https://www.op.gg/>)",inline=True)
                        embed.add_field(name="ㅤ",value="[League of Legends](<https://na.leagueoflegends.com/ko-kr/>)",inline=True)
                        await message.channel.send(embed=embed)
                    except:
                        await message.channel.send(embed=discord.Embed(title="Error", description = "검색결과가 없거나 `!가입`을 사용해주세요!\nIf there are no search results, please use '!Join'",color=0xff0000))


                else: 
                    for i in range(1, 1001):
                        if sheet["B" + str(i)].value == str(ui):
                            if sheet["Y" + str(i)].value == 0:
                                slev = "소환사 레벨"
                                srank = "랭크"
                                stier = "티어"
                                ssch = "전적 검색 결과 입니다!"
                                sdeta = "자세히보기"
                                sgame = "최근 게임"

                                sall = "전"
                                swin = "승"
                                slose = "패"
                                swinp = "승률"


                            elif sheet["Y" + str(i)].value == 1:
                                slev = "Summoner Level"
                                srank = "Rank"
                                stier = "Tear"
                                ssch = "Player information search results!"
                                sdeta = "Read more"
                                sgame = "Recent games"

                                sall = " Out of"
                                swin = " Wins"
                                slose = " Losses"
                                swinp = "Win rate"

                            elif sheet["Y" + str(i)].value == 2:
                                slev = "召喚師のレベル"
                                srank = "ランク"
                                stier = "ティア"
                                ssch = "プレイヤー情報検索結果です！"
                                sdeta = "続きを読む"
                                sgame = "最近のゲーム"

                                sall = "全"
                                swin = "勝"
                                slose = "敗"
                                swinp = "勝率"

                            elif sheet["Y" + str(i)].value == 3:
                                slev = "召唤师等级"
                                srank = "秩"
                                stier = "眼泪"
                                ssch = "玩家信息搜索结果！"
                                sdeta = "阅读更多"
                                sgame = "最近的比赛"

                                sall = "负"
                                swin = "胜"
                                slose = "败"
                                swinp = "赢率"

                    try:
                        embed = discord.Embed(title=Name, description="\n{}: {}".format(slev, level), timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xededed)
                        embed.add_field(name=srank, value="{}: ".format(stier) + Tier[0].strip('\n\t') + " | "+ LP[0] + "\n" + Wins[0] + " " + Losses[0] + " ({}: ".format(swinp) + Ratio[0]+")", inline=False)
                        embed.add_field(name=sgame, value="{}: {}{} {}{} {}{}\n{}/{}/{}\n{} ({})".format(sgame, cm, sall, cw, swin, cl, slose, kill, death, assist, KDARatio, KDARatiop), inline=False)
                        embed.set_thumbnail(url=p) # 프사
                        embed.set_author(name="OP.GG League of Legends {}".format(ssch), icon_url=t) # 맨위
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        embed.add_field(name="ㅤ",value="[{}](<{}>)".format(sdeta,urla),inline=True)
                        embed.add_field(name="ㅤ",value="[OP.GG](<https://www.op.gg/>)",inline=True)
                        embed.add_field(name="ㅤ",value="[League of Legends](<https://na.leagueoflegends.com/ko-kr/>)",inline=True)
                        await message.channel.send(embed=embed)
                    except:
                        await message.channel.send(embed=discord.Embed(title="Error", description = "검색결과가 없거나 `!가입`을 사용해주세요!\nIf there are no search results, please use '!Join'",color=0xff0000))

        

        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = "검색결과가 없습니다.\nNo results were found for your search.",color=0xff0000))


    if msg == "!롤" or msg == "!롤 " or msg == "!リーグ・オブ・レジェンド" or msg == "!リーグ・オブ・レジェンド " or msg == "!lol" or msg == "!Lol" or msg == "!LOL" or msg == "!LoL":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    w = "전적 검색법"


                elif sheet["Y" + str(i)].value == 1:
                    w = "How to search play history"


                elif sheet["Y" + str(i)].value == 2:
                    w = "プレイ履歴を検索する方法"


                elif sheet["Y" + str(i)].value == 3:
                    w = "如何搜索播放历史"

        try:
            embed = discord.Embed(title="OP.GG League of Legends {}".format(w), description="!롤 Nickname\n!lol Nickname", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xededed)
            embed.add_field(name="ㅤ",value="[OP.GG](<https://www.op.gg/>)",inline=True)
            embed.add_field(name="ㅤ",value="[League of Legends](<https://na.leagueoflegends.com/ko-kr/>)",inline=True)
            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
            await message.channel.send(embed=embed)

        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)


# 정보

    if msg == "!봇 정보" or msg == "!봇정보" or msg == "!봇 정보 " or msg == "!봇정보 " or msg == "botinfo" or msg == "!botinfo" or msg == "!Botinfo" or msg == "!ボットについて" or msg == "!ボット情報" or msg == "!机器人信息":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        
        try:

            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    s = len(client.guilds)
                    
                    V = cpuinfo.get_cpu_info()

                    cpu_b = V['brand_raw'] # 모델
                    cpu_c = V['hz_actual_friendly'] # 클럭

                    cpu_percent = psutil.cpu_percent() #사용률

                    cpu_count = psutil.cpu_count() # cpu 코어 + 쓰레드
                    cpu_count_m =psutil.cpu_count(logical=False) # cpu 물리 코어

                    os = platform.system() # os
                    os_v = platform.version() # os 버전

                    mem = psutil.virtual_memory() 
                    mem_t = round(mem.total/1024**3) # 메모리 총
                    mem_u = round(mem.used/1024**3) # 메모리 사용중
                    mem_p = mem.percent # 메모리 사용률

                    if mem_u == 0:
                        mem_u = 0.3
                    else:
                        pass

                    discord_v = discord.__version__

                    if sheet["Y" + str(i)].value == 0:
                        test_ping = "핑 테스트중..."

                        ping = round(client.latency*1000) # 핑
                        start = timen.getnow()  
                        msg = await message.channel.send(test_ping)
                        end = timen.getnow()  
                        await msg.delete()
                        api_ping = end - start
                        api_ping = round(api_ping * 1000) # api 핑

                        if ping < 200:
                            ping_s = "🟢"
                        elif ping >= 200 and ping < 500:
                            ping_s = "🟡"
                        else:
                            ping_s = "🔴"

                        if api_ping < 200:
                            a_ping_s = "🟢"
                        elif api_ping >= 200 and api_ping < 500:
                            a_ping_s = "🟡"
                        else:
                            a_ping_s = "🔴"

                        embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                        embed.add_field(name="• 이름", value="M BOT new!", inline=True)
                        embed.add_field(name="• 생성일", value="2021년 4월 24일", inline=True)
                        embed.add_field(name="• 사용중인 서버", value=f'{(s)}', inline=True)
                        embed.add_field(name="• 사용중인 서버의 유저", value=4344, inline=True)

                        embed.add_field(name="• 지원 언어", value="한국어, English, 日本, 中文", inline=True)
                        embed.add_field(name="• 개발자", value="마인잡지#0001", inline=True)

                        embed.add_field(name="• CPU", value="```{} {}```".format(cpu_b, cpu_c), inline=False)
                        embed.add_field(name="• CPU 사용률", value=f'`{(cpu_percent)}%`', inline=True)
                        embed.add_field(name="• CPU 코어", value="`{}/{}`".format(cpu_count_m,cpu_count), inline=True)
                        embed.add_field(name="• 시스템 OS", value="`{} | {}`".format(os,os_v), inline=True)

                        embed.add_field(name="• 메모리", value="`{}GB/{}GB`".format(mem_u, mem_t), inline=True)
                        embed.add_field(name="• 메모리 사용률", value="`{}%`".format(mem_p), inline=True)

                        embed.add_field(name="• 라이브러리", value="`Discord.py`", inline=True)
                        embed.add_field(name="• Discord.py 버전", value="`{}`".format(discord_v), inline=True)

                        embed.add_field(name="• Ping", value="`{}ms` {}".format(ping,ping_s), inline=True)
                        embed.add_field(name="• API Ping", value="`{}ms` {}".format(api_ping, a_ping_s), inline=True)

                        embed.add_field(name="• 개발자 이메일", value="Minemagazinebe@gmail.com", inline=False)
                        embed.set_author(name="M BOT 정보", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        embed.add_field(name="ㅤ",value="[서포터 서버](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                        embed.add_field(name="ㅤ",value="[초대하기!](<https://discord.com/oauth2/authorize?client_id=859608662213787679&permissions=0&scope=applications.commands%20bot>)",inline=True)
                        embed.add_field(name="ㅤ",value="[웹사이트](<https://discord-mbot.netlify.app>)",inline=True)

                    elif sheet["Y" + str(i)].value == 1:
                        test_ping = "Ping test..."

                        ping = round(client.latency*1000) # 핑
                        start = timen.getnow()  
                        msg = await message.channel.send(test_ping)
                        end = timen.getnow()  
                        await msg.delete()
                        api_ping = end - start
                        api_ping = round(api_ping * 1000) # api 핑

                        if ping < 200:
                            ping_s = "🟢"
                        elif ping >= 200 and ping < 500:
                            ping_s = "🟡"
                        else:
                            ping_s = "🔴"

                        if api_ping < 200:
                            a_ping_s = "🟢"
                        elif api_ping >= 200 and api_ping < 500:
                            a_ping_s = "🟡"
                        else:
                            a_ping_s = "🔴"

                        embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                        embed.add_field(name="• Name", value="M BOT new!", inline=True)
                        embed.add_field(name="• Creation date", value="4/24/2021", inline=True)
                        embed.add_field(name="• Server in use", value=f'{(s)}', inline=True)
                        embed.add_field(name="• User of the server in use", value=4344, inline=True)

                        embed.add_field(name="• Supported languages", value="한국어, English, 日本, 中文", inline=True)
                        embed.add_field(name="• Developer", value="마인잡지#0001", inline=True)

                        embed.add_field(name="• CPU", value="```{} {}```".format(cpu_b, cpu_c), inline=False)
                        embed.add_field(name="• CPU utilization", value=f'`{(cpu_percent)}%`', inline=True)
                        embed.add_field(name="• CPU core", value="`{}/{}`".format(cpu_count_m,cpu_count), inline=True)
                        embed.add_field(name="• System OS", value="`{} | {}`".format(os,os_v), inline=True)

                        embed.add_field(name="• Memory", value="`{}GB/{}GB`".format(mem_u, mem_t), inline=True)
                        embed.add_field(name="• Memory utilization", value="`{}%`".format(mem_p), inline=True)

                        embed.add_field(name="• Library", value="`Discord.py`", inline=True)
                        embed.add_field(name="• Discord.py version", value="`{}`".format(discord_v), inline=True)

                        embed.add_field(name="• Ping", value="`{}ms` {}".format(ping,ping_s), inline=True)
                        embed.add_field(name="• API Ping", value="`{}ms` {}".format(api_ping, a_ping_s), inline=True)

                        embed.add_field(name="• Developer Email", value="Minemagazinebe@gmail.com", inline=False)
                        embed.set_author(name="M BOT Information", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        embed.add_field(name="ㅤ",value="[support server](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                        embed.add_field(name="ㅤ",value="[Invite!](<https://discord.com/oauth2/authorize?client_id=859608662213787679&permissions=0&scope=applications.commands%20bot>)",inline=True)
                        embed.add_field(name="ㅤ",value="[Website](<https://discord-mbot.netlify.app>)",inline=True)

                    elif sheet["Y" + str(i)].value == 2:
                        test_ping = "Pingテスト中..."

                        ping = round(client.latency*1000) # 핑
                        start = timen.getnow()  
                        msg = await message.channel.send(test_ping)
                        end = timen.getnow()  
                        await msg.delete()
                        api_ping = end - start
                        api_ping = round(api_ping * 1000) # api 핑

                        if ping < 200:
                            ping_s = "🟢"
                        elif ping >= 200 and ping < 500:
                            ping_s = "🟡"
                        else:
                            ping_s = "🔴"

                        if api_ping < 200:
                            a_ping_s = "🟢"
                        elif api_ping >= 200 and api_ping < 500:
                            a_ping_s = "🟡"
                        else:
                            a_ping_s = "🔴"

                        embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                        embed.add_field(name="• 名前", value="M BOT new!", inline=True)
                        embed.add_field(name="• 生成日", value="2021年4月24日", inline=True)
                        embed.add_field(name="• 使用しているサーバー", value=f'{(s)}', inline=True)
                        embed.add_field(name="• 使用しているサーバーのユーザー", value=4344, inline=True)

                        embed.add_field(name="• サポートされる言語", value="한국어, English, 日本, 中文", inline=True)
                        embed.add_field(name="• 開発者", value="마인잡지#0001", inline=True)

                        embed.add_field(name="• CPU", value="```{} {}```".format(cpu_b, cpu_c), inline=False)
                        embed.add_field(name="• CPU 使用率", value=f'`{(cpu_percent)}%`', inline=True)
                        embed.add_field(name="• CPU コア", value="`{}/{}`".format(cpu_count_m,cpu_count), inline=True)
                        embed.add_field(name="• システム OS", value="`{} | {}`".format(os,os_v), inline=True)

                        embed.add_field(name="• メモリ", value="`{}GB/{}GB`".format(mem_u, mem_t), inline=True)
                        embed.add_field(name="• メモリの使用率", value="`{}%`".format(mem_p), inline=True)

                        embed.add_field(name="• ライブラリ", value="`Discord.py`", inline=True)
                        embed.add_field(name="• Discord.py バージョン", value="`{}`".format(discord_v), inline=True)

                        embed.add_field(name="• Ping", value="`{}ms` {}".format(ping,ping_s), inline=True)
                        embed.add_field(name="• API Ping", value="`{}ms` {}".format(api_ping, a_ping_s), inline=True)

                        embed.add_field(name="• 開発者の電子メール", value="Minemagazinebe@gmail.com", inline=False)
                        embed.set_author(name="M BOT 情報", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        embed.add_field(name="ㅤ",value="[サポーターサーバー](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                        embed.add_field(name="ㅤ",value="[招待!](<https://discord.com/oauth2/authorize?client_id=859608662213787679&permissions=0&scope=applications.commands%20bot>)",inline=True)
                        embed.add_field(name="ㅤ",value="[ウェブサイト](<https://discord-mbot.netlify.app>)",inline=True)

                    elif sheet["Y" + str(i)].value == 3:
                        test_ping = "ping测试..."

                        ping = round(client.latency*1000) # 핑
                        start = timen.getnow()  
                        msg = await message.channel.send(test_ping)
                        end = timen.getnow()  
                        await msg.delete()
                        api_ping = end - start
                        api_ping = round(api_ping * 1000) # api 핑

                        if ping < 200:
                            ping_s = "🟢"
                        elif ping >= 200 and ping < 500:
                            ping_s = "🟡"
                        else:
                            ping_s = "🔴"

                        if api_ping < 200:
                            a_ping_s = "🟢"
                        elif api_ping >= 200 and api_ping < 500:
                            a_ping_s = "🟡"
                        else:
                            a_ping_s = "🔴"

                        embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                        embed.add_field(name="• 名称", value="M BOT new!", inline=True)
                        embed.add_field(name="• 创立日期", value="2021 年 4 月 24 日", inline=True)
                        embed.add_field(name="• 使用中的服务器", value=f'{(s)}', inline=True)
                        embed.add_field(name="• 正在使用的服务器的用户", value=4344, inline=True)

                        embed.add_field(name="• 支持的语言", value="한국어, English, 日本, 中文", inline=True)
                        embed.add_field(name="• 开发商", value="마인잡지#0001", inline=True)

                        embed.add_field(name="• CPU", value="```{} {}```".format(cpu_b, cpu_c), inline=False)
                        embed.add_field(name="• CPU 利用率", value=f'`{(cpu_percent)}%`', inline=True)
                        embed.add_field(name="• CPU 核", value="`{}/{}`".format(cpu_count_m,cpu_count), inline=True)
                        embed.add_field(name="• 系统 OS", value="`{} | {}`".format(os,os_v), inline=True)

                        embed.add_field(name="• 记忆", value="`{}GB/{}GB`".format(mem_u, mem_t), inline=True)
                        embed.add_field(name="• 内存利用率", value="`{}%`".format(mem_p), inline=True)

                        embed.add_field(name="• 图书馆", value="`Discord.py`", inline=True)
                        embed.add_field(name="• Discord.py 版本", value="`{}`".format(discord_v), inline=True)

                        embed.add_field(name="• Ping", value="`{}ms` {}".format(ping,ping_s), inline=True)
                        embed.add_field(name="• API Ping", value="`{}ms` {}".format(api_ping, a_ping_s), inline=True)

                        embed.add_field(name="• 开发者邮箱", value="Minemagazinebe@gmail.com", inline=False)
                        embed.set_author(name="M BOT 정보", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        embed.add_field(name="ㅤ",value="[支持服务器](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                        embed.add_field(name="ㅤ",value="[邀请!](<https://discord.com/oauth2/authorize?client_id=859608662213787679&permissions=0&scope=applications.commands%20bot>)",inline=True)
                        embed.add_field(name="ㅤ",value="[网站](<https://discord-mbot.netlify.app>)",inline=True)

            try:
                await message.channel.send(embed=embed)
            except:
                file = openpyxl.load_workbook("user.xlsx")
                sheet = file.active
                un = message.author.name
                ui = message.author.id
                us = message.author.avatar_url
                now = datetime.datetime.now()
                ny = now.year
                nm = now.month
                nd = now.day
                nh = now.hour
                nmm = now.minute
                embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                msg = await message.channel.send(message.author.mention,embed=embed)
                await msg.add_reaction("<a:check:848042146044575767>")
                await msg.add_reaction("❌")
                def check(reaction, user):
                    if user.bot == 1:  # 봇이면 패스
                        return None
                    return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                try:
                    reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                except asyncio.TimeoutError:
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)
                else:
                    if str(reaction.emoji) == "<a:check:848042146044575767>":
                        for i in range(1, 1001):
                            if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                sheet["A" + str(i)].value = str(un)
                                sheet["B" + str(i)].value = str(ui)
                                sheet["C" + str(i)].value = 5000
                                sheet["D" + str(i)].value = "안녕하세요"
                                sheet["E" + str(i)].value = str(us)
                                sheet["F" + str(i)].value = ny
                                sheet["G" + str(i)].value = nm
                                sheet["H" + str(i)].value = nd
                                sheet["I" + str(i)].value = ny
                                sheet["J" + str(i)].value = nm
                                sheet["K" + str(i)].value = nd
                                sheet["L" + str(i)].value = nh
                                sheet["M" + str(i)].value = nmm
                                sheet["N" + str(i)].value = 0
                                sheet["O" + str(i)].value = 0
                                sheet["AA" + str(i)].value = 0
                                sheet["AB" + str(i)].value = 0
                                sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                await message.channel.send(message.author.mention, embed=embed)
                                sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                file.save("user.xlsx")
                                break
                    elif str(reaction.emoji) == "❌":
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)


        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


    if message.content.startswith("!내정보") or message.content.startswith("!내 정보") or msg == "!userinfo" or msg == "!user info" or msg == "!私の情報" or msg == "!我的信息":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        user = message.author
        date = datetime.datetime.utcfromtimestamp(((int(user.id) >> 22 ) + 1420070400000 ) / 1000)

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    user_name = message.author
                    user_id = user.id
                    user_server_name = user.display_name
                    user_p = user.avatar_url
                    user_year = date.year
                    user_month = date.month
                    user_day = date.day
                    user_highest_role = user.roles[-1].name
                    user_status = user.status

                    if message.author.bot:
                        user_bot = "True"
                    else:
                        user_bot = "False"

                    embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                    embed.add_field(name="• 이름", value=user_name, inline=True)
                    embed.add_field(name="• 닉네임", value=user_server_name, inline=True)
                    embed.add_field(name="• 아이디", value=user_id, inline=True)
                    embed.add_field(name="• 디스코드 가입일", value="{}년{}월{}일".format(user_year,user_month,user_day), inline=True)
                    embed.add_field(name="• 상태", value=user_status, inline=True)
                    embed.add_field(name="• 최고역할", value=user_highest_role, inline=True)
                    embed.add_field(name="• 봇 여부", value=user_bot, inline=True)

                    embed.set_author(name="{}님의 정보".format(user_name), icon_url=user_p) # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embed.set_thumbnail(url=user_p)

                elif sheet["Y" + str(i)].value == 1:
                    user_name = message.author
                    user_id = user.id
                    user_server_name = user.display_name
                    user_p = user.avatar_url
                    user_year = date.year
                    user_month = date.month
                    user_day = date.day
                    user_highest_role = user.roles[-1].name
                    user_status = user.status

                    if message.author.bot:
                        user_bot = "True"
                    else:
                        user_bot = "False"

                    embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                    embed.add_field(name="• Name", value=user_name, inline=True)
                    embed.add_field(name="• Nickname", value=user_server_name, inline=True)
                    embed.add_field(name="• ID", value=user_id, inline=True)
                    embed.add_field(name="• Discord sign up", value="{}/{}/{}/".format(user_month,user_day,user_year), inline=True)
                    embed.add_field(name="• State", value=user_status, inline=True)
                    embed.add_field(name="• Top role", value=user_highest_role, inline=True)
                    embed.add_field(name="• Bot or not", value=user_bot, inline=True)

                    embed.set_author(name="{}'s information".format(user_name), icon_url=user_p) # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embed.set_thumbnail(url=user_p)


                elif sheet["Y" + str(i)].value == 2:
                    user_name = message.author
                    user_id = user.id
                    user_server_name = user.display_name
                    user_p = user.avatar_url
                    user_year = date.year
                    user_month = date.month
                    user_day = date.day
                    user_highest_role = user.roles[-1].name
                    user_status = user.status

                    if message.author.bot:
                        user_bot = "True"
                    else:
                        user_bot = "False"

                    embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                    embed.add_field(name="• 名前", value=user_name, inline=True)
                    embed.add_field(name="• ニックネーム", value=user_server_name, inline=True)
                    embed.add_field(name="• ユーザ名", value=user_id, inline=True)
                    embed.add_field(name="• ディスコード日", value="{}年{}月{}日".format(user_year,user_month,user_day), inline=True)
                    embed.add_field(name="• 状態", value=user_status, inline=True)
                    embed.add_field(name="• 最高の役割", value=user_highest_role, inline=True)
                    embed.add_field(name="• ボットかどうか", value=user_bot, inline=True)

                    embed.set_author(name="{}さんの情報".format(user_name), icon_url=user_p) # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embed.set_thumbnail(url=user_p)


                elif sheet["Y" + str(i)].value == 3:
                    user_name = message.author
                    user_id = user.id
                    user_server_name = user.display_name
                    user_p = user.avatar_url
                    user_year = date.year
                    user_month = date.month
                    user_day = date.day
                    user_highest_role = user.roles[-1].name
                    user_status = user.status

                    if message.author.bot:
                        user_bot = "True"
                    else:
                        user_bot = "False"

                    embed = discord.Embed(title=" ", description=" ", timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x90c2ff)
                    embed.add_field(name="• 名称", value=user_name, inline=True)
                    embed.add_field(name="• 昵称", value=user_server_name, inline=True)
                    embed.add_field(name="• ID", value=user_id, inline=True)
                    embed.add_field(name="• DISCORD 会员自", value="{}년{}월{}일".format(user_year,user_month,user_day), inline=True)
                    embed.add_field(name="• 状态", value=user_status, inline=True)
                    embed.add_field(name="• 顶级角色", value=user_highest_role, inline=True)
                    embed.add_field(name="• 机器人与否", value=user_bot, inline=True)

                    embed.set_author(name="{}的资料".format(user_name), icon_url=user_p) # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    embed.set_thumbnail(url=user_p)

        try:
            await message.channel.send(embed=embed)
        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)


# 채광

    if message.content.startswith("!채광") or message.content.startswith("!광질") or message.content.startswith("!mining") or message.content.startswith("!採鉱") or message.content.startswith("!矿业"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        un = message.author.name
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    mining = "채광"
                    mineral = "채광 결과"
                    sell = "판매가격"

                    stone = "돌"
                    coal = "석탄"
                    copper = "구리"
                    steel = "철"
                    silver = "은"
                    gold = "금"
                    platinum = "백금"
                    iridium = "이리듐"

                    emerald = "에메랄드"
                    diamond = "다이아몬드"
                    ruby = "루비"
                    sapphire = "사파이어"
                    m_discord  = "DISCORD"

                elif sheet["Y" + str(i)].value == 1:
                    mining = "Mining"
                    mineral = "Mining results"
                    sell = "Sale Price"

                    stone = "stone"
                    coal = "coal"
                    copper = "copper"
                    steel = "steel"
                    silver = "silver"
                    gold = "gold"
                    platinum = "platinum"
                    iridium = "iridium"

                    emerald = "emerald"
                    diamond = "diamond"
                    ruby = "ruby"
                    sapphire = "sapphire"
                    m_discord  = "DISCORD"

                elif sheet["Y" + str(i)].value == 2:
                    mining = "採鉱"
                    mineral = "採光結果"
                    sell = "販売価格"

                    stone = "石"
                    coal = "石炭"
                    copper = "銅"
                    steel = "鉄"
                    silver = "は"
                    gold = "金の"
                    platinum = "白金"
                    iridium = "イリジウム"

                    emerald = "エメラルド"
                    diamond = "ダイヤモンド"
                    ruby = "ルビー"
                    sapphire = "サファイア"
                    m_discord  = "DISCORD"

                elif sheet["Y" + str(i)].value == 3:
                    mining = "矿业"
                    mineral = "挖掘结果"
                    sell = "销售价格"

                    stone = "岩石"
                    coal = "煤炭"
                    copper = "铜"
                    steel = "钢"
                    silver = "银"
                    gold = "金子"
                    platinum = "铂"
                    iridium = "铱"

                    emerald = "翠"
                    diamond = "钻石"
                    ruby = "红宝石"
                    sapphire = "蓝宝石"
                    m_discord  = "DISCORD"

                sum = sheet["C" + str(i)].value
                sheet["X" + str(i)].value = sheet["X" + str(i)].value + 1
                r = (1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,4,4,4,4,4,5,5,5,5,5,5,6,6,6,6,6,7,7,7,8,8,9,9,9,10,10,10,11,11,11,13)
                rr = str(random.sample(r, 1))
                rr = rr[1:-1]
                rr = int(rr)
                u = sheet["W" + str(i)].value
                if "『광부』" in u:
                    pass
                else:
                    embed = discord.Embed(title="**《업적달성》**",description="『광부』 - 채광 하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    await message.channel.send(message.author.mention,embed=embed)
                    sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『광부』 "

                if sheet["X" + str(i)].value == 100:
                    embed = discord.Embed(title="**《업적달성》**",description="『프로광부』 - 채광 100번 하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    await message.channel.send(message.author.mention,embed=embed)
                    sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『프로광부』 "

                if rr == 1:
                    a = random.randrange(900, 1101)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, stone, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x5a5a5a)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 2:
                    a = random.randrange(1300, 1701)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, coal, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x262626)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 3:
                    a = random.randrange(1500, 2501)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, copper, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff9920)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 4:
                    a = random.randrange(3000, 4001)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, steel, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xededed)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 5:
                    a = random.randrange(3500, 4501)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, silver, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xd7d7d7)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 6:
                    a = random.randrange(4500, 5501)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, gold, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff720)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 7:
                    a = random.randrange(40000, 60000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, platinum, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xe2e2e2)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 8:
                    a = random.randrange(80000, 120000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, iridium, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xbfbfbf)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 9:
                    a = random.randrange(35000, 55000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, emerald, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x4aff20)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 10:
                    a = random.randrange(60000, 80000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, diamond, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x20f2ff)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 11:
                    a = random.randrange(40000, 60000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, ruby, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff2020)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 12:
                    a = random.randrange(40000, 60000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, sapphire, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x2020ff)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    break

                if rr == 13:
                    a = random.randrange(500000, 750000)
                    embed = discord.Embed(title="**{}**".format(mining), description="{} : {}\n{} : {}".format(mineral, m_discord, sell, a),timesamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    await message.channel.send(message.author.mention, embed=embed)
                    summ = sheet["C" + str(i)].value = sheet["C" + str(i)].value + a
                    if "『M-DISCORD』" in u:
                        pass
                    else:
                        embed = discord.Embed(title="**《업적달성》**",description="『M-DISCORD』 - `!채광`에서 DISCORD 채광하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        await message.channel.send(message.author.mention,embed=embed)
                        sheet["W" + str(i)].value = sheet["W" + str(i)].value + "『M-DISCORD』 "
                    break

                
                

        file.save("user.xlsx")


# 한강물

    if msg == "!한강물" or msg == "!한강" or msg == "!한강 기온" or msg == "!한강온도" or msg == "!한강기온" or msg == "!서울 한강" or msg == "!HanRiver" or msg == "!Han River" or msg == "!han river" or msg == "!hanriver" or msg == "!漢江" or msg == "!汉江":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    w = "기준 서울 한강 온도"


                elif sheet["Y" + str(i)].value == 1:
                    w = "Standard Seoul Han River temperature"


                elif sheet["Y" + str(i)].value == 2:
                    w = "基準ソウル漢江温度"


                elif sheet["Y" + str(i)].value == 3:
                    w = "汉江标准汉江温度"

        try:
            t = requests.get("https://api.hangang.msub.kr/").json()['temp']
            f = float(t) * 9 / 5 +32
            time = requests.get("https://api.hangang.msub.kr/").json()['time']
            embed = discord.Embed(title = "{}{}".format(time, w), description="{}℃/{}℉".format(t,f), timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x579aff)
            await message.channel.send(embed = embed)

        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)


# 계산

    if message.content.startswith("!계산") or message.content.startswith("!計算") or message.content.startswith("!计算"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            a = message.content[4:]
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    if sheet["Y" + str(i)].value == 0:
                        w = "계산기"
                        input_value = "입력 값"
                        output_value = "출력 값"
                        error_value = "숫자 (+ - * / ^) 숫자"


                    elif sheet["Y" + str(i)].value == 1:
                        w = "Calculator"
                        input_value = "Input value"
                        output_value = "Output value"
                        error_value = "number (+ - * / ^) number"


                    elif sheet["Y" + str(i)].value == 2:
                        w = "計算機"
                        input_value = "入力値"
                        output_value = "出力値"
                        error_value = "数字 (+ - * / ^) 数字"


                    elif sheet["Y" + str(i)].value == 3:
                        w = "计算器"
                        input_value = "输入值"
                        output_value = "产值"
                        error_value = "数字 (+ - * / ^) 数字"

            if "+" in a:
                as1 = a.split("+")
                o = int(as1[0]) + int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "-" in a:
                as1 = a.split("-")
                o = int(as1[0]) - int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "*" in a:
                as1 = a.split("*")
                o = int(as1[0]) * int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "/" in a:
                as1 = a.split("/")
                o = int(as1[0]) / int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "^" in a:
                as1 = a.split("^")
                if int(as1[0]) and int(as1[1]) < 10000:
                    o = int(as1[0]) ** int(as1[1])
                    try:
                        embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        await message.channel.send(embed=embed)
                    except:
                        file = openpyxl.load_workbook("user.xlsx")
                        sheet = file.active
                        un = message.author.name
                        ui = message.author.id
                        us = message.author.avatar_url
                        now = datetime.datetime.now()
                        ny = now.year
                        nm = now.month
                        nd = now.day
                        nh = now.hour
                        nmm = now.minute
                        embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        msg = await message.channel.send(message.author.mention,embed=embed)
                        await msg.add_reaction("<a:check:848042146044575767>")
                        await msg.add_reaction("❌")
                        def check(reaction, user):
                            if user.bot == 1:  # 봇이면 패스
                                return None
                            return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                        except asyncio.TimeoutError:
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)
                        else:
                            if str(reaction.emoji) == "<a:check:848042146044575767>":
                                for i in range(1, 1001):
                                    if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                        sheet["A" + str(i)].value = str(un)
                                        sheet["B" + str(i)].value = str(ui)
                                        sheet["C" + str(i)].value = 5000
                                        sheet["D" + str(i)].value = "안녕하세요"
                                        sheet["E" + str(i)].value = str(us)
                                        sheet["F" + str(i)].value = ny
                                        sheet["G" + str(i)].value = nm
                                        sheet["H" + str(i)].value = nd
                                        sheet["I" + str(i)].value = ny
                                        sheet["J" + str(i)].value = nm
                                        sheet["K" + str(i)].value = nd
                                        sheet["L" + str(i)].value = nh
                                        sheet["M" + str(i)].value = nmm
                                        sheet["N" + str(i)].value = 0
                                        sheet["O" + str(i)].value = 0
                                        sheet["AA" + str(i)].value = 0
                                        sheet["AB" + str(i)].value = 0
                                        sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                        embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                        file.save("user.xlsx")
                                        break
                            elif str(reaction.emoji) == "❌":
                                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                                await msg.delete()
                                await message.channel.send(embed=embed)

                else:
                    await message.channel.send(embed=discord.Embed(title="Error", description = "overload", color = 0xff0000))
            
            else:
                await message.channel.send(message.author.mention + error_value)
        
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return

    if message.content.startswith("!calculate"):
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        try:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            a = message.content[11:]
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    if sheet["Y" + str(i)].value == 0:
                        w = "계산기"
                        input_value = "입력 값"
                        output_value = "출력 값"
                        error_value = "숫자 (+ - * / ^) 숫자"


                    elif sheet["Y" + str(i)].value == 1:
                        w = "Calculator"
                        input_value = "Input value"
                        output_value = "Output value"
                        error_value = "number (+ - * / ^) number"


                    elif sheet["Y" + str(i)].value == 2:
                        w = "計算機"
                        input_value = "入力値"
                        output_value = "出力値"
                        error_value = "数字 (+ - * / ^) 数字"


                    elif sheet["Y" + str(i)].value == 3:
                        w = "计算器"
                        input_value = "输入值"
                        output_value = "产值"
                        error_value = "数字 (+ - * / ^) 数字"

            if "+" in a:
                as1 = a.split("+")
                o = int(as1[0]) + int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "-" in a:
                as1 = a.split("-")
                o = int(as1[0]) - int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "*" in a:
                as1 = a.split("*")
                o = int(as1[0]) * int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "/" in a:
                as1 = a.split("/")
                o = int(as1[0]) / int(as1[1])
                try:
                    embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)
                except:
                    file = openpyxl.load_workbook("user.xlsx")
                    sheet = file.active
                    un = message.author.name
                    ui = message.author.id
                    us = message.author.avatar_url
                    now = datetime.datetime.now()
                    ny = now.year
                    nm = now.month
                    nd = now.day
                    nh = now.hour
                    nmm = now.minute
                    embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                    msg = await message.channel.send(message.author.mention,embed=embed)
                    await msg.add_reaction("<a:check:848042146044575767>")
                    await msg.add_reaction("❌")
                    def check(reaction, user):
                        if user.bot == 1:  # 봇이면 패스
                            return None
                        return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                    try:
                        reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                    except asyncio.TimeoutError:
                        embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                        await msg.delete()
                        await message.channel.send(embed=embed)
                    else:
                        if str(reaction.emoji) == "<a:check:848042146044575767>":
                            for i in range(1, 1001):
                                if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                    sheet["A" + str(i)].value = str(un)
                                    sheet["B" + str(i)].value = str(ui)
                                    sheet["C" + str(i)].value = 5000
                                    sheet["D" + str(i)].value = "안녕하세요"
                                    sheet["E" + str(i)].value = str(us)
                                    sheet["F" + str(i)].value = ny
                                    sheet["G" + str(i)].value = nm
                                    sheet["H" + str(i)].value = nd
                                    sheet["I" + str(i)].value = ny
                                    sheet["J" + str(i)].value = nm
                                    sheet["K" + str(i)].value = nd
                                    sheet["L" + str(i)].value = nh
                                    sheet["M" + str(i)].value = nmm
                                    sheet["N" + str(i)].value = 0
                                    sheet["O" + str(i)].value = 0
                                    sheet["AA" + str(i)].value = 0
                                    sheet["AB" + str(i)].value = 0
                                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                    embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                    await message.channel.send(message.author.mention, embed=embed)
                                    sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                    file.save("user.xlsx")
                                    break
                        elif str(reaction.emoji) == "❌":
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)

            elif "^" in a:
                as1 = a.split("^")
                if int(as1[0]) and int(as1[1]) < 10000:
                    o = int(as1[0]) ** int(as1[1])
                    try:
                        embed = discord.Embed(title="{}".format(w),description="{}\n```{}```\n{}\n```{}```".format(input_value, a, output_value, o),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
                        embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                        await message.channel.send(embed=embed)
                    except:
                        file = openpyxl.load_workbook("user.xlsx")
                        sheet = file.active
                        un = message.author.name
                        ui = message.author.id
                        us = message.author.avatar_url
                        now = datetime.datetime.now()
                        ny = now.year
                        nm = now.month
                        nd = now.day
                        nh = now.hour
                        nmm = now.minute
                        embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                        msg = await message.channel.send(message.author.mention,embed=embed)
                        await msg.add_reaction("<a:check:848042146044575767>")
                        await msg.add_reaction("❌")
                        def check(reaction, user):
                            if user.bot == 1:  # 봇이면 패스
                                return None
                            return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
                        try:
                            reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
                        except asyncio.TimeoutError:
                            embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                            await msg.delete()
                            await message.channel.send(embed=embed)
                        else:
                            if str(reaction.emoji) == "<a:check:848042146044575767>":
                                for i in range(1, 1001):
                                    if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                                        sheet["A" + str(i)].value = str(un)
                                        sheet["B" + str(i)].value = str(ui)
                                        sheet["C" + str(i)].value = 5000
                                        sheet["D" + str(i)].value = "안녕하세요"
                                        sheet["E" + str(i)].value = str(us)
                                        sheet["F" + str(i)].value = ny
                                        sheet["G" + str(i)].value = nm
                                        sheet["H" + str(i)].value = nd
                                        sheet["I" + str(i)].value = ny
                                        sheet["J" + str(i)].value = nm
                                        sheet["K" + str(i)].value = nd
                                        sheet["L" + str(i)].value = nh
                                        sheet["M" + str(i)].value = nmm
                                        sheet["N" + str(i)].value = 0
                                        sheet["O" + str(i)].value = 0
                                        sheet["AA" + str(i)].value = 0
                                        sheet["AB" + str(i)].value = 0
                                        sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                                        embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                                        await message.channel.send(message.author.mention, embed=embed)
                                        sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                                        file.save("user.xlsx")
                                        break
                            elif str(reaction.emoji) == "❌":
                                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                                await msg.delete()
                                await message.channel.send(embed=embed)

                else:
                    await message.channel.send(embed=discord.Embed(title="Error", description = "overload", color = 0xff0000))
            
            else:
                await message.channel.send(message.author.mention + error_value)
        
        except Exception as e:
            await message.channel.send(embed=discord.Embed(title="Error", description = str(e), color = 0xff0000))
            return


    if message.content.startswith("!원주율") or msg == "!Pi" or msg == "!perimeter" or msg == "!Perimeter" or msg == "!円周率" or msg == "!周长" or msg == "!pi" or msg == "!PI":

        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        un = message.author.name
        ui = message.author.id
        a = message.content[11:]
        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    w = "원주율"
                    ii = "소수점 아래 1000자리"


                elif sheet["Y" + str(i)].value == 1:
                    w = "Perimeter"
                    ii = "1000 decimal places"


                elif sheet["Y" + str(i)].value == 2:
                    w = "円周率"
                    ii = "小数点以下1000桁"


                elif sheet["Y" + str(i)].value == 3:
                    w = "周长"
                    ii = "1000 位小数"

        r = 3.141592653589793238462643383279502884197169399375105820974944592307816406286208998628034825342117067
        try:

            embed = discord.Embed(title = w, description="{}\n```3.141592653589793238462643383279502884197169399375105820974944592307816406286208998628034825342117067982148086513282306670938446095505822317253594081284811174502841027019385211055596446229489549303819644288109756659334461284756482337867831652712019091456485669234603486104543266482133936072602491412737245870066063155881748815209209628292540917153643678925903600113305305488204665213841469519415116094330572703657595919530921861173819326117931051185480744623799627495673518857527248912279381830119491298336733624406566430860213949463952247371907021798609437027705392171762931767523846748184676694051320005681271452635608277857713427577896091736371787214684409012249534301465495853710507922796892589235420199561121290219608640344181598136297747713099605187072113499999983729780499510597317328160963185950244594553469083026425223082533446850352619311881710100031378387528865875332083814206171776691473035982534904287554687311595628638823537875937519577818577805321712268066130019278766111959092164201989```".format(ii),timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x545dcf)
            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
            await message.channel.send(embed=embed)

        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)


# 언어설정

    if msg == "!언어" or msg == "!language" or msg == "!言語" or msg == "!语" or msg == "!语言":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id

        for i in range(1, 1001):
            if sheet["B" + str(i)].value == str(ui):
                if sheet["Y" + str(i)].value == 0:
                    w = "현재 사용중인 언어"
                    input_value = "한국어"


                elif sheet["Y" + str(i)].value == 1:
                    w = "language currently in use"
                    input_value = "English"


                elif sheet["Y" + str(i)].value == 2:
                    w = "現在使用中の言語"
                    input_value = "日本"


                elif sheet["Y" + str(i)].value == 3:
                    w = "当前使用的语言"
                    input_value = "中文"

        try:
            embed = discord.Embed(title="Set Language", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x000746)
            embed.add_field(name="**한국어 🇰🇷**", value="M BOT의 기본 언어를 한국어로 설정합니다", inline=False)
            embed.add_field(name="**English 🇺🇸**", value="Set the default language of M BOT to English", inline=False)
            embed.add_field(name="**日本 🇯🇵**", value="M BOTのデフォルト言語を日本語に設定します", inline=False)
            embed.add_field(name="**中文 🇨🇳**", value="将M BOT的默认语言设置为中文", inline=False)
            embed.add_field(name="**ㅤ**", value="ㅤ", inline=False)
            embed.add_field(name="**{}**".format(w), value=input_value, inline=False)
            embed.add_field(name="**ㅤ**", value="ㅤ", inline=False)
            embed.add_field(name="**⚠️Warning!⚠️**", value="`!가입` 하셔야 언어설정을 사용하실수 있습니다\nYou must sign up for `!join` to use the language settings.\n `!登録` が必要言語の設定を使用し出来ます\n您必须注册 `!加入` 才能使用语言设置。", inline=False)
            embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
            embed.add_field(name="ㅤ",value="[homepage](<https://discord-mbot.netlify.app/>)",inline=True)
            embed.set_author(name="M BOT Language", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
            msg = await message.channel.send(embed=embed)

            await msg.add_reaction("🇰🇷")
            await msg.add_reaction("🇺🇸")
            await msg.add_reaction("🇯🇵")
            await msg.add_reaction("🇨🇳")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '🇰🇷' or user == message.author and str(reaction.emoji) == '🇺🇸' or user == message.author and str(reaction.emoji) == '🇯🇵' or user == message.author and str(reaction.emoji) == '🇨🇳' or user == message.author and str(reaction.emoji) == '❌'


            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="Set Language", description="Cancel - 취소 - キャンセル - 取消", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)

            else:
                if str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="Set Language", description="Cancel - 취소 - キャンセル - 取消", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)


                elif str(reaction.emoji) == "🇰🇷":
                    await msg.delete()

                    for i in range(1, 1001):
                        if sheet["B" + str(i)].value == str(ui):
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                            embed = discord.Embed(title="언어설정", description="M BOT의 기본언어를 한국어로 설정하였습니다.",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                            embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                            embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
                            embed.set_author(name="M BOT 언어", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                            await message.channel.send(embed=embed)

                            file.save("user.xlsx")
                            break

                elif str(reaction.emoji) == "🇺🇸":
                    await msg.delete()

                    for i in range(1, 1001):
                        if sheet["B" + str(i)].value == str(ui):
                            sheet["Y" + str(i)].value = 1 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                            embed = discord.Embed(title="Set Language", description="The default language of M BOT is set to English.",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                            embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                            embed.add_field(name="ㅤ",value="[Homepage](<https://discord-mbot.netlify.app/>)",inline=True)
                            embed.set_author(name="M BOT Language", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                            await message.channel.send(embed=embed)

                            file.save("user.xlsx")
                            break

                elif str(reaction.emoji) == "🇯🇵":
                    await msg.delete()

                    for i in range(1, 1001):
                        if sheet["B" + str(i)].value == str(ui):
                            sheet["Y" + str(i)].value = 2 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                            embed = discord.Embed(title="言語設定", description="M BOTのデフォルト言語を日本語に設定しました。",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                            embed.add_field(name="ㅤ",value="[ディスコード](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                            embed.add_field(name="ㅤ",value="[ホームページ](<https://discord-mbot.netlify.app/>)",inline=True)
                            embed.set_author(name="M BOT 言語", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                            await message.channel.send(embed=embed)

                            file.save("user.xlsx")
                            break

                elif str(reaction.emoji) == "🇨🇳":
                    await msg.delete()

                    for i in range(1, 1001):
                        if sheet["B" + str(i)].value == str(ui):
                            sheet["Y" + str(i)].value = 3 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                            embed = discord.Embed(title="语言设置", description="M BOT 的默认语言设置为中文。",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                            embed.add_field(name="ㅤ",value="[不和谐(discord)](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                            embed.add_field(name="ㅤ",value="[主页](<https://discord-mbot.netlify.app/>)",inline=True)
                            embed.set_author(name="M BOT 语言", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                            await message.channel.send(embed=embed)

                            file.save("user.xlsx")
                            break


        except:
            file = openpyxl.load_workbook("user.xlsx")
            sheet = file.active
            un = message.author.name
            ui = message.author.id
            us = message.author.avatar_url
            now = datetime.datetime.now()
            ny = now.year
            nm = now.month
            nd = now.day
            nh = now.hour
            nmm = now.minute
            embed = discord.Embed(title="**M BOT 가입**",description="M BOT의 [이용약관](<https://discord-mbot.netlify.app/use.html>)을 읽으신후\n60초 내로 <a:check:848042146044575767>를 눌러주시면 가입이 완료됩니다!", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
            msg = await message.channel.send(message.author.mention,embed=embed)
            await msg.add_reaction("<a:check:848042146044575767>")
            await msg.add_reaction("❌")
            def check(reaction, user):
                if user.bot == 1:  # 봇이면 패스
                    return None
                return user == message.author and str(reaction.emoji) == '<a:check:848042146044575767>' or user == message.author and str(reaction.emoji) == '❌'
            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                await msg.delete()
                await message.channel.send(embed=embed)
            else:
                if str(reaction.emoji) == "<a:check:848042146044575767>":
                    for i in range(1, 1001):
                        if sheet["A" + str(i)].value == "-" and sheet["B" + str(i)].value == "-":
                            sheet["A" + str(i)].value = str(un)
                            sheet["B" + str(i)].value = str(ui)
                            sheet["C" + str(i)].value = 5000
                            sheet["D" + str(i)].value = "안녕하세요"
                            sheet["E" + str(i)].value = str(us)
                            sheet["F" + str(i)].value = ny
                            sheet["G" + str(i)].value = nm
                            sheet["H" + str(i)].value = nd
                            sheet["I" + str(i)].value = ny
                            sheet["J" + str(i)].value = nm
                            sheet["K" + str(i)].value = nd
                            sheet["L" + str(i)].value = nh
                            sheet["M" + str(i)].value = nmm
                            sheet["N" + str(i)].value = 0
                            sheet["O" + str(i)].value = 0
                            sheet["AA" + str(i)].value = 0
                            sheet["AB" + str(i)].value = 0
                            sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3
                            embed = discord.Embed(title="Complete",description="M BOT 서비스에 가입되셨습니다.\nYou have signed up for M BOT service.\nM BOTサービスに登録完了。\n您已注册 M BOT 服务。", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            embed = discord.Embed(title="**《업적달성》**",description="『유저』 - 가입하기! `!업적`", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xfff957)
                            await message.channel.send(message.author.mention, embed=embed)
                            sheet["W" + str(i)].value = sheet["W" + str(i)].value = "『유저』 "
                            file.save("user.xlsx")
                            break
                elif str(reaction.emoji) == "❌":
                    embed = discord.Embed(title="M BOT 가입", description="M BOT 가입에 취소하셨습니다.\nYou have canceled your M BOT service subscription.", timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0xff0000)
                    await msg.delete()
                    await message.channel.send(embed=embed)

    
    if message.content.startswith("!언어") and not msg == "!언어" and not msg == "!language" and not msg == "!言語" and not msg == "!语" and not msg == "!语言":
        print("\n이름 : {} | 아이디 : {} | 메세지 : {}\n채널 : {} | 채널아이디 : {}\n서버 : {} | 서버아이디 : {}\n사용시간 : {}".format(message.author.name, message.author.id, message.content, message.channel, message.channel.id,message.guild.name, message.guild.id, datetime.datetime.today()))
        file = openpyxl.load_workbook("user.xlsx")
        sheet = file.active
        ui = message.author.id
        l = message.content.split(" ")
        l = l[1]
        if l == "korean" or l == "Korean" or l == "KOREAN" or l == "한국어" or l == "한국" or l == "韓国語" or l =="韓国" or l == "韩国" or l == "KR" or l =="KOR" or l == "kor" or l == "kr":
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sheet["Y" + str(i)].value = 0 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                    embed = discord.Embed(title="언어설정", description="M BOT의 기본언어를 한국어로 설정하였습니다.",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                    embed.add_field(name="ㅤ",value="[디스코드](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[홈페이지](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 언어", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)

                    file.save("user.xlsx")
                    break

        elif l == "日本" or l == "日本の" or l == "の日本" or l == "일본어" or l == "일본" or l == "japan" or l == "Japan" or l == "Japanese" or l == "JP" or l == "jp" or l == "Jp" or l == "日语":
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sheet["Y" + str(i)].value = 2 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                    embed = discord.Embed(title="言語設定", description="M BOTのデフォルト言語を日本語に設定しました。",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                    embed.add_field(name="ㅤ",value="[ディスコード](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[ホームページ](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 言語", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)

                    file.save("user.xlsx")
                    break

        elif l == "中文" or l == "中国人" or l == "중국" or l == "중국어" or l == "China" or l == "CHINA" or l == "china" or l == "CH" or l == "Ch" or l == "ch" or l == "中国の" or l == "中国" or l == "Chinese" or l == "Chinese":
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sheet["Y" + str(i)].value = 3 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                    embed = discord.Embed(title="语言设置", description="M BOT 的默认语言设置为中文。",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                    embed.add_field(name="ㅤ",value="[不和谐(discord)](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[主页](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT 语言", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)

                    file.save("user.xlsx")
                    break

        elif l == "US" or l == "us" or l == "English" or l == "english" or l == "영어" or l == "英語" or l == "英语" or l == "USA" or l =="usa":
            for i in range(1, 1001):
                if sheet["B" + str(i)].value == str(ui):
                    sheet["Y" + str(i)].value = 1 # 언어 - 한국 : 0, 영어 : 1, 일본어 : 2, 중국어 : 3

                    embed = discord.Embed(title="Set Language", description="The default language of M BOT is set to English.",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x6a0000)
                    embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
                    embed.add_field(name="ㅤ",value="[Homepage](<https://discord-mbot.netlify.app/>)",inline=True)
                    embed.set_author(name="M BOT Language", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
                    embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
                    await message.channel.send(embed=embed)

                    file.save("user.xlsx")
                    break

        else:
            embed = discord.Embed(title="Language setting help", description="",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x000746)
            embed.add_field(name="**How to use**", value="`!언어 <언어이름>`\n`!language <language name>`\n`!言語 <言語名>\n`!语 语言名称`", inline=False)
            embed.add_field(name="**ㅤ**", value="ㅤ", inline=False)
            embed.add_field(name="**⚠️Warning!⚠️**", value="`!가입` 하셔야 언어설정을 사용하실수 있습니다\nYou must sign up for `!join` to use the language settings.\n `!登録` が必要言語の設定を使用し出来ます\n您必须注册 `!加入` 才能使用语言设置。", inline=False)
            embed.add_field(name="ㅤ",value="[Discord](<https://discord.gg/fEEHHM4eHq>)",inline=True)
            embed.add_field(name="ㅤ",value="[homepage](<https://discord-mbot.netlify.app/>)",inline=True)
            embed.set_author(name="M BOT Language", icon_url="https://cdn.discordapp.com/attachments/832849028873191516/858274320463691796/-_5.png") # 맨위
            embed.set_footer(text=message.author, icon_url=message.author.avatar_url) # 맨 아래
            msg = await message.channel.send(embed=embed)



        








token = "user token"
client.run(token)
